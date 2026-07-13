"""
FTTH Auto BOM - Excel Export
Author: Mustafa M M Elaham
"""

def export_bom_to_excel(bom_items, params, output_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    except ImportError:
        return False, "openpyxl not installed"

    wb = Workbook()
    hdr_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    sec_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    sec_font = Font(color="FFFFFF", bold=True, size=12, name="Calibri")
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin = Border(left=Side(style='thin', color="B4C6E7"), right=Side(style='thin', color="B4C6E7"),
        top=Side(style='thin', color="B4C6E7"), bottom=Side(style='thin', color="B4C6E7"))
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # SHEET 1: Project Info
    ws = wb.active; ws.title = "Project Info"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 35; ws.column_dimensions["D"].width = 40

    ws.merge_cells("B2:D2")
    ws["B2"] = "FTTH Auto BOM - Generated from QGIS Layers"
    ws["B2"].font = Font(color="1F4E78", bold=True, size=18, name="Calibri")
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 35

    ws.merge_cells("B3:D3")
    ws["B3"] = f"Project: {params.get('project_name', '')}  |  Zone: {params.get('zone_code', '')}"
    ws["B3"].font = Font(color="666666", size=11, name="Calibri")
    ws["B3"].alignment = Alignment(horizontal="center", vertical="center")

    info = [("Number of Blocks", params.get("num_blocks", "Auto")),
        ("Number of MDUs", params.get("num_mdus", "Auto")),
        ("Number of Hubs", "Auto from layer"),
        ("Number of Poles", "Auto from layer"),
        ("Aerial Deployment", "Yes"),
        ("OLT Termination", "Yes" if params.get("is_olt_termination") else "No"),
        ("Core Manhole", "Yes" if params.get("is_core_mh") else "No"),
        ("LMJ Status", "Existing" if params.get("is_lmj_existing") else "New"),
        ("Generation Method", "AUTO - From QGIS Layer Data"),
        ("Direction Changes", "Auto-calculated from cable geometry"),]

    r = 5
    ws.merge_cells(f"B{r}:D{r}")
    ws.cell(row=r, column=2, value="PROJECT SUMMARY").font = sec_font
    ws.cell(row=r, column=2).fill = sec_fill
    for c in range(2, 5): ws.cell(row=r, column=c).fill = sec_fill; ws.cell(row=r, column=c).border = thin

    for i, (label, value) in enumerate(info):
        rr = r + 1 + i
        ws.cell(row=rr, column=2, value=label).font = Font(bold=True, size=10, name="Calibri")
        ws.cell(row=rr, column=2).alignment = left; ws.cell(row=rr, column=2).border = thin
        ws.cell(row=rr, column=3, value=value).alignment = center; ws.cell(row=rr, column=3).border = thin
        ws.cell(row=rr, column=4).border = thin
        if i % 2 == 1:
            for c in range(2, 5): ws.cell(row=rr, column=c).fill = alt_fill

    # Key Relations
    r2 = r + len(info) + 3
    ws.merge_cells(f"B{r2}:D{r2}")
    ws.cell(row=r2, column=2, value="AUTO-CALCULATION RELATIONS").font = sec_font
    ws.cell(row=r2, column=2).fill = sec_fill
    for c in range(2, 5): ws.cell(row=r2, column=c).fill = sec_fill; ws.cell(row=r2, column=c).border = thin

    relations = [
        "DE/Tangent: counted from poles each cable passes through",
        "3-Way Hook: counted from direction changes > 30 deg in cable geometry",
        "Huawei Mount = Plum Ring = Pole 6M (1:1:1)",
        "Wedge Clamp = 2 x Pole 6M",
        "V-Shape Slack = 1 per hub with MDUs",
        "M8 Microloop + Slack Bracket + Universal Bracket = 1 per MDU",
        "Pigtails: ONLY for standalone blocks (NOT MDUs)",
        "FAT 8-core (40%) + FAT 9-core (60%) = ONLY for standalone blocks",
    ]
    for i, rel in enumerate(relations):
        ws.merge_cells(f"B{r2+1+i}:D{r2+1+i}")
        ws.cell(row=r2+1+i, column=2, value=f"  {i+1}. {rel}").font = Font(size=9, name="Calibri")

    # SHEET 2: BOM Detail
    ws2 = wb.create_sheet("BOM Detail")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 3; ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 55; ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 10; ws2.column_dimensions["F"].width = 45

    ws2.merge_cells("B2:F2")
    ws2["B2"] = f"Auto BOM - {params.get('project_name', 'FTTH')}"
    ws2["B2"].font = Font(color="1F4E78", bold=True, size=16, name="Calibri")
    ws2["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[2].height = 30

    for j, h in enumerate(["Item Code", "Description", "Qty", "Unit", "Notes"]):
        ws2.cell(row=4, column=2+j, value=h)
        ws2.cell(row=4, column=2+j).fill = hdr_fill; ws2.cell(row=4, column=2+j).font = hdr_font
        ws2.cell(row=4, column=2+j).alignment = center; ws2.cell(row=4, column=2+j).border = thin

    # Import items
    try:
        from .auto_calculator import ITEMS
    except ImportError:
        from auto_calculator import ITEMS

    tier_labels = {
        "OLT": "OLT SHELTER TIER", "LMJ": "LMJ MANHOLE TIER",
        "Manhole": "MANHOLE TIER", "HUB": "HUB TIER",
        "Block": "BLOCK TIER (FAT)", "MDU": "MDU TIER",
        "Pole": "POLES & AERIAL ACCESSORIES", "Drop": "DROP CABLES",
    }

    current_tier = None
    r = 5; alt = False
    for item_code, qty, note in bom_items:
        tier = ITEMS.get(item_code, {}).get("tier", "Other")
        if item_code.startswith("SP_"): tier = "Splice Protector"
        elif item_code.startswith("14137938"): tier = "Drop"
        elif item_code in ["01524339", "01524342", "01524343", "01524344", "152433", "34543", "01524337", "01524341"]:
            tier = "Cable Hardware"

        tier_label = tier_labels.get(tier, tier)
        if tier_label != current_tier:
            ws2.merge_cells(f"B{r}:F{r}")
            ws2.cell(row=r, column=2, value=tier_label).font = sec_font
            ws2.cell(row=r, column=2).fill = sec_fill
            for c in range(2, 7): ws2.cell(row=r, column=c).fill = sec_fill; ws2.cell(row=r, column=c).border = thin
            current_tier = tier_label; r += 1; alt = False

        desc = ITEMS.get(item_code, {}).get("desc", note)
        ws2.cell(row=r, column=2, value=item_code).alignment = center; ws2.cell(row=r, column=2).border = thin
        ws2.cell(row=r, column=3, value=desc).alignment = left; ws2.cell(row=r, column=3).border = thin
        ws2.cell(row=r, column=4, value=qty).alignment = center; ws2.cell(row=r, column=4).border = thin
        ws2.cell(row=r, column=5, value="EA").alignment = center; ws2.cell(row=r, column=5).border = thin
        ws2.cell(row=r, column=6, value=note).alignment = left; ws2.cell(row=r, column=6).border = thin
        if alt:
            for c in range(2, 7): ws2.cell(row=r, column=c).fill = alt_fill
        alt = not alt; r += 1

    # Total
    ws2.cell(row=r, column=2, value="").border = thin
    ws2.cell(row=r, column=3, value="TOTAL LINE ITEMS").font = Font(bold=True, size=11, name="Calibri")
    ws2.cell(row=r, column=3).alignment = left; ws2.cell(row=r, column=3).border = thin
    ws2.cell(row=r, column=3).fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    ws2.cell(row=r, column=4, value=len(bom_items)).font = Font(bold=True, size=11, name="Calibri")
    ws2.cell(row=r, column=4).alignment = center; ws2.cell(row=r, column=4).border = thin
    ws2.cell(row=r, column=4).fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    for c in [5, 6]: ws2.cell(row=r, column=c).border = thin; ws2.cell(row=r, column=c).fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    wb.save(output_path)
    return True, f"BOM exported to: {output_path}"
