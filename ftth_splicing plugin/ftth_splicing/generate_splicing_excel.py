#!/usr/bin/env python3
"""
generate_splicing_excel.py  v4.0

Generates FTTH Splicing Diagram Excel from the plugin's enhanced CSV.
Uses only built-in Python modules + openpyxl (no pandas).

KEY FEATURES (v4.0):
  - Dynamic DC segments: exactly N DC segments per block where N = actual
    number of DJs in that block (was previously fixed at 4).
  - ORIGINAL DC labels preserved from CSV (no longer rewritten).
  - Per-block numbering: DJ001-DJ00N, DC{block}.{pos}, SP_001-SP_00N
    resets for every block independently.
  - Staircase link-fibre pattern: DJ at position P shows Leg 9 in
    DC segments 1..(P-1) and Leg 1-8 in DC segment P.
  - 8 rows per DJ (Leg 1-8), houses sorted alphabetically.
  - Handles blocks with 1-4 DJs (including MDU with 1 DJ).

Usage:
    python generate_splicing_excel.py <input_csv> <output_xlsx>
"""

import sys
import re
import os
import csv

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed.")
    print("Install it:  python -m pip install openpyxl")
    print("Or use QGIS Python:  C:\\Program Files\\QGIS 3.40\\bin\\python.exe -m pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
TITLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
BOLD_FONT = Font(bold=True, size=10)
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=12)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

COLS_PER_DC = 13
TRAILING_COLS = 3


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def parse_args():
    if len(sys.argv) < 3:
        print(__doc__)
        print("Usage: python generate_splicing_excel.py <input_csv> <output_xlsx>")
        sys.exit(1)
    return sys.argv[1], sys.argv[2]


def _val(row, key, default=""):
    """Safely extract a cleaned string value from a CSV row dict."""
    v = row.get(key, default)
    if v is None:
        return default
    v = str(v).strip()
    return "" if v.lower() in ("", "nan", "none", "null") else v


def get_block_from_ag(ag_name):
    """GMGZ2_AG01 -> B001"""
    if not ag_name:
        return "B001"
    m = re.search(r"AG(\d+)", str(ag_name))
    return f"B{int(m.group(1)):03d}" if m else "B001"


def get_block_num(block_name):
    """'B001' -> 1, 'B013' -> 13"""
    if not block_name:
        return 0
    m = re.search(r"(\d+)", str(block_name))
    return int(m.group(1)) if m else 0


def get_fc_from_fj(fj_name):
    """'GMGZ2_FJ01_AG13' -> 'FC01'"""
    if not fj_name:
        return "FC01"
    m = re.search(r"FJ(\d+)", str(fj_name))
    return f"FC{int(m.group(1)):02d}" if m else "FC01"


def extract_short_area(dj_name):
    """Extract short area code from DJ name. 'GMGZ2_DJ205_1:8' -> 'GMGZ2'"""
    if not dj_name:
        return ""
    m = re.match(r"([A-Za-z0-9]+)_DJ", str(dj_name))
    return m.group(1) if m else ""


def extract_dc_info(dc_name):
    """Extract size, length, length+slack from DC name."""
    if not dc_name:
        return {"size": "", "length": "", "length_slack": ""}
    name = str(dc_name)
    size_m = re.search(r"_(\d+F)_", name)
    size = size_m.group(1) if size_m else ""
    len_m = re.search(r"\((\d+)m\s*-\s*(\d+)m\)", name)
    if len_m:
        return {
            "size": size,
            "length": int(len_m.group(1)),
            "length_slack": int(len_m.group(2)),
        }
    len_m2 = re.search(r"\((\d+)m\)", name)
    if len_m2:
        d = int(len_m2.group(1))
        return {"size": size, "length": d, "length_slack": d + 10}
    return {"size": size, "length": "", "length_slack": ""}


# ---------------------------------------------------------------------------
# Data Loading
# ---------------------------------------------------------------------------

def load_and_group(csv_path):
    """
    Load CSV and group rows by (ag_name -> block -> position).

    Returns:
        {ag_name: {block_name: [dj_entry, ...]}}

    Each dj_entry:
        {
            "position": int,           # 1, 2, 3, or 4 (within block)
            "original_dj_name": str,   # e.g. "GMGZ2_DJ205_1:8"
            "original_dc_label": str,  # e.g. "VTN_HHS_GMGZ2_AG13_DC053.4_..."
            "pole_number": str,        # e.g. "GMGZ2_P0147"
            "fj_name": str,            # e.g. "GMGZ2_FJ01_AG13"
            "fj_pole": str,
            "ag_name": str,            # e.g. "GMGZ2_AG13"
            "splitter": str,           # e.g. "1:8"  (from CSV splitter column)
            "short_area": str,         # e.g. "GMGZ2"
            "houses": [str, ...],      # sorted list of house names
        }
    """
    rows = []
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)

    # Group by (ag_name, block_name, position)
    ag_groups = {}  # {ag_name: {block_name: {position: dj_dict}}}

    for row in rows:
        ag = _val(row, "ag_name", "Unknown")
        block = get_block_from_ag(ag)
        pos_str = _val(row, "position", "1")
        try:
            pos = int(pos_str)
        except ValueError:
            pos = 1

        if ag not in ag_groups:
            ag_groups[ag] = {}
        if block not in ag_groups[ag]:
            ag_groups[ag][block] = {}

        if pos not in ag_groups[ag][block]:
            original_dj = _val(row, "name_3")
            original_dc = _val(row, "dc_name")
            ag_groups[ag][block][pos] = {
                "position": pos,
                "original_dj_name": original_dj,
                "original_dc_label": original_dc,
                "pole_number": _val(row, "name_2_2"),
                "fj_name": _val(row, "fj_name"),
                "fj_pole": _val(row, "fj_pole"),
                "ag_name": ag,
                "splitter": _val(row, "splitter", "1:8"),
                "short_area": extract_short_area(original_dj),
                "houses": [],
            }

        house = _val(row, "Name_2")
        if house:
            ag_groups[ag][block][pos]["houses"].append(house)

    # Sort houses alphabetically within each DJ, convert to sorted list
    for ag in ag_groups:
        for block in ag_groups[ag]:
            for pos in ag_groups[ag][block]:
                ag_groups[ag][block][pos]["houses"] = sorted(
                    set(ag_groups[ag][block][pos]["houses"])
                )

    # Convert inner dicts to sorted lists
    for ag in ag_groups:
        for block in ag_groups[ag]:
            positions = sorted(ag_groups[ag][block].keys())
            ag_groups[ag][block] = [
                ag_groups[ag][block][p] for p in positions
            ]

    return ag_groups


def compute_per_block_names(dj_entry, block_num):
    """
    Given a DJ entry and its block number, compute all per-block names.

    Returns dict with:
        dj_name, splitter_number, dc_number, fc, ag, pole,
        dc_label (ORIGINAL), dc_size, dc_length, dc_slack
    """
    pos = dj_entry["position"]
    short = dj_entry["short_area"]
    splitter = dj_entry["splitter"]

    # Per-block DJ name: GMGZ2_DJ001_1:8
    dj_name = f"{short}_DJ{pos:03d}_{splitter}"

    # Per-block splitter number: SP_001
    splitter_number = f"SP_{pos:03d}"

    # Per-block DC number: DC13.4
    dc_number = f"DC{block_num}.{pos}"

    # ORIGINAL DC label preserved from CSV (NOT rewritten)
    dc_label = dj_entry["original_dc_label"]

    # FC from FJ
    fc = get_fc_from_fj(dj_entry["fj_name"])

    # AG as-is
    ag = dj_entry["ag_name"]

    # Pole as-is
    pole = dj_entry["pole_number"]

    # DC size and lengths from original DC label
    dc_info = extract_dc_info(dj_entry["original_dc_label"])

    return {
        "dj_name": dj_name,
        "splitter_number": splitter_number,
        "dc_number": dc_number,
        "dc_label": dc_label,
        "fc": fc,
        "ag": ag,
        "pole": pole,
        "dc_size": dc_info["size"],
        "dc_length": dc_info["length"],
        "dc_slack": dc_info["length_slack"],
        "splitter": splitter,
    }


# ---------------------------------------------------------------------------
# Excel Generation
# ---------------------------------------------------------------------------

def build_excel(ag_groups, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Splicing Diagram"

    # Determine the max DJs across all blocks for header sizing
    max_djs_per_block = 0
    block_dj_counts = []  # (ag_name, block_name, num_djs)
    for ag_name in sorted(ag_groups.keys()):
        for block_name in sorted(ag_groups[ag_name].keys()):
            num_djs = len(ag_groups[ag_name][block_name])
            block_dj_counts.append((ag_name, block_name, num_djs))
            if num_djs > max_djs_per_block:
                max_djs_per_block = num_djs

    # Total columns based on max DJs (for header row spanning)
    total_dc_cols_max = max_djs_per_block * COLS_PER_DC
    total_cols_max = total_dc_cols_max + TRAILING_COLS
    slack_col_max = total_dc_cols_max + 1

    # --- Row 1: Title row ---
    ws.merge_cells(
        start_row=1, start_column=1,
        end_row=1, end_column=total_dc_cols_max,
    )
    c = ws.cell(row=1, column=1, value="Distribution Layer 1 (A - B)")
    c.font = TITLE_FONT
    c.alignment = CENTER_ALIGNMENT
    c.fill = HEADER_FILL

    ws.merge_cells(
        start_row=1, start_column=slack_col_max,
        end_row=1, end_column=total_cols_max,
    )
    c = ws.cell(
        row=1, column=slack_col_max,
        value="PREMIS DATA (MUST COMPLETE ALL COLUMS & DUPLICATE INFO REVIEWED)",
    )
    c.font = TITLE_FONT
    c.alignment = CENTER_ALIGNMENT
    c.fill = TITLE_FILL

    for col in range(1, total_cols_max + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL if col <= total_dc_cols_max else TITLE_FILL
        cell.border = THIN_BORDER

    # --- Row 2: Category headers ---
    for dc_idx in range(max_djs_per_block):
        base = dc_idx * COLS_PER_DC + 1

        # Feeder / Link (cols 1-3)
        ws.merge_cells(
            start_row=2, start_column=base,
            end_row=2, end_column=base + 2,
        )
        c = ws.cell(row=2, column=base, value="Feeder / Link")
        c.font = BOLD_FONT
        c.alignment = CENTER_ALIGNMENT
        c.fill = SUBHEADER_FILL
        c.border = THIN_BORDER
        for o in range(3):
            ws.cell(row=2, column=base + o).border = THIN_BORDER
            ws.cell(row=2, column=base + o).fill = SUBHEADER_FILL

        # Cable Route ID (AERIAL)_N (cols 4-6)
        ws.merge_cells(
            start_row=2, start_column=base + 3,
            end_row=2, end_column=base + 5,
        )
        c = ws.cell(
            row=2, column=base + 3,
            value=f"Cable Route ID (AERIAL)_{dc_idx + 1}",
        )
        c.font = BOLD_FONT
        c.alignment = CENTER_ALIGNMENT
        c.fill = SUBHEADER_FILL
        c.border = THIN_BORDER
        for o in range(3):
            ws.cell(row=2, column=base + 3 + o).border = THIN_BORDER
            ws.cell(row=2, column=base + 3 + o).fill = SUBHEADER_FILL

        # Fibre (col 8 -> index 7)
        c = ws.cell(row=2, column=base + 7, value="Fibre")
        c.font = BOLD_FONT
        c.alignment = CENTER_ALIGNMENT
        c.fill = SUBHEADER_FILL
        c.border = THIN_BORDER

        # Splitter (col 10 -> index 9)
        c = ws.cell(row=2, column=base + 9, value="Splitter")
        c.font = BOLD_FONT
        c.alignment = CENTER_ALIGNMENT
        c.fill = SUBHEADER_FILL
        c.border = THIN_BORDER

        # Access Box & Joint (cols 11-12 -> indices 10-11)
        ws.merge_cells(
            start_row=2, start_column=base + 10,
            end_row=2, end_column=base + 11,
        )
        c = ws.cell(row=2, column=base + 10, value="Access Box & Joint")
        c.font = BOLD_FONT
        c.alignment = CENTER_ALIGNMENT
        c.fill = SUBHEADER_FILL
        c.border = THIN_BORDER
        for o in range(2):
            ws.cell(row=2, column=base + 10 + o).border = THIN_BORDER
            ws.cell(row=2, column=base + 10 + o).fill = SUBHEADER_FILL

        # Distance (col 13 -> index 12)
        c = ws.cell(row=2, column=base + 12, value="Distance")
        c.font = BOLD_FONT
        c.alignment = CENTER_ALIGNMENT
        c.fill = SUBHEADER_FILL
        c.border = THIN_BORDER

    # Trailing category headers
    ws.merge_cells(
        start_row=2, start_column=slack_col_max,
        end_row=2, end_column=slack_col_max + 1,
    )
    c = ws.cell(row=2, column=slack_col_max, value="+ Slack")
    c.font = BOLD_FONT
    c.alignment = CENTER_ALIGNMENT
    c.fill = SUBHEADER_FILL
    c.border = THIN_BORDER
    ws.cell(row=2, column=slack_col_max + 1).border = THIN_BORDER
    ws.cell(row=2, column=slack_col_max + 1).fill = SUBHEADER_FILL

    c = ws.cell(row=2, column=slack_col_max + 2, value="Block")
    c.font = BOLD_FONT
    c.alignment = CENTER_ALIGNMENT
    c.fill = SUBHEADER_FILL
    c.border = THIN_BORDER

    # --- Row 3: Detailed column headers ---
    detail_headers = [
        "Pole", "FEEDER", "AG",
        "DC label", "DC number", "Size",
        "No.", "splitter number", "Leg",
        "Pole number", "DJ number",
        "DC length", "DC length+10(slack)",
    ]
    for dc_idx in range(max_djs_per_block):
        base = dc_idx * COLS_PER_DC + 1
        for hi, hdr in enumerate(detail_headers):
            c = ws.cell(row=3, column=base + hi, value=hdr)
            c.font = BOLD_FONT
            c.alignment = CENTER_ALIGNMENT
            c.fill = HEADER_FILL
            c.border = THIN_BORDER

    for off, hdr in enumerate(["slack length", "premise data name", "Block"]):
        c = ws.cell(row=3, column=slack_col_max + off, value=hdr)
        c.font = BOLD_FONT
        c.alignment = CENTER_ALIGNMENT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER

    # --- Data Rows ---
    current_row = 4

    for ag_name, block_name, num_djs in block_dj_counts:
        djs = ag_groups[ag_name][block_name]
        block_num = get_block_num(block_name)

        # Pre-compute per-block names for all DJs in this block
        dj_info_list = []
        for dj_entry in djs:
            info = compute_per_block_names(dj_entry, block_num)
            dj_info_list.append(info)

        # For each DJ in the block (each contributes 8 rows: Leg 1-8)
        for dj_idx in range(num_djs):
            dj_entry = djs[dj_idx]
            houses = dj_entry["houses"]

            for leg in range(1, 9):  # 8 rows per DJ
                # For each DC segment that ACTUALLY exists in this block:
                #   If segment S < DJ position: show Leg 9 (link fiber)
                #   If segment S == DJ position: show Leg 1-8 (house fiber)
                #   If segment S > DJ position: empty (not visible yet)
                for seg_idx in range(num_djs):
                    base = seg_idx * COLS_PER_DC + 1
                    source_info = dj_info_list[seg_idx]

                    if seg_idx < dj_idx:
                        # Upstream DJ: show Leg 9 (link fiber)
                        show_leg = "9"
                    elif seg_idx == dj_idx:
                        # Current DJ: show the actual leg (1-8)
                        show_leg = str(leg)
                    else:
                        # Downstream DJ: not visible yet, skip
                        continue

                    # Populate all 13 columns for this DC segment
                    ws.cell(row=current_row, column=base,
                            value=source_info["pole"])
                    ws.cell(row=current_row, column=base + 1,
                            value=source_info["fc"])
                    ws.cell(row=current_row, column=base + 2,
                            value=source_info["ag"])
                    ws.cell(row=current_row, column=base + 3,
                            value=source_info["dc_label"])
                    ws.cell(row=current_row, column=base + 4,
                            value=source_info["dc_number"])
                    ws.cell(row=current_row, column=base + 5,
                            value=source_info["dc_size"])
                    ws.cell(row=current_row, column=base + 6,
                            value="1")  # No. = 1 (1F)
                    ws.cell(row=current_row, column=base + 7,
                            value=source_info["splitter_number"])
                    ws.cell(row=current_row, column=base + 8,
                            value=show_leg)
                    ws.cell(row=current_row, column=base + 9,
                            value=source_info["pole"])
                    ws.cell(row=current_row, column=base + 10,
                            value=source_info["dj_name"])
                    # Format lengths with 4 digits (zero-padded)
                    _dl = source_info["dc_length"]
                    _ds = source_info["dc_slack"]
                    ws.cell(row=current_row, column=base + 11,
                            value=f"{int(_dl):04d}" if _dl != '' else '')
                    ws.cell(row=current_row, column=base + 12,
                            value=f"{int(_ds):04d}" if _ds != '' else '')

                # Trailing columns (positioned based on this block's DJ count)
                block_total_dc_cols = num_djs * COLS_PER_DC
                block_slack_col = block_total_dc_cols + 1

                # House name: assign from current DJ's houses, empty if no house
                if leg <= len(houses):
                    house_name = houses[leg - 1]
                else:
                    house_name = ""

                ws.cell(row=current_row, column=block_slack_col, value="")
                ws.cell(row=current_row, column=block_slack_col + 1,
                        value=house_name)
                ws.cell(row=current_row, column=block_slack_col + 2,
                        value=block_name)

                # Apply borders and alignment to all populated columns
                for col in range(1, block_slack_col + 3):
                    ws.cell(row=current_row, column=col).border = THIN_BORDER
                    ws.cell(row=current_row, column=col).alignment = CENTER_ALIGNMENT

                current_row += 1

    # --- Column widths ---
    for dc_idx in range(max_djs_per_block):
        base = dc_idx * COLS_PER_DC + 1
        ws.column_dimensions[get_column_letter(base)].width = 12       # Pole
        ws.column_dimensions[get_column_letter(base + 1)].width = 10   # FEEDER
        ws.column_dimensions[get_column_letter(base + 2)].width = 12   # AG
        ws.column_dimensions[get_column_letter(base + 3)].width = 55   # DC label
        ws.column_dimensions[get_column_letter(base + 4)].width = 10   # DC number
        ws.column_dimensions[get_column_letter(base + 5)].width = 8    # Size
        ws.column_dimensions[get_column_letter(base + 6)].width = 6    # No.
        ws.column_dimensions[get_column_letter(base + 7)].width = 18   # splitter number
        ws.column_dimensions[get_column_letter(base + 8)].width = 8    # Leg
        ws.column_dimensions[get_column_letter(base + 9)].width = 12   # Pole number
        ws.column_dimensions[get_column_letter(base + 10)].width = 25  # DJ number
        ws.column_dimensions[get_column_letter(base + 11)].width = 12  # DC length
        ws.column_dimensions[get_column_letter(base + 12)].width = 20  # DC length+slack

    ws.column_dimensions[get_column_letter(slack_col_max)].width = 14      # slack length
    ws.column_dimensions[get_column_letter(slack_col_max + 1)].width = 30  # premise data name
    ws.column_dimensions[get_column_letter(slack_col_max + 2)].width = 10  # Block

    ws.freeze_panes = "A4"
    wb.save(output_path)

    total_rows = current_row - 4
    print(f"Saved: {output_path}")
    print(f"  Data rows: {total_rows}")
    print(f"  Max DJs per block: {max_djs_per_block}")
    print(f"  Total header cols: {total_cols_max}")


def main():
    input_csv, output_xlsx = parse_args()
    if not os.path.exists(input_csv):
        print(f"Error: File not found: {input_csv}")
        sys.exit(1)
    out_dir = os.path.dirname(output_xlsx)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    print(f"Reading: {input_csv}")
    ag_groups = load_and_group(input_csv)
    total_blocks = sum(len(bl) for ag in ag_groups.values() for bl in ag.values())
    total_djs = sum(
        len(bl) for ag in ag_groups.values() for bl in ag.values()
    )
    print(f"Found {len(ag_groups)} AGs, {total_blocks} blocks, {total_djs} DJ groups")

    print(f"Writing: {output_xlsx}")
    build_excel(ag_groups, output_xlsx)
    print("Done!")


if __name__ == "__main__":
    main()
