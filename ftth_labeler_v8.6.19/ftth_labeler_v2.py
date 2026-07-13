# -*- coding: utf-8 -*-
"""
FTTH Labeler v8.6 (internal) — Q1.0 release
Copyright (c) Mustafa M M Ellaham. All rights reserved.

Proprietary and confidential. Unauthorized copying, distribution,
modification, reverse engineering, or use of this plugin, in whole
or in part, is strictly prohibited without express written permission
from Mustafa M M Ellaham.

By Mustafa M M Ellaham

STRICT INPUT (you must provide):
  - AG Polygons: ID column with AG numbers (1, 2, 3...)
  - Block Polygons: ID column with Block numbers (1, 2, 3...)
  - FJ Points: any table (geometry only matters)
  - DJ Points: any table (geometry only matters)
  - DC Lines: individual segments between 2 nodes, any direction

FLEXIBLE INPUT (plugin auto-generates columns):
  - All layers except AG/Block: plugin creates 'name' and all other needed columns
  - AG/Block: plugin auto-creates 'name' column if missing

PARAMETERS (you enter when running):
  - Area Code: any length (type YOUR actual area code)
  - Zone Code: e.g., Z01, Z02 (optional — leave empty for no zone)
  - AG ID Field Name: defaults to 'ID'
  - Block ID Field Name: defaults to 'ID'

OUTPUT (plugin generates and writes to layers):
  - AG.name = AG01, AG02... (2 digits)
  - Block.name = B001, B002... (3 digits) + ag_parent
  - FJ.name = {AREA}_FJ01_{AG}...
  - DJ.name = {AREA}_DJ001_1:9 + splitter + position + dc_name
  - DC.name = {AREA}_{ZONE}_{AG}_DC{block:03d}.{seq}_1F_ADSS_G.657.A1 + from_node + to_node
  - Pole.name = {AREA}_P0001...
"""

import os
import re
import json
import csv
import datetime
from qgis.PyQt.QtCore import QVariant
from qgis.core import (
    QgsApplication, QgsProcessing,
    QgsProject, QgsVectorLayer, QgsFeature, QgsGeometry, QgsPointXY,
    QgsField, QgsFields, QgsWkbTypes,
    QgsProcessingAlgorithm, QgsProcessingParameterVectorLayer,
    QgsProcessingParameterString, QgsProcessingParameterBoolean,
    QgsProcessingParameterNumber, QgsProcessingParameterFolderDestination,
    QgsProcessingParameterFileDestination, QgsProcessingParameterFile,
    QgsProcessingParameterEnum,
    QgsProcessingOutputFolder,
    QgsProcessingProvider,
    QgsVectorFileWriter, QgsCoordinateTransformContext
)
from qgis import processing

from .trace_engine import FTTHTraceEngine, DJPositionMapper
from .bom_engine import calculate_bom

# openpyxl for Excel output (bundled with QGIS)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class FTTHLabelerProvider(QgsProcessingProvider):
    def loadAlgorithms(self, *args, **kwargs):
        self.addAlgorithm(FTTHLabelAlgorithm())
        self.addAlgorithm(FTTHValidatorAlgorithm())
        self.addAlgorithm(FTTHSplicingPlanAlgorithm())
        self.addAlgorithm(FTTHBOMAlgorithm())
    def id(self, *args, **kwargs):
        return 'ftth_labeler'
    def name(self, *args, **kwargs):
        return 'FTTH Tools'
    def icon(self):
        return QgsProcessingProvider.icon(self)


class FTTHLabelerV2Plugin:
    def __init__(self, iface):
        self.iface = iface
        self.provider = FTTHLabelerProvider()
    def initGui(self):
        QgsApplication.processingRegistry().addProvider(self.provider)
    def unload(self):
        QgsApplication.processingRegistry().removeProvider(self.provider)


class FTTHLabelAlgorithm(QgsProcessingAlgorithm):
    IN_AG = 'IN_AG'
    IN_BLOCKS = 'IN_BLOCKS'
    IN_FJ = 'IN_FJ'
    IN_DJ = 'IN_DJ'
    IN_DC = 'IN_DC'
    IN_POLES = 'IN_POLES'
    IN_FC = 'IN_FC'
    
    PARAM_AREA = 'PARAM_AREA'
    PARAM_ZONE = 'PARAM_ZONE'
    PARAM_AG_ID = 'PARAM_AG_ID'
    PARAM_BLOCK_ID = 'PARAM_BLOCK_ID'
    PARAM_FC_ID = 'PARAM_FC_ID'
    PARAM_TARGET = 'PARAM_TARGET'
    PARAM_OVERWRITE = 'PARAM_OVERWRITE'
    PARAM_BUFFER = 'PARAM_BUFFER'
    PARAM_OUTPUT = 'PARAM_OUTPUT'
    
    # Layer selection checkboxes
    PARAM_LABEL_AG   = 'PARAM_LABEL_AG'
    PARAM_LABEL_BLOCK = 'PARAM_LABEL_BLOCK'
    PARAM_LABEL_FJ   = 'PARAM_LABEL_FJ'
    PARAM_LABEL_DJ   = 'PARAM_LABEL_DJ'
    PARAM_LABEL_DC   = 'PARAM_LABEL_DC'
    PARAM_LABEL_POLE = 'PARAM_LABEL_POLE'
    PARAM_LABEL_FC   = 'PARAM_LABEL_FC'
    
    def name(self):
        return 'ftth_labeler_v2'
    def displayName(self):
        return 'FTTH Labeler v3 (QGIS-Native)'
    def group(self):
        return 'FTTH Tools'
    def groupId(self):
        return 'ftth_tools'
    def createInstance(self):
        return FTTHLabelAlgorithm()
    
    def shortHelpString(self):
        return """
        <h3>FTTH Labeler v3.9 — By Mustafa M M Ellaham</h3>
        
        <h4>STRICT INPUT (you must provide):</h4>
        <ul>
        <li><b>AG Polygons</b>: ID column with AG numbers (1,2,3...)</li>
        <li><b>Block Polygons</b>: ID column with block numbers (1,2,3...)</li>
        <li><b>FJ, DJ, DC</b>: any table structure — plugin auto-creates all columns</li>
        </ul>
        
        <h4>OPTIONAL INPUT:</h4>
        <ul>
        <li><b>Feeder Cable (FC) Lines</b>: one line per cable route. ID column with cable numbers. Plugin auto-detects start/end AGs and FJ count along route.</li>
        </ul>
        
        <h4>PARAMETERS (you enter when running):</h4>
        <ul>
        <li><b>Area Code</b>: any length (type YOUR actual area code)</li>
        <li><b>Zone Code</b>: e.g., Z01, Z02 (optional — leave empty for no zone)</li>
        <li><b>AG ID Field</b>: column name with AG numbers (default: ID)</li>
        <li><b>Block ID Field</b>: column name with block numbers (default: ID)</li>
        <li><b>FC ID Field</b>: column name with cable numbers (default: ID)</li>
        <li><b>FC Size/Type columns</b>: Cable size and type (CC0x/FC/LC) read from FC layer</li>
        </ul>
        
        <h4>Huawei Pre-Con Splitter Logic:</h4>
        <ul>
        <li>Position 4 = always 1:8 (terminal)</li>
        <li>Positions 1,2,3 = always 1:9 (intermediate)</li>
        </ul>
        
        <h4>FC Label Format:</h4>
        <p>{Area}_{Zone}_{StartAG}-{EndAG}_FC{ID:02d}_{Size}_ADSS_G.657.A1({Length}m-{Length+Slack}m)</p>
        <p>Slack = 16m per FJ along the route</p>
        """
    
    def initAlgorithm(self, config=None):
        # Required layers
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_AG, 'AG Polygons (MUST have ID column with AG numbers)',
            [QgsProcessing.TypeVectorPolygon]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_BLOCKS, 'Block Polygons (MUST have ID column with block numbers)',
            [QgsProcessing.TypeVectorPolygon]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_FJ, 'Feeder Joint (FJ) Points (any table structure)',
            [QgsProcessing.TypeVectorPoint]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_DJ, 'Distribution Joint (DJ) Points (any table structure)',
            [QgsProcessing.TypeVectorPoint]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_DC, 'Distribution Cable (DC) Lines — individual segments',
            [QgsProcessing.TypeVectorLine]
        ))
        # Optional
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_POLES, 'Pole Points (optional, any table)',
            [QgsProcessing.TypeVectorPoint], optional=True
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_FC, 'Feeder Cable (FC) Lines — optional, one line per cable route',
            [QgsProcessing.TypeVectorLine], optional=True
        ))
        # Parameters
        self.addParameter(QgsProcessingParameterString(
            self.PARAM_AREA, 'Area Code (e.g., VTN_HHS_GMG — type YOUR actual area code)',
            defaultValue=''
        ))
        self.addParameter(QgsProcessingParameterString(
            self.PARAM_ZONE, 'Zone Code (optional, e.g., Z02 — leave empty for no zone)',
            defaultValue='', optional=True
        ))
        self.addParameter(QgsProcessingParameterString(
            self.PARAM_AG_ID, 'AG ID Field Name (your AG number column)',
            defaultValue='ID'
        ))
        self.addParameter(QgsProcessingParameterString(
            self.PARAM_BLOCK_ID, 'Block ID Field Name (your block number column)',
            defaultValue='ID'
        ))
        self.addParameter(QgsProcessingParameterString(
            self.PARAM_FC_ID, 'FC ID Field Name (your feeder cable number column)',
            defaultValue='ID'
        ))
        self.addParameter(QgsProcessingParameterString(
            self.PARAM_TARGET, 'Target Field Name for Labels',
            defaultValue='name'
        ))
        self.addParameter(QgsProcessingParameterBoolean(
            self.PARAM_OVERWRITE, 'Overwrite existing values',
            defaultValue=False
        ))
        # Layer selection checkboxes
        self.addParameter(QgsProcessingParameterBoolean(
            self.PARAM_LABEL_AG, 'Label AG layers', defaultValue=True
        ))
        self.addParameter(QgsProcessingParameterBoolean(
            self.PARAM_LABEL_BLOCK, 'Label Block layers', defaultValue=True
        ))
        self.addParameter(QgsProcessingParameterBoolean(
            self.PARAM_LABEL_FJ, 'Label FJ layers', defaultValue=True
        ))
        self.addParameter(QgsProcessingParameterBoolean(
            self.PARAM_LABEL_DJ, 'Label DJ layers', defaultValue=True
        ))
        self.addParameter(QgsProcessingParameterBoolean(
            self.PARAM_LABEL_DC, 'Label DC cables', defaultValue=True
        ))
        self.addParameter(QgsProcessingParameterBoolean(
            self.PARAM_LABEL_POLE, 'Label Poles', defaultValue=True
        ))
        self.addParameter(QgsProcessingParameterBoolean(
            self.PARAM_LABEL_FC, 'Label FC cables', defaultValue=True
        ))
        self.addParameter(QgsProcessingParameterNumber(
            self.PARAM_BUFFER, 'Node Snap Tolerance (meters)',
            type=QgsProcessingParameterNumber.Double,
            defaultValue=2.0, minValue=0.1, maxValue=50.0
        ))
        self.addParameter(QgsProcessingParameterFolderDestination(
            self.PARAM_OUTPUT, 'Output Folder'
        ))
    
    # ---- Static Helpers ----
    
    @staticmethod
    def _find_field(layer, field_names):
        """Find field index, case-insensitive. Returns -1 if not found."""
        all_names = [f.name() for f in layer.fields()]
        for fn in field_names:
            fn_lower = fn.lower()
            for actual_name in all_names:
                if actual_name.lower() == fn_lower:
                    return layer.fields().indexFromName(actual_name)
        return -1
    
    @staticmethod
    def _list_fields(layer):
        return [(f.name(), f.typeName()) for f in layer.fields()]
    
    @staticmethod
    def _ensure_field(layer, field_name, field_type=QVariant.String, length=100):
        """Create field via dataProvider (safe when not editing)."""
        idx = FTTHLabelAlgorithm._find_field(layer, [field_name])
        if idx >= 0:
            return idx
        try:
            dp = layer.dataProvider()
            dp.addAttributes([QgsField(field_name, field_type, len=length)])
            layer.updateFields()
            return FTTHLabelAlgorithm._find_field(layer, [field_name])
        except Exception:
            return -1
    
    @staticmethod
    def _safe_write(layer, feat_id, field_idx, value, overwrite=False):
        """Write attribute safely. Handles QGIS NULL (QVariant) properly."""
        if field_idx < 0 or feat_id < 0:
            return False
        try:
            # CRITICAL: Must be in edit mode for changeAttributeValue
            if not layer.isEditable():
                return False
            # ALWAYS validate feature exists before writing — prevents QGIS crash
            feat = layer.getFeature(feat_id)
            if not feat.isValid():
                return False
            if overwrite:
                layer.changeAttributeValue(feat_id, field_idx, value)
                return True
            current = feat.attribute(field_idx)
            if current is None:
                layer.changeAttributeValue(feat_id, field_idx, value)
                return True
            try:
                from qgis.core import QgsVariantUtils
                if QgsVariantUtils.isNull(current):
                    layer.changeAttributeValue(feat_id, field_idx, value)
                    return True
            except Exception:
                pass
            if str(current).strip() in ('', 'NULL', 'None'):
                layer.changeAttributeValue(feat_id, field_idx, value)
                return True
            return False
        except Exception:
            return False
    
    @staticmethod
    def _batch_write_via_dataprovider(layer, updates, field_idx=None):
        """Batch write attributes via dataProvider — bypasses edit buffer.
        updates: {feat_id: value} dict
        field_idx: target field index (auto-detected if None)
        Returns: number of successful writes
        """
        if not updates:
            return 0
        dp = layer.dataProvider()
        if field_idx is None:
            field_idx = FTTHLabelAlgorithm._find_field(layer, ['name', 'Name', 'NAME'])
        if field_idx < 0:
            return 0
        changes = {}
        for feat_id, value in updates.items():
            changes[feat_id] = {field_idx: value}
        try:
            ok = dp.changeAttributeValues(changes)
            return len(changes) if ok else 0
        except Exception:
            return 0
    
    # ---- Main Processing ----
    
    def processAlgorithm(self, parameters, context, feedback):
        area_code = self.parameterAsString(parameters, self.PARAM_AREA, context)
        zone = self.parameterAsString(parameters, self.PARAM_ZONE, context)
        ag_id_field = self.parameterAsString(parameters, self.PARAM_AG_ID, context) or 'ID'
        block_id_field = self.parameterAsString(parameters, self.PARAM_BLOCK_ID, context) or 'ID'
        fc_id_field = self.parameterAsString(parameters, self.PARAM_FC_ID, context) or 'ID'
        fc_size_default = '144F'  # fallback when FC layer has no 'size' column
        target_field = self.parameterAsString(parameters, self.PARAM_TARGET, context) or 'name'
        overwrite = self.parameterAsBoolean(parameters, self.PARAM_OVERWRITE, context)
        # Read layer selection flags
        label_ag   = self.parameterAsBoolean(parameters, self.PARAM_LABEL_AG, context)
        label_block = self.parameterAsBoolean(parameters, self.PARAM_LABEL_BLOCK, context)
        label_fj   = self.parameterAsBoolean(parameters, self.PARAM_LABEL_FJ, context)
        label_dj   = self.parameterAsBoolean(parameters, self.PARAM_LABEL_DJ, context)
        label_dc   = self.parameterAsBoolean(parameters, self.PARAM_LABEL_DC, context)
        label_pole = self.parameterAsBoolean(parameters, self.PARAM_LABEL_POLE, context)
        label_fc   = self.parameterAsBoolean(parameters, self.PARAM_LABEL_FC, context)
        tolerance = self.parameterAsDouble(parameters, self.PARAM_BUFFER, context)
        output_folder = self.parameterAsString(parameters, self.PARAM_OUTPUT, context)
        
        ag_layer = self.parameterAsVectorLayer(parameters, self.IN_AG, context)
        blocks_layer = self.parameterAsVectorLayer(parameters, self.IN_BLOCKS, context)
        fj_layer = self.parameterAsVectorLayer(parameters, self.IN_FJ, context)
        dj_layer = self.parameterAsVectorLayer(parameters, self.IN_DJ, context)
        dc_layer = self.parameterAsVectorLayer(parameters, self.IN_DC, context)
        poles_layer = self.parameterAsVectorLayer(parameters, self.IN_POLES, context)
        fc_layer = self.parameterAsVectorLayer(parameters, self.IN_FC, context)
        
        feedback.pushInfo("=" * 60)
        feedback.pushInfo("FTTH Labeler V1.0 (build 8.6.19) — By Mustafa M M Ellaham")
        feedback.pushInfo("=" * 60)
        feedback.pushInfo(f"Area Code: {area_code}")
        feedback.pushInfo(f"Zone: {zone}")
        feedback.pushInfo(f"AG ID Field: {ag_id_field}")
        feedback.pushInfo(f"Block ID Field: {block_id_field}")
        if fc_layer:
            feedback.pushInfo(f"FC ID Field: {fc_id_field}")
            feedback.pushInfo(f"FC Cable Size: from FC layer 'size' column (fallback: {fc_size_default})")
        feedback.pushInfo(f"Target Field: {target_field}")
        feedback.pushInfo(f"Snap Tolerance: {tolerance}m")
        feedback.pushInfo("")
        feedback.pushInfo("Layer selection:")
        feedback.pushInfo("  AG:   {}".format('YES' if label_ag else 'SKIPPED'))
        feedback.pushInfo("  Block: {}".format('YES' if label_block else 'SKIPPED'))
        feedback.pushInfo("  FJ:   {}".format('YES' if label_fj else 'SKIPPED'))
        feedback.pushInfo("  DJ:   {}".format('YES' if label_dj else 'SKIPPED'))
        feedback.pushInfo("  DC:   {}".format('YES' if label_dc else 'SKIPPED'))
        feedback.pushInfo("  Pole: {}".format('YES' if label_pole else 'SKIPPED'))
        feedback.pushInfo("  FC:   {}".format('YES' if label_fc else 'SKIPPED'))
        
        # ===== PHASE 0: Read AG and Block IDs =====
        feedback.setProgress(5)
        feedback.pushInfo("[Phase 0] Reading AG and Block IDs...")
        
        # Create engine FIRST (needed for name generation in logs)
        engine = FTTHTraceEngine(area_code, zone)
        
        # Read AGs — store as (feat_id, feat, ag_num) with SHORT name AG01
        ag_id_idx = self._find_field(ag_layer, [ag_id_field, 'id', 'ID', 'Id', 'fid', 'FID'])
        feedback.pushInfo(f"  AG layer fields: {self._list_fields(ag_layer)}")
        feedback.pushInfo(f"  AG ID field index: {ag_id_idx}")
        
        ag_data = []
        if ag_id_idx >= 0:
            for feat in ag_layer.getFeatures():
                val = feat.attribute(ag_id_idx)
                if val is not None:
                    try:
                        ag_num = int(val)
                        ag_data.append((feat.id(), feat, ag_num))
                    except:
                        pass
        
        feedback.pushInfo(f"  Found {len(ag_data)} AGs:")
        for _, _, num in ag_data:
            feedback.pushInfo(f"    ID={num} -> AG{num:02d} (full: {engine.generate_ag_name(num)})")
        
        if len(ag_data) == 0:
            feedback.pushInfo("  ERROR: No AGs found. Check AG ID field name.")
            return {self.PARAM_OUTPUT: output_folder}
        
        # Read Blocks — store as (feat_id, feat, block_num, ag_short_name)
        block_id_idx = self._find_field(blocks_layer, [block_id_field, 'id', 'ID', 'Id', 'fid', 'FID'])
        feedback.pushInfo(f"  Block layer fields: {self._list_fields(blocks_layer)}")
        feedback.pushInfo(f"  Block ID field index: {block_id_idx}")
        
        block_data = []
        if block_id_idx >= 0:
            for feat in blocks_layer.getFeatures():
                val = feat.attribute(block_id_idx)
                if val is not None:
                    try:
                        b_num = int(val)
                        # Find parent AG by point-in-polygon
                        ag_short = 'AG01'  # default
                        centroid = feat.geometry().centroid()
                        for _, ag_feat, ag_num in ag_data:
                            if ag_feat.geometry().contains(centroid):
                                ag_short = f"AG{ag_num:02d}"
                                break
                        block_data.append((feat.id(), feat, b_num, f"B{b_num:03d}", ag_short))
                    except:
                        pass
        
        feedback.pushInfo(f"  Found {len(block_data)} Blocks:")
        for _, _, num, bname, ag_short in block_data:
            feedback.pushInfo(f"    ID={num} -> {bname} (AG: {ag_short}, full: {engine.generate_block_name(num)})")
        
        if len(block_data) == 0:
            feedback.pushInfo("  ERROR: No Blocks found. Check Block ID field name.")
            return {self.PARAM_OUTPUT: output_folder}
        
        # Read FC data (optional)
        fc_data = []
        fc_labels_map = {}
        if fc_layer:
            feedback.pushInfo("")
            feedback.pushInfo("[FC] Reading Feeder Cable lines...")
            fc_id_idx = self._find_field(fc_layer, [fc_id_field, 'id', 'ID', 'Id'])
            fc_size_idx = self._find_field(fc_layer, ['size', 'Size', 'SIZE', 'cable_size'])
            fc_type_idx = self._find_field(fc_layer, ['type', 'Type', 'TYPE', 'cable_type'])
            feedback.pushInfo(f"  FC layer fields: {self._list_fields(fc_layer)}")
            feedback.pushInfo(f"  FC ID field index: {fc_id_idx}")
            if fc_size_idx >= 0:
                feedback.pushInfo(f"  FC size field index: {fc_size_idx}")
            if fc_type_idx >= 0:
                feedback.pushInfo(f"  FC type field index: {fc_type_idx}")
            
            fc_raw = []
            if fc_id_idx >= 0:
                for feat in fc_layer.getFeatures():
                    val = feat.attribute(fc_id_idx)
                    if val is not None:
                        try:
                            fc_num = int(val)
                            # Read size and type from FC layer if available
                            fc_size_val = fc_size_default
                            if fc_size_idx >= 0:
                                s = feat.attribute(fc_size_idx)
                                if s is not None and str(s).strip():
                                    fc_size_val = f"{int(s)}F"
                            fc_type_val = 'FC'
                            if fc_type_idx >= 0:
                                t = feat.attribute(fc_type_idx)
                                if t is not None and str(t).strip():
                                    fc_type_val = str(t).strip()
                            fc_raw.append((feat.id(), feat, fc_num, fc_size_val, fc_type_val))
                        except:
                            pass
            
            # Sort by ID, then build per-type counters
            fc_raw.sort(key=lambda x: x[2])  # sort by ID
            type_counters = {}  # e.g., {'CC0x': 0, 'FC': 0, 'LC': 0}
            
            # Pre-index FJ geometries for intersection counting
            fj_geoms = [(f.id(), f.geometry()) for f in fj_layer.getFeatures()]
            
            # Pre-index AG geometries with their numbers
            ag_geom_index = [(ag_num, ag_feat.geometry()) for _, ag_feat, ag_num in ag_data]
            
            for feat_id, feat, fc_num, cable_size_from_layer, fc_type in fc_raw:
                fc_geom = feat.geometry()
                route_length = fc_geom.length()
                
                # Per-type counter: each type group starts from 1
                if fc_type not in type_counters:
                    type_counters[fc_type] = 0
                type_counters[fc_type] += 1
                type_num = type_counters[fc_type]
                
                # Find AGs intersected by this FC line
                intersected_ags = []
                for ag_num, ag_geom in ag_geom_index:
                    if fc_geom.intersects(ag_geom):
                        intersected_ags.append(ag_num)
                
                if len(intersected_ags) >= 2:
                    ag_start = min(intersected_ags)
                    ag_end = max(intersected_ags)
                elif len(intersected_ags) == 1:
                    ag_start = ag_end = intersected_ags[0]
                else:
                    # Fallback: nearest AG at start and end of line
                    start_pt = fc_geom.interpolate(0)
                    end_pt = fc_geom.interpolate(fc_geom.length())
                    ag_start = ag_geom_index[0][0] if ag_geom_index else 1
                    ag_end = ag_geom_index[-1][0] if ag_geom_index else 1
                    min_start = min_end = float('inf')
                    for ag_num, ag_geom in ag_geom_index:
                        d_start = start_pt.distance(ag_geom)
                        d_end = end_pt.distance(ag_geom)
                        if d_start < min_start:
                            min_start = d_start
                            ag_start = ag_num
                        if d_end < min_end:
                            min_end = d_end
                            ag_end = ag_num
                
                ag_start_short = f"AG{ag_start:02d}"
                ag_end_short = f"AG{ag_end:02d}"
                
                # Count FJs that this FC line passes through (intersects within tolerance)
                fj_count = 0
                for _, fj_geom in fj_geoms:
                    if fc_geom.distance(fj_geom) <= tolerance:
                        fj_count += 1
                
                fc_name = engine.generate_fc_name(
                    ag_start_short, ag_end_short, type_num,
                    cable_size=cable_size_from_layer, route_length=route_length,
                    fj_count=fj_count, fc_type=fc_type
                )
                
                fc_slack = int(route_length * 0.08)
                fc_data.append((feat_id, feat, fc_num))
                fc_labels_map[feat_id] = {
                    'fc_name': fc_name,
                    'fc_num': type_num,
                    'fc_type': fc_type,
                    'ag_start': ag_start_short,
                    'ag_end': ag_end_short,
                    'route_length': route_length,
                    'fj_count': fj_count,
                    'slack': fc_slack
                }
                display_prefix = fc_type if fc_type.upper() == 'CC0X' else f"{fc_type}{type_num:02d}"
                feedback.pushInfo(f"    {display_prefix}: {ag_start_short}-{ag_end_short}, "
                                  f"{route_length:.0f}m, {fj_count} FJs, "
                                  f"slack={fc_slack}m (8%), size={cable_size_from_layer}")
            
            feedback.pushInfo(f"  Found {len(fc_data)} Feeder Cable(s)")
        
        # ===== PHASE 1: Read geometry layers (no ID needed) =====
        feedback.setProgress(15)
        feedback.pushInfo("")
        feedback.pushInfo("[Phase 1] Reading FJ, DJ, DC geometry layers...")
        
        fj_data = [(f.id(), f, '') for f in fj_layer.getFeatures()]
        # Assign AG to each FJ by POINT-IN-POLYGON (correct method)
        # FJ must be INSIDE its AG polygon — centroid distance is wrong for irregular shapes
        for i, (fid, feat, _) in enumerate(fj_data):
            ag_assigned = None
            # Step 1: Check if FJ point is INSIDE any AG polygon
            fj_geom = feat.geometry()
            for _, ag_feat, ag_num in ag_data:
                if ag_feat.geometry().contains(fj_geom):
                    ag_assigned = f"AG{ag_num:02d}"
                    break
            # Step 2: Fallback — nearest centroid distance (only if not inside any AG)
            if ag_assigned is None:
                min_dist = float('inf')
                ag_assigned = f"AG{ag_data[0][2]:02d}" if ag_data else 'AG01'
                for _, ag_feat, ag_num in ag_data:
                    d = ag_feat.geometry().centroid().distance(fj_geom)
                    if d < min_dist:
                        min_dist = d
                        ag_assigned = f"AG{ag_num:02d}"
            fj_data[i] = (fid, feat, ag_assigned)
        
        dj_data = [(f.id(), f) for f in dj_layer.getFeatures()]
        dc_data = [(f.id(), f) for f in dc_layer.getFeatures()]
        poles_data = [(f.id(), f) for f in poles_layer.getFeatures()] if poles_layer else []
        
        feedback.pushInfo(f"  FJs: {len(fj_data)}")
        feedback.pushInfo(f"  DJs: {len(dj_data)}")
        feedback.pushInfo(f"  DC segments: {len(dc_data)}")
        feedback.pushInfo(f"  Poles: {len(poles_data)}")
        
        # ===== PHASE 2: Run tracing engine =====
        feedback.setProgress(30)
        feedback.pushInfo("")
        feedback.pushInfo("[Phase 2] Tracing DC network and generating labels...")
        
        results = engine.run_full_labeling(
            ag_data, block_data, fj_data, dj_data, dc_data, 
            poles_data=None, debug_log=feedback.pushInfo
        )
        
        # Log results
        for ag_name, fj_result in results['ags'].items():
            feedback.pushInfo(f"  {fj_result['fj_name']} (AG: {ag_name}):")
            for chain in fj_result['block_chains']:
                names = [d['dj_name'] for d in chain['djs']]
                positions = [str(d['position']) for d in chain['djs']]
                ratios = [d['ratio'] for d in chain['djs']]
                feedback.pushInfo(f"    Block {chain['block_name']} ({chain['dj_count']} DJs):")
                feedback.pushInfo(f"      DJs: {', '.join(names)}")
                feedback.pushInfo(f"      Positions: {', '.join(positions)}")
                feedback.pushInfo(f"      Ratios: {', '.join(ratios)}")
        
        # ===== PHASE 2.5: Three-stage Pole Labeling =====
        # Stage 1: Label poles along FC cables (FC01, FC02, ...)
        # Stage 2: Label poles along DC cables (AG01 B001 DC001.1, DC001.2, ... B002 DC002.1, ...)
        # Stage 3: Nearest-line fallback for poles not directly on any cable
        # Skip poles already labeled in earlier stages (FC/DC may share poles)
        pole_labels = {}  # {pole_feat_id: (pole_name, projected_dist)}
        pole_stage1_count = 0
        pole_stage2_count = 0
        pole_stage3_count = 0
        
        if label_pole:
            if not label_dc:
                feedback.pushInfo("")
                feedback.pushInfo("[Phase 2.5] Labeling poles (3-stage)...")
                feedback.reportError("  ERROR: Cannot label poles without DC labels. Enable 'Label DC cables' or skip poles.")
            elif poles_layer and poles_data:
                feedback.pushInfo("")
                feedback.pushInfo("[Phase 2.5] Labeling poles (3-stage)...")
            
                def _find_poles_on_line(line_geom, pole_candidates, tol=2.0):
                    """Return list of (pole_feat_id, projected_dist) for poles on/near line."""
                    on_line = []
                    for p_fid, p_feat in pole_candidates:
                        p_geom = p_feat.geometry()
                        dist = line_geom.distance(p_geom)
                        if dist <= tol:
                            projected_dist = line_geom.lineLocatePoint(p_geom)
                            on_line.append((p_fid, projected_dist))
                    on_line.sort(key=lambda x: x[1])
                    return on_line
            
                # Collect all line geometries with their processing priority for fallback
                all_lines = []  # (line_geom, stage_priority, sort_key)
            
                # ---- STAGE 1: FC Poles ----
                if fc_data and fc_layer:
                    fc_sorted = sorted(fc_data, key=lambda x: x[2])
                    for fc_fid, fc_feat, fc_num in fc_sorted:
                        fc_geom = fc_feat.geometry()
                        all_lines.append((fc_geom, 1, fc_num))
                        poles_on_fc = _find_poles_on_line(fc_geom, poles_data, tolerance)
                        for p_fid, proj_dist in poles_on_fc:
                            if p_fid not in pole_labels:
                                pole_labels[p_fid] = (engine.generate_pole_name(), proj_dist, 1)
                                pole_stage1_count += 1
                    feedback.pushInfo(f"  Stage 1 (FC): {pole_stage1_count} poles on FC cables")
            
                # ---- STAGE 2: DC Poles ----
                # Trace from FJ outward: for each AG -> each block -> each DC cable in chain order
                # For each DC segment, sort poles from upstream (FJ-side) to downstream (DJ-side)
                def _ag_sort_key(ag_name):
                    try:
                        return int(re.search(r'\d+', ag_name).group())
                    except Exception:
                        return 0
            
                # Build DJ geometry lookup for node-based direction detection
                dj_geom_lookup = {}
                for d_fid, d_feat in dj_data:
                    dj_geom_lookup[d_fid] = d_feat.geometry()
                
                # Build FJ geometry lookup
                fj_geom_lookup = {}
                for f_fid, f_feat, f_ag in fj_data:
                    fj_geom_lookup[f_fid] = f_feat.geometry()
                
                # Build DJ name -> feat_id mapping from results (for node-based pole direction)
                dj_name_to_fid = {}
                for agn, fj_res in results['ags'].items():
                    for chain in fj_res['block_chains']:
                        for dj in chain['djs']:
                            dj_name_to_fid[dj['dj_name']] = dj['feat_id']
                
                dc_geom_lookup = {fid: feat.geometry() for fid, feat in dc_data}
            
                for ag_name in sorted(results['ags'].keys(), key=_ag_sort_key):
                    fj_result = results['ags'][ag_name]
                    for chain in fj_result['block_chains']:
                        for dc in chain['dc_cables']:
                            dc_fid = dc.get('dc_feat_id')
                            if dc_fid is None or dc_fid not in dc_geom_lookup:
                                continue
                            dc_geom = dc_geom_lookup[dc_fid]
                            all_lines.append((dc_geom, 2, len(all_lines)))
                        
                            # Find poles on this DC segment
                            poles_on_dc = _find_poles_on_line(dc_geom, poles_data, tolerance)
                        
                            if len(poles_on_dc) == 0:
                                continue
                        
                            # === NODE-BASED DIRECTION (independent of line drawing direction) ===
                            # Project FROM and TO node geometries onto the line.
                            # Pole order = distance from FROM node along the line.
                            from_node_name = dc.get('from_node', '')
                            to_node_name = dc.get('to_node', '')
                            
                            # Find the FROM node geometry (FJ or previous DJ)
                            from_geom = None
                            to_geom = None
                            
                            # Check if from_node is an FJ
                            for f_fid, f_feat, f_ag in fj_data:
                                fj_display = engine.generate_fj_name(f_ag)
                                if fj_display == from_node_name:
                                    from_geom = f_feat.geometry()
                                    break
                            
                            # Check if from_node is a DJ
                            if from_geom is None:
                                d_fid = dj_name_to_fid.get(from_node_name)
                                if d_fid and d_fid in dj_geom_lookup:
                                    from_geom = dj_geom_lookup[d_fid]
                            
                            # Find the TO node geometry (DJ)
                            d_fid = dj_name_to_fid.get(to_node_name)
                            if d_fid and d_fid in dj_geom_lookup:
                                to_geom = dj_geom_lookup[d_fid]
                            
                            # Determine direction: project both nodes onto the line
                            if from_geom is not None and to_geom is not None:
                                from_proj = dc_geom.lineLocatePoint(from_geom)
                                to_proj = dc_geom.lineLocatePoint(to_geom)
                                # If FROM projects closer to the END of the line, the line is "reversed"
                                # We want ascending order = FROM → TO regardless of drawing direction
                                if from_proj > to_proj:
                                    # Line drawn TO→FROM: invert lineLocatePoint
                                    max_dist = dc_geom.length()
                                    poles_on_dc = [(p_fid, max_dist - proj_dist) 
                                                   for p_fid, proj_dist in poles_on_dc]
                                    poles_on_dc.sort(key=lambda x: x[1])
                                # else: from_proj <= to_proj, line drawn FROM→TO: use as-is
                            # If we can't determine direction, use lineLocatePoint as-is
                        
                            for p_fid, _ in poles_on_dc:
                                if p_fid not in pole_labels:
                                    pole_labels[p_fid] = (engine.generate_pole_name(), 0, 2)
                                    pole_stage2_count += 1
                feedback.pushInfo(f"  Stage 2 (DC): {pole_stage2_count} poles on DC cables (always FJ->DJ direction)")
            
                # ---- STAGE 3: Nearest-line fallback ----
                unlabeled = [(pf, pf) for pf, _ in poles_data if pf not in pole_labels]
                if unlabeled:
                    feedback.pushInfo(f"  Stage 3 (Fallback): {len(unlabeled)} poles not on any cable within {tolerance}m...")
                    for p_fid, _ in unlabeled:
                        p_geom = None
                        for pf, pft in poles_data:
                            if pf == p_fid:
                                p_geom = pft.geometry()
                                break
                        if p_geom is None:
                            continue
                        best_dist = float('inf')
                        best_geom = None
                        best_priority = 999
                        for line_geom, priority, _ in all_lines:
                            d = line_geom.distance(p_geom)
                            if d < best_dist or (abs(d - best_dist) < 0.001 and priority < best_priority):
                                best_dist = d
                                best_geom = line_geom
                                best_priority = priority
                        if best_geom is not None:
                            proj_dist = best_geom.lineLocatePoint(p_geom)
                            pole_labels[p_fid] = (engine.generate_pole_name(), proj_dist, 3)
                            pole_stage3_count += 1
                    feedback.pushInfo(f"  Stage 3 (Fallback): {pole_stage3_count} assigned to nearest line")
            
                feedback.pushInfo(f"  Total: {len(pole_labels)} / {len(poles_data)} poles labeled")
                if len(pole_labels) < len(poles_data):
                    feedback.pushInfo(f"  WARNING: {len(poles_data) - len(pole_labels)} poles unlabeled")
        else:
            feedback.pushInfo("")
            feedback.pushInfo("[Phase 2.5] Pole labeling SKIPPED")
        
        # ===== PHASE 3: Create ALL fields first =====
        feedback.setProgress(60)
        feedback.pushInfo("")
        feedback.pushInfo("[Phase 3] Creating fields...")
        
        # AG: name
        ag_name_idx = -1
        if label_ag:
            ag_name_idx = self._ensure_field(ag_layer, target_field)
            feedback.pushInfo(f"  AG '{target_field}': {'OK' if ag_name_idx >= 0 else 'FAILED'}")
        
        # Block: name, ag_parent
        block_name_idx = -1
        block_ag_idx = -1
        if label_block:
            block_name_idx = self._ensure_field(blocks_layer, target_field)
            block_ag_idx = self._ensure_field(blocks_layer, 'ag_parent')
            feedback.pushInfo(f"  Block '{target_field}': {'OK' if block_name_idx >= 0 else 'FAILED'}")
            feedback.pushInfo(f"  Block 'ag_parent': {'OK' if block_ag_idx >= 0 else 'FAILED'}")
        
        # FJ: name
        fj_name_idx = -1
        if label_fj:
            fj_name_idx = self._ensure_field(fj_layer, target_field)
            feedback.pushInfo(f"  FJ '{target_field}': {'OK' if fj_name_idx >= 0 else 'FAILED'}")
        
        # DJ: name, splitter, position, dc_name
        dj_name_idx = -1
        dj_split_idx = -1
        dj_pos_idx = -1
        dj_dc_idx = -1
        if label_dj:
            dj_name_idx = self._ensure_field(dj_layer, target_field)
            dj_split_idx = self._ensure_field(dj_layer, 'splitter', QVariant.String, 10)
            dj_pos_idx = self._ensure_field(dj_layer, 'position', QVariant.Int)
            dj_dc_idx = self._ensure_field(dj_layer, 'dc_name', QVariant.String, 200)
            feedback.pushInfo(f"  DJ '{target_field}': {'OK' if dj_name_idx >= 0 else 'FAILED'}")
            feedback.pushInfo(f"  DJ 'splitter': {'OK' if dj_split_idx >= 0 else 'FAILED'}")
            feedback.pushInfo(f"  DJ 'position': {'OK' if dj_pos_idx >= 0 else 'FAILED'}")
            feedback.pushInfo(f"  DJ 'dc_name': {'OK' if dj_dc_idx >= 0 else 'FAILED'}")
        
        # DC: name, from_node, to_node
        dc_name_idx = -1
        dc_from_idx = -1
        dc_to_idx = -1
        if label_dc:
            dc_name_idx = self._ensure_field(dc_layer, target_field)
            dc_from_idx = self._ensure_field(dc_layer, 'from_node', QVariant.String, 100)
            dc_to_idx = self._ensure_field(dc_layer, 'to_node', QVariant.String, 100)
            feedback.pushInfo(f"  DC '{target_field}': {'OK' if dc_name_idx >= 0 else 'FAILED'}")
            feedback.pushInfo(f"  DC 'from_node': {'OK' if dc_from_idx >= 0 else 'FAILED'}")
            feedback.pushInfo(f"  DC 'to_node': {'OK' if dc_to_idx >= 0 else 'FAILED'}")
        
        # Pole: name
        pole_name_idx = -1
        if label_pole and poles_layer:
            pole_name_idx = self._ensure_field(poles_layer, target_field)
            feedback.pushInfo(f"  Pole '{target_field}': {'OK' if pole_name_idx >= 0 else 'FAILED'}")
        
        # FC: name
        fc_name_idx = -1
        if label_fc and fc_layer:
            fc_name_idx = self._ensure_field(fc_layer, target_field)
            feedback.pushInfo(f"  FC '{target_field}': {'OK' if fc_name_idx >= 0 else 'FAILED'}")
        
        # ===== VALIDATION GATE =====
        # Check if a pre-flight validation report exists with critical errors.
        # If so, STOP and tell the user to fix issues first.
        validation_report = os.path.join(output_folder, 'preflight_validation_report.txt')
        if os.path.exists(validation_report):
            try:
                with open(validation_report, 'r', encoding='utf-8') as f:
                    content = f.read()
                if 'CRITICAL issues' in content or 'CRITICAL' in content:
                    feedback.reportError("=" * 60)
                    feedback.reportError("VALIDATION GATE BLOCKED")
                    feedback.reportError("=" * 60)
                    feedback.reportError("A pre-flight validation report with CRITICAL issues was found.")
                    feedback.reportError("File: " + validation_report)
                    feedback.reportError("")
                    feedback.reportError("Please either:")
                    feedback.reportError("  1. Fix the critical issues and delete the validation report, OR")
                    feedback.reportError("  2. Run the FTTH Validator again to re-check and clear the report.")
                    feedback.reportError("")
                    feedback.reportError("The labeler will NOT proceed to prevent data corruption.")
                    feedback.reportError("=" * 60)
                    return {self.PARAM_OUTPUT: output_folder}
                elif 'warnings' in content.lower() and 'ALL CHECKS PASSED' not in content:
                    feedback.pushInfo("[Validation Gate] Non-critical warnings found in previous validation.")
                    feedback.pushInfo("  File: " + validation_report)
                    feedback.pushInfo("  Proceeding with caution...")
                    feedback.pushInfo("")
            except Exception:
                pass

        # ===== BUILD LABEL LOOKUPS (must be before Phase 4) =====
        ag_labels = {fid: engine.generate_ag_name(num) for fid, _, num in ag_data}
        block_labels = {fid: engine.generate_block_name(bnum) for fid, _, bnum, _, _ in block_data}
        block_ag_map = {fid: f"{engine._az()}_{ag_short}" for fid, _, _, _, ag_short in block_data}

        dj_labels = {}
        dc_labels = {}
        dc_none_count = 0
        dc_dup_count = 0
        seen_dc_fids = set()
        for ag_name, fj_result in results['ags'].items():
            for chain in fj_result['block_chains']:
                for dj in chain['djs']:
                    dj_labels[dj['feat_id']] = dj
                for dc in chain['dc_cables']:
                    dc_fid = dc.get('dc_feat_id')
                    if dc_fid is not None:
                        if dc_fid in seen_dc_fids:
                            dc_dup_count += 1
                        else:
                            seen_dc_fids.add(dc_fid)
                            dc_labels[dc_fid] = dc
                    else:
                        dc_none_count += 1
        if dc_none_count > 0:
            feedback.pushInfo(f"  WARNING: {dc_none_count} DC cable(s) have no feature ID")
        if dc_dup_count > 0:
            feedback.pushInfo(f"  WARNING: {dc_dup_count} DC cable(s) have duplicate feature IDs")

        fj_labels = {}
        for fid, _, ag_name in fj_data:
            for agn, fj_res in results['ags'].items():
                if agn == ag_name:
                    fj_labels[fid] = fj_res['fj_name']
                    break

        # ===== PHASE 4: Write data (dataProvider batch — NO edit buffer) =====
        feedback.setProgress(75)
        feedback.pushInfo("")
        feedback.pushInfo("[Phase 4] Writing labels to QGIS layers (batch mode)...")

        def _batch_write_attrs(layer, attr_map, layer_name):
            """Write attributes via dataProvider — bypasses edit buffer completely.
            attr_map: {feat_id: {field_idx: value, ...}, ...}
            Returns (written_count, error_msg or None)."""
            if not layer or not layer.isValid():
                return 0, "layer invalid"
            if not attr_map:
                return 0, None
            try:
                # Roll back any stale edit session first
                if layer.isEditable():
                    try:
                        layer.rollBack()
                    except Exception:
                        pass
                    time.sleep(0.2)
                ok = layer.dataProvider().changeAttributeValues(attr_map)
                if ok:
                    layer.triggerRepaint()
                    return len(attr_map), None
                else:
                    return 0, "dataProvider.changeAttributeValues returned False"
            except Exception as e:
                return 0, str(e)[:200]

        # Write AG
        if label_ag and ag_name_idx >= 0 and ag_labels:
            updates = {}
            for feat in ag_layer.getFeatures():
                fid = feat.id()
                if fid in ag_labels:
                    updates[fid] = {ag_name_idx: ag_labels[fid]}
            written, err = _batch_write_attrs(ag_layer, updates, "AG")
            if err:
                feedback.pushInfo(f"  AG layer: ERROR — {err}")
            else:
                feedback.pushInfo(f"  AG layer: {written} written")

        # Write Block
        if label_block and block_name_idx >= 0 and block_labels:
            updates = {}
            for feat in blocks_layer.getFeatures():
                fid = feat.id()
                if fid in block_labels:
                    up = {block_name_idx: block_labels[fid]}
                    if fid in block_ag_map and block_ag_idx >= 0:
                        up[block_ag_idx] = block_ag_map[fid]
                    updates[fid] = up
            written, err = _batch_write_attrs(blocks_layer, updates, "Block")
            if err:
                feedback.pushInfo(f"  Block layer: ERROR — {err}")
            else:
                feedback.pushInfo(f"  Block layer: {written} written")

        # Write FJ
        if label_fj and fj_name_idx >= 0 and fj_labels:
            updates = {}
            for feat in fj_layer.getFeatures():
                fid = feat.id()
                if fid in fj_labels:
                    updates[fid] = {fj_name_idx: fj_labels[fid]}
            written, err = _batch_write_attrs(fj_layer, updates, "FJ")
            if err:
                feedback.pushInfo(f"  FJ layer: ERROR — {err}")
            else:
                feedback.pushInfo(f"  FJ layer: {written} written")

        # Write DJ (multi-field: name, splitter, position, dc_name)
        if label_dj and dj_name_idx >= 0 and dj_labels:
            updates = {}
            for feat in dj_layer.getFeatures():
                if feedback.isCanceled():
                    return {self.PARAM_OUTPUT: output_folder}
                fid = feat.id()
                if fid in dj_labels:
                    info = dj_labels[fid]
                    up = {dj_name_idx: info["dj_name"]}
                    if dj_split_idx >= 0:
                        up[dj_split_idx] = info["ratio"]
                    if dj_pos_idx >= 0:
                        up[dj_pos_idx] = info["position"]
                    if dj_dc_idx >= 0:
                        up[dj_dc_idx] = info["dc_name"]
                    updates[fid] = up
            written, err = _batch_write_attrs(dj_layer, updates, "DJ")
            if err:
                feedback.pushInfo(f"  DJ layer: ERROR — {err}")
            else:
                feedback.pushInfo(f"  DJ layer: {written} written")

        # Write DC (multi-field: name, from_node, to_node)
        if label_dc and dc_name_idx >= 0 and dc_labels:
            updates = {}
            for feat in dc_layer.getFeatures():
                if feedback.isCanceled():
                    return {self.PARAM_OUTPUT: output_folder}
                fid = feat.id()
                if fid in dc_labels:
                    info = dc_labels[fid]
                    up = {dc_name_idx: info["dc_name"]}
                    if dc_from_idx >= 0:
                        up[dc_from_idx] = info["from_node"]
                    if dc_to_idx >= 0:
                        up[dc_to_idx] = info["to_node"]
                    updates[fid] = up
            written, err = _batch_write_attrs(dc_layer, updates, "DC")
            if err:
                feedback.pushInfo(f"  DC layer: ERROR — {err}")
            else:
                feedback.pushInfo(f"  DC layer: {written} written")

        # Write Poles (single-field: name)
        if label_pole and pole_name_idx >= 0 and pole_labels:
            updates = {}
            for feat in poles_layer.getFeatures():
                if feedback.isCanceled():
                    return {self.PARAM_OUTPUT: output_folder}
                fid = feat.id()
                if fid in pole_labels:
                    updates[fid] = {pole_name_idx: pole_labels[fid][0]}
            written, err = _batch_write_attrs(poles_layer, updates, "Pole")
            if err:
                feedback.pushInfo(f"  Pole layer: ERROR — {err}")
            else:
                feedback.pushInfo(f"  Pole layer: {written} written (FC:{pole_stage1_count} DC:{pole_stage2_count} Fallback:{pole_stage3_count})")

        # Write FC
        if label_fc and fc_name_idx >= 0 and fc_labels_map:
            updates = {}
            for feat in fc_layer.getFeatures():
                fid = feat.id()
                if fid in fc_labels_map:
                    updates[fid] = {fc_name_idx: fc_labels_map[fid]["fc_name"]}
            written, err = _batch_write_attrs(fc_layer, updates, "FC")
            if err:
                feedback.pushInfo(f"  FC layer: ERROR — {err}")
            else:
                feedback.pushInfo(f"  FC layer: {written} written")

        # ===== PHASE 5: Write output files =====
        feedback.setProgress(90)
        feedback.pushInfo("")
        feedback.pushInfo("[Phase 5] Writing output files...")
        
        os.makedirs(output_folder, exist_ok=True)
        
        try:
            json_path = os.path.join(output_folder, 'labeled_network.json')
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2, ensure_ascii=False, default=str)
            feedback.pushInfo(f"  [OK] labeled_network.json")
        except Exception as e:
            feedback.pushInfo(f"  [ERROR] JSON: {e}")
        
        try:
            csv_path = os.path.join(output_folder, 'naming_summary.csv')
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['FJ', 'Block', 'DJ_Name', 'Position', 'Ratio', 'DC_Name', 'From', 'To'])
                for ag_name, fj_result in results['ags'].items():
                    for chain in fj_result['block_chains']:
                        for i, dj in enumerate(chain['djs']):
                            dc = chain['dc_cables'][i] if i < len(chain['dc_cables']) else {}
                            writer.writerow([fj_result['fj_name'], chain['block_name'],
                                           dj['dj_name'], dj['position'], dj['ratio'],
                                           dc.get('dc_name', ''),
                                           dc.get('from_node', ''), dc.get('to_node', '')])
            feedback.pushInfo(f"  [OK] naming_summary.csv")
        except Exception as e:
            feedback.pushInfo(f"  [ERROR] CSV: {e}")
        
        # FC summary CSV
        if fc_labels_map:
            try:
                fc_csv_path = os.path.join(output_folder, 'fc_summary.csv')
                with open(fc_csv_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(['FC_Name', 'FC_Num', 'Start_AG', 'End_AG', 
                                     'Route_Length_m', 'FJ_Count', 'Slack_m'])
                    for fid, info in fc_labels_map.items():
                        writer.writerow([info['fc_name'], info['fc_num'],
                                         info['ag_start'], info['ag_end'],
                                         f"{info['route_length']:.0f}",
                                         info['fj_count'], info['slack']])
                feedback.pushInfo(f"  [OK] fc_summary.csv")
            except Exception as e:
                feedback.pushInfo(f"  [ERROR] FC CSV: {e}")
        
        try:
            report_path = os.path.join(output_folder, 'validation_report.txt')
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write("FTTH Labeler V1.0 — Report\n")
                f.write("=" * 60 + "\n\n")
                f.write(f"Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Area: {area_code}, Zone: {zone}\n\n")
                f.write(f"AGs: {results['summary']['total_ags']}\n")
                f.write(f"Blocks: {results['summary']['total_blocks']}\n")
                f.write(f"DJs: {results['summary']['total_djs']}\n")
                f.write(f"DC Cables: {results['summary']['total_dc']}\n")
                if fc_labels_map:
                    f.write(f"FC Cables: {len(fc_labels_map)}\n")
                if results['summary']['total_poles'] > 0:
                    f.write(f"Poles: {results['summary']['total_poles']}\n")
            feedback.pushInfo(f"  [OK] validation_report.txt")
        except Exception as e:
            feedback.pushInfo(f"  [ERROR] Report: {e}")
        
        # ===== DONE =====
        feedback.setProgress(100)
        feedback.pushInfo("")
        feedback.pushInfo("=" * 60)
        feedback.pushInfo("LABELING COMPLETE")
        feedback.pushInfo("=" * 60)
        feedback.pushInfo(f"Output: {output_folder}")
        
        return {self.PARAM_OUTPUT: output_folder}


# =============================================================================
# FTTH PRE-FLIGHT VALIDATOR
# =============================================================================

class FTTHValidatorAlgorithm(QgsProcessingAlgorithm):
    """Pre-flight topology validator for FTTH network layers.
    
    Checks all requirements before running the FTTH Labeler:
    1. CRS consistency across all layers
    2. Required ID columns exist
    3. DC lines snap to DJ points
    4. DC lines snap to FJ points (origin)
    5. DJ points inside Block polygons
    6. Block polygons inside AG polygons
    7. FJ points inside AG polygons
    8. Poles near DC or FC lines
    9. FC lines intersect AG polygons
    10. No duplicate geometries
    11. DC network connectivity (no orphaned DJs)
    12. FJ count matches AG count
    """
    
    IN_AG = 'IN_AG'
    IN_BLOCKS = 'IN_BLOCKS'
    IN_FJ = 'IN_FJ'
    IN_DJ = 'IN_DJ'
    IN_DC = 'IN_DC'
    IN_POLES = 'IN_POLES'
    IN_FC = 'IN_FC'
    
    PARAM_AG_ID = 'PARAM_AG_ID'
    PARAM_BLOCK_ID = 'PARAM_BLOCK_ID'
    PARAM_TOLERANCE = 'PARAM_TOLERANCE'
    PARAM_OUTPUT = 'PARAM_OUTPUT'
    
    def name(self):
        return 'ftth_validator'
    def displayName(self):
        return 'FTTH Pre-Flight Validator'
    def group(self):
        return 'FTTH Tools'
    def groupId(self):
        return 'ftth_tools'
    def createInstance(self):
        return FTTHValidatorAlgorithm()
    
    def shortHelpString(self):
        return """
        <h3>FTTH Pre-Flight Validator</h3>
        <p>Validates topology and requirements before running the FTTH Labeler.</p>
        
        <h4>Checks Performed:</h4>
        <ol>
        <li><b>CRS Consistency</b> — all layers use the same CRS</li>
        <li><b>Required Columns</b> — AG and Block have ID columns</li>
        <li><b>DC→DJ Snap</b> — DC line endpoints snap to DJ points</li>
        <li><b>DC→FJ Snap</b> — first DC segment originates from FJ</li>
        <li><b>DJ in Block</b> — DJ points fall inside Block polygons</li>
        <li><b>Block in AG</b> — Block polygons fall inside AG polygons</li>
        <li><b>FJ in AG</b> — FJ points fall inside AG polygons</li>
        <li><b>Pole on Cable</b> — poles are near DC or FC lines</li>
        <li><b>FC→AG</b> — FC lines intersect AG polygons</li>
        <li><b>No Duplicates</b> — no duplicate geometries</li>
        <li><b>Network Connectivity</b> — DJs are reachable from FJs via DC</li>
        <li><b>FJ Count</b> — at least one FJ per AG</li>
        </ol>
        
        <p>Output: validation report with PASS/FAIL for each check + issue counts.</p>
        """
    
    def initAlgorithm(self, config=None):
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_AG, 'AG Polygons (MUST have ID column)',
            [QgsProcessing.TypeVectorPolygon]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_BLOCKS, 'Block Polygons (MUST have ID column)',
            [QgsProcessing.TypeVectorPolygon]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_FJ, 'Feeder Joint (FJ) Points',
            [QgsProcessing.TypeVectorPoint]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_DJ, 'Distribution Joint (DJ) Points',
            [QgsProcessing.TypeVectorPoint]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_DC, 'Distribution Cable (DC) Lines',
            [QgsProcessing.TypeVectorLine]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_POLES, 'Pole Points (optional)',
            [QgsProcessing.TypeVectorPoint], optional=True
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_FC, 'Feeder Cable (FC) Lines (optional)',
            [QgsProcessing.TypeVectorLine], optional=True
        ))
        self.addParameter(QgsProcessingParameterString(
            self.PARAM_AG_ID, 'AG ID Field Name', defaultValue='ID'
        ))
        self.addParameter(QgsProcessingParameterString(
            self.PARAM_BLOCK_ID, 'Block ID Field Name', defaultValue='ID'
        ))
        self.addParameter(QgsProcessingParameterNumber(
            self.PARAM_TOLERANCE, 'Snap Tolerance (meters)',
            type=QgsProcessingParameterNumber.Double,
            defaultValue=2.0, minValue=0.1, maxValue=50.0
        ))
        self.addParameter(QgsProcessingParameterFolderDestination(
            self.PARAM_OUTPUT, 'Output Folder'
        ))
    
    @staticmethod
    def _find_field(layer, field_names):
        all_names = [f.name() for f in layer.fields()]
        for fn in field_names:
            fn_lower = fn.lower()
            for actual_name in all_names:
                if actual_name.lower() == fn_lower:
                    return layer.fields().indexFromName(actual_name)
        return -1
    
    @staticmethod
    def _count_features(layer):
        return sum(1 for _ in layer.getFeatures())
    
    @staticmethod
    def _get_crs_authid(layer):
        return layer.crs().authid()
    
    @staticmethod
    def _get_line_endpoints(line_geom):
        """Return (start_point, end_point) as QgsPointXY."""
        if line_geom.isMultipart():
            parts = line_geom.asMultiPolyline()
            if parts:
                # Use first part's start and last part's end
                first_part = parts[0]
                last_part = parts[-1]
                if first_part and last_part:
                    return (QgsPointXY(first_part[0]), QgsPointXY(last_part[-1]))
        else:
            pts = line_geom.asPolyline()
            if pts and len(pts) >= 2:
                return (QgsPointXY(pts[0]), QgsPointXY(pts[-1]))
        return (None, None)
    
    @staticmethod
    def _point_distance_to_point(pt1, pt2):
        """Distance between two QgsPointXY."""
        return ((pt1.x() - pt2.x())**2 + (pt1.y() - pt2.y())**2)**0.5
    
    def processAlgorithm(self, parameters, context, feedback):
        tolerance = self.parameterAsDouble(parameters, self.PARAM_TOLERANCE, context)
        ag_id_field = self.parameterAsString(parameters, self.PARAM_AG_ID, context) or 'ID'
        block_id_field = self.parameterAsString(parameters, self.PARAM_BLOCK_ID, context) or 'ID'
        output_folder = self.parameterAsString(parameters, self.PARAM_OUTPUT, context)
        
        ag_layer = self.parameterAsVectorLayer(parameters, self.IN_AG, context)
        blocks_layer = self.parameterAsVectorLayer(parameters, self.IN_BLOCKS, context)
        fj_layer = self.parameterAsVectorLayer(parameters, self.IN_FJ, context)
        dj_layer = self.parameterAsVectorLayer(parameters, self.IN_DJ, context)
        dc_layer = self.parameterAsVectorLayer(parameters, self.IN_DC, context)
        poles_layer = self.parameterAsVectorLayer(parameters, self.IN_POLES, context)
        fc_layer = self.parameterAsVectorLayer(parameters, self.IN_FC, context)
        
        # Layer summary
        ag_count = self._count_features(ag_layer)
        block_count = self._count_features(blocks_layer)
        fj_count = self._count_features(fj_layer)
        dj_count = self._count_features(dj_layer)
        dc_count = self._count_features(dc_layer)
        pole_count = self._count_features(poles_layer) if poles_layer else 0
        fc_count = self._count_features(fc_layer) if fc_layer else 0
        
        feedback.pushInfo("=" * 60)
        feedback.pushInfo("FTTH Pre-Flight Validator")
        feedback.pushInfo("=" * 60)
        feedback.pushInfo(f"Snap Tolerance: {tolerance}m")
        feedback.pushInfo("")
        feedback.pushInfo("Layer Summary:")
        feedback.pushInfo(f"  AG:     {ag_count} features")
        feedback.pushInfo(f"  Block:  {block_count} features")
        feedback.pushInfo(f"  FJ:     {fj_count} features")
        feedback.pushInfo(f"  DJ:     {dj_count} features")
        feedback.pushInfo(f"  DC:     {dc_count} features")
        if poles_layer:
            feedback.pushInfo(f"  Pole:   {pole_count} features")
        if fc_layer:
            feedback.pushInfo(f"  FC:     {fc_count} features")
        feedback.pushInfo("")
        
        report_lines = []
        report_lines.append("FTTH Pre-Flight Validation Report")
        report_lines.append("=" * 60)
        report_lines.append(f"Snap Tolerance: {tolerance}m")
        report_lines.append("")
        report_lines.append("Layer Summary:")
        report_lines.append(f"  AG: {ag_count}, Block: {block_count}, FJ: {fj_count}, DJ: {dj_count}, DC: {dc_count}, Pole: {pole_count}, FC: {fc_count}")
        report_lines.append("")
        
        total_issues = 0
        critical_issues = 0
        
        # ===== CHECK 1: CRS Consistency =====
        feedback.setProgress(5)
        feedback.pushInfo("[Check 1] CRS Consistency...")
        report_lines.append("[1] CRS CONSISTENCY")
        
        layers_to_check = [
            ('AG', ag_layer), ('Block', blocks_layer), ('FJ', fj_layer),
            ('DJ', dj_layer), ('DC', dc_layer)
        ]
        if poles_layer:
            layers_to_check.append(('Pole', poles_layer))
        if fc_layer:
            layers_to_check.append(('FC', fc_layer))
        
        crses = {}
        for name, lyr in layers_to_check:
            crses[name] = self._get_crs_authid(lyr)
        
        unique_crs = set(crses.values())
        if len(unique_crs) == 1:
            feedback.pushInfo(f"  PASS: All layers use {list(unique_crs)[0]}")
            report_lines.append(f"  PASS: All layers use {list(unique_crs)[0]}")
        else:
            feedback.reportError(f"  FAIL: Multiple CRS found!")
            report_lines.append(f"  FAIL: Multiple CRS found!")
            for name, crs in crses.items():
                line = f"    {name}: {crs}"
                feedback.pushInfo(line)
                report_lines.append(line)
            total_issues += 1
            critical_issues += 1
        
        # ===== CHECK 2: Required Columns =====
        feedback.setProgress(10)
        feedback.pushInfo("")
        feedback.pushInfo("[Check 2] Required ID Columns...")
        report_lines.append("")
        report_lines.append("[2] REQUIRED ID COLUMNS")
        
        ag_id_idx = self._find_field(ag_layer, [ag_id_field, 'id', 'ID', 'Id'])
        block_id_idx = self._find_field(blocks_layer, [block_id_field, 'id', 'ID', 'Id'])
        
        if ag_id_idx >= 0:
            feedback.pushInfo(f"  PASS: AG layer has '{ag_id_field}' column")
            report_lines.append(f"  PASS: AG layer has '{ag_id_field}' column")
        else:
            feedback.reportError(f"  FAIL: AG layer missing '{ag_id_field}' column!")
            report_lines.append(f"  FAIL: AG layer missing '{ag_id_field}' column!")
            total_issues += 1
            critical_issues += 1
        
        if block_id_idx >= 0:
            feedback.pushInfo(f"  PASS: Block layer has '{block_id_field}' column")
            report_lines.append(f"  PASS: Block layer has '{block_id_field}' column")
        else:
            feedback.reportError(f"  FAIL: Block layer missing '{block_id_field}' column!")
            report_lines.append(f"  FAIL: Block layer missing '{block_id_field}' column!")
            total_issues += 1
            critical_issues += 1
        
        # ===== CHECK 3: DC→DJ Snap =====
        feedback.setProgress(20)
        feedback.pushInfo("")
        feedback.pushInfo("[Check 3] DC lines snap to DJ points...")
        report_lines.append("")
        report_lines.append("[3] DC LINES SNAP TO DJ POINTS")
        
        # Build DJ point index
        dj_points = []
        for feat in dj_layer.getFeatures():
            dj_points.append((feat.id(), feat.geometry().asPoint()))
        
        dc_no_dj = 0
        dc_dj_snap = 0
        for feat in dc_layer.getFeatures():
            start_pt, end_pt = self._get_line_endpoints(feat.geometry())
            if start_pt is None:
                continue
            # Check both endpoints
            start_has_dj = any(self._point_distance_to_point(start_pt, dj_pt) <= tolerance
                             for _, dj_pt in dj_points)
            end_has_dj = any(self._point_distance_to_point(end_pt, dj_pt) <= tolerance
                           for _, dj_pt in dj_points)
            if start_has_dj or end_has_dj:
                dc_dj_snap += 1
            else:
                dc_no_dj += 1
        
        total_dc = dc_dj_snap + dc_no_dj
        if dc_no_dj == 0:
            feedback.pushInfo(f"  PASS: All {total_dc} DC segments snap to DJ")
            report_lines.append(f"  PASS: All {total_dc} DC segments snap to DJ")
        else:
            feedback.reportError(f"  FAIL: {dc_no_dj}/{total_dc} DC segments do NOT snap to any DJ!")
            report_lines.append(f"  FAIL: {dc_no_dj}/{total_dc} DC segments do NOT snap to any DJ!")
            total_issues += dc_no_dj
            critical_issues += 1
        
        # ===== CHECK 4: Each block has at least one DJ with DC→FJ connection =====
        feedback.setProgress(30)
        feedback.pushInfo("")
        feedback.pushInfo("[Check 4] Each block's first DJ connects to FJ...")
        report_lines.append("")
        report_lines.append("[4] BLOCK FIRST DJ -> FJ CONNECTION")
        
        # Build FJ point index
        fj_points = []
        for feat in fj_layer.getFeatures():
            fj_points.append((feat.id(), feat.geometry().asPoint()))
        
        # Build block polygon index
        block_geoms = [(feat.id(), feat.geometry()) for feat in blocks_layer.getFeatures()]
        
        # For each block: find DJs inside it, check if any has a DC that touches an FJ
        block_ok = 0
        block_fail = 0
        for b_fid, b_geom in block_geoms:
            # Find all DJs inside this block
            djs_in_block = []
            for feat in dj_layer.getFeatures():
                if b_geom.contains(feat.geometry()):
                    djs_in_block.append((feat.id(), feat.geometry().asPoint()))
            
            # Check if any DJ in this block has a DC that also touches an FJ
            has_fj_connection = False
            for dj_fid, dj_pt in djs_in_block:
                for dc_feat in dc_layer.getFeatures():
                    s_pt, e_pt = self._get_line_endpoints(dc_feat.geometry())
                    if s_pt is None:
                        continue
                    # DC must touch this DJ
                    touches_dj = (self._point_distance_to_point(s_pt, dj_pt) <= tolerance or
                                 self._point_distance_to_point(e_pt, dj_pt) <= tolerance)
                    if not touches_dj:
                        continue
                    # DC must also touch an FJ
                    for _, fj_pt in fj_points:
                        touches_fj = (self._point_distance_to_point(s_pt, fj_pt) <= tolerance or
                                     self._point_distance_to_point(e_pt, fj_pt) <= tolerance)
                        if touches_fj:
                            has_fj_connection = True
                            break
                    if has_fj_connection:
                        break
                if has_fj_connection:
                    break
            
            if has_fj_connection:
                block_ok += 1
            else:
                block_fail += 1
        
        total_block_check = block_ok + block_fail
        if block_fail == 0:
            feedback.pushInfo(f"  PASS: All {total_block_check} blocks have DJ->FJ connection")
            report_lines.append(f"  PASS: All {total_block_check} blocks have DJ->FJ connection")
        else:
            feedback.reportError(f"  FAIL: {block_fail}/{total_block_check} blocks have NO DJ->FJ connection!")
            report_lines.append(f"  FAIL: {block_fail}/{total_block_check} blocks have NO DJ->FJ connection!")
            total_issues += block_fail
            critical_issues += 1
        
        # ===== CHECK 5: DJ inside Block =====
        feedback.setProgress(40)
        feedback.pushInfo("")
        feedback.pushInfo("[Check 5] DJ points inside Block polygons...")
        report_lines.append("")
        report_lines.append("[5] DJ POINTS INSIDE BLOCK POLYGONS")
        
        # Build block polygon index
        block_geoms = [(feat.id(), feat.geometry()) for feat in blocks_layer.getFeatures()]
        
        dj_outside = 0
        dj_inside = 0
        for feat in dj_layer.getFeatures():
            dj_geom = feat.geometry()
            in_any_block = any(block_geom.contains(dj_geom) for _, block_geom in block_geoms)
            if in_any_block:
                dj_inside += 1
            else:
                dj_outside += 1
        
        total_dj = dj_inside + dj_outside
        if dj_outside == 0:
            feedback.pushInfo(f"  PASS: All {total_dj} DJs are inside blocks")
            report_lines.append(f"  PASS: All {total_dj} DJs are inside blocks")
        else:
            feedback.reportError(f"  FAIL: {dj_outside}/{total_dj} DJs are OUTSIDE all blocks!")
            report_lines.append(f"  FAIL: {dj_outside}/{total_dj} DJs are OUTSIDE all blocks!")
            total_issues += dj_outside
            critical_issues += 1
        
        # ===== CHECK 6: Block inside AG =====
        feedback.setProgress(50)
        feedback.pushInfo("")
        feedback.pushInfo("[Check 6] Block polygons inside AG polygons...")
        report_lines.append("")
        report_lines.append("[6] BLOCK POLYGONS INSIDE AG POLYGONS")
        
        ag_geoms = [(feat.id(), feat.geometry()) for feat in ag_layer.getFeatures()]
        
        block_outside = 0
        block_inside = 0
        for feat in blocks_layer.getFeatures():
            block_geom = feat.geometry()
            centroid = block_geom.centroid()
            in_any_ag = any(ag_geom.contains(centroid) for _, ag_geom in ag_geoms)
            if in_any_ag:
                block_inside += 1
            else:
                block_outside += 1
        
        total_block = block_inside + block_outside
        if block_outside == 0:
            feedback.pushInfo(f"  PASS: All {total_block} blocks are inside AGs")
            report_lines.append(f"  PASS: All {total_block} blocks are inside AGs")
        else:
            feedback.reportError(f"  FAIL: {block_outside}/{total_block} blocks are OUTSIDE all AGs!")
            report_lines.append(f"  FAIL: {block_outside}/{total_block} blocks are OUTSIDE all AGs!")
            total_issues += block_outside
            critical_issues += 1
        
        # ===== CHECK 7: FJ inside AG =====
        feedback.setProgress(55)
        feedback.pushInfo("")
        feedback.pushInfo("[Check 7] FJ points inside AG polygons...")
        report_lines.append("")
        report_lines.append("[7] FJ POINTS INSIDE AG POLYGONS")
        
        fj_outside = 0
        fj_inside = 0
        for feat in fj_layer.getFeatures():
            fj_geom = feat.geometry()
            in_any_ag = any(ag_geom.contains(fj_geom) for _, ag_geom in ag_geoms)
            if in_any_ag:
                fj_inside += 1
            else:
                fj_outside += 1
        
        if fj_outside == 0:
            feedback.pushInfo(f"  PASS: All {fj_inside} FJs are inside AGs")
            report_lines.append(f"  PASS: All {fj_inside} FJs are inside AGs")
        else:
            feedback.reportError(f"  FAIL: {fj_outside}/{fj_inside+fj_outside} FJs are outside AGs!")
            report_lines.append(f"  FAIL: {fj_outside}/{fj_inside+fj_outside} FJs are outside AGs!")
            total_issues += fj_outside
        
        # ===== CHECK 8: Poles on Cables =====
        feedback.setProgress(60)
        isolated_pole_ids = []  # collect for error layer
        if poles_layer:
            feedback.pushInfo("")
            feedback.pushInfo("[Check 8] Poles near DC or FC lines...")
            report_lines.append("")
            report_lines.append("[8] POLES NEAR CABLES (DC or FC)")
            
            # Build line geometry list
            line_geoms = [feat.geometry() for feat in dc_layer.getFeatures()]
            if fc_layer:
                line_geoms.extend([feat.geometry() for feat in fc_layer.getFeatures()])
            
            pole_isolated = 0
            pole_on_cable = 0
            for feat in poles_layer.getFeatures():
                pole_geom = feat.geometry()
                near_any = any(line_geom.distance(pole_geom) <= tolerance for line_geom in line_geoms)
                if near_any:
                    pole_on_cable += 1
                else:
                    pole_isolated += 1
                    isolated_pole_ids.append(feat.id())
            
            total_pole = pole_on_cable + pole_isolated
            if pole_isolated == 0:
                feedback.pushInfo(f"  PASS: All {total_pole} poles are on cables")
                report_lines.append(f"  PASS: All {total_pole} poles are on cables")
            else:
                feedback.pushInfo(f"  WARN: {pole_isolated}/{total_pole} poles are NOT near any cable")
                report_lines.append(f"  WARN: {pole_isolated}/{total_pole} poles are NOT near any cable")
                report_lines.append(f"  Isolated pole Feature IDs: {isolated_pole_ids}")
                report_lines.append(f"  (These will use Stage 3 fallback — nearest line assignment)")
                total_issues += pole_isolated
                
                # Create error layer for isolated poles
                try:
                    crs_str = poles_layer.crs().authid()
                    err_layer = QgsVectorLayer(f"Point?crs={crs_str}", "isolated_poles", "memory")
                    dp = err_layer.dataProvider()
                    dp.addAttributes([QgsField("pole_fid", QVariant.Int),
                                      QgsField("issue", QVariant.String, len=100)])
                    err_layer.updateFields()
                    
                    for feat in poles_layer.getFeatures():
                        if feat.id() in isolated_pole_ids:
                            err_feat = QgsFeature()
                            err_feat.setGeometry(feat.geometry())
                            err_feat.setAttributes([feat.id(), "Not near any DC or FC line"])
                            dp.addFeature(err_feat)
                    
                    err_path = os.path.join(output_folder, 'isolated_poles.gpkg')
                    QgsVectorFileWriter.writeAsVectorFormat(err_layer, err_path, 'UTF-8',
                        err_layer.crs(), 'GPKG')
                    feedback.pushInfo(f"  [OK] Error layer saved: {err_path}")
                    report_lines.append(f"  [OK] Error layer saved: {err_path}")
                except Exception as e:
                    feedback.pushInfo(f"  [WARN] Could not save error layer: {e}")
        else:
            feedback.pushInfo("[Check 8] No pole layer — skipped")
            report_lines.append("")
            report_lines.append("[8] POLES: No pole layer provided — skipped")
        
        # ===== CHECK 9: FC intersects AG =====
        feedback.setProgress(70)
        if fc_layer:
            feedback.pushInfo("")
            feedback.pushInfo("[Check 9] FC lines intersect AG polygons...")
            report_lines.append("")
            report_lines.append("[9] FC LINES INTERSECT AG POLYGONS")
            
            fc_no_ag = 0
            fc_has_ag = 0
            for feat in fc_layer.getFeatures():
                fc_geom = feat.geometry()
                intersects = any(fc_geom.intersects(ag_geom) for _, ag_geom in ag_geoms)
                if intersects:
                    fc_has_ag += 1
                else:
                    fc_no_ag += 1
            
            total_fc = fc_has_ag + fc_no_ag
            if fc_no_ag == 0:
                feedback.pushInfo(f"  PASS: All {total_fc} FC lines intersect AGs")
                report_lines.append(f"  PASS: All {total_fc} FC lines intersect AGs")
            else:
                feedback.reportError(f"  FAIL: {fc_no_ag}/{total_fc} FC lines do NOT intersect any AG!")
                report_lines.append(f"  FAIL: {fc_no_ag}/{total_fc} FC lines do NOT intersect any AG!")
                total_issues += fc_no_ag
                critical_issues += 1
        else:
            feedback.pushInfo("[Check 9] No FC layer — skipped")
            report_lines.append("")
            report_lines.append("[9] FC: No FC layer provided — skipped")
        
        # ===== CHECK 10: Duplicate Geometries =====
        feedback.setProgress(80)
        feedback.pushInfo("")
        feedback.pushInfo("[Check 10] Duplicate geometries...")
        report_lines.append("")
        report_lines.append("[10] DUPLICATE GEOMETRIES")
        
        dup_found = False
        for name, lyr in [('AG', ag_layer), ('Block', blocks_layer), ('FJ', fj_layer),
                          ('DJ', dj_layer), ('DC', dc_layer)]:
            geoms = []
            for feat in lyr.getFeatures():
                wkt = feat.geometry().asWkt(6)  # 6 decimal places
                geoms.append(wkt)
            seen = set()
            dups = 0
            for wkt in geoms:
                if wkt in seen:
                    dups += 1
                seen.add(wkt)
            if dups > 0:
                line = f"  WARN: {name} layer has {dups} duplicate geometries"
                feedback.pushInfo(line)
                report_lines.append(line)
                total_issues += dups
                dup_found = True
        
        if not dup_found:
            feedback.pushInfo(f"  PASS: No duplicate geometries found")
            report_lines.append(f"  PASS: No duplicate geometries found")
        
        # ===== CHECK 11: Network Connectivity =====
        feedback.setProgress(90)
        feedback.pushInfo("")
        feedback.pushInfo("[Check 11] DC network connectivity...")
        report_lines.append("")
        report_lines.append("[11] NETWORK CONNECTIVITY (DJ reachable from FJ)")
        
        # Build simple undirected graph
        from collections import defaultdict
        graph = defaultdict(set)
        # Add DJ nodes
        for feat in dj_layer.getFeatures():
            label = f"DJ_{feat.id()}"
        # This is a simplified check — full BFS is in trace_engine
        # Just check if each DJ has at least one DC segment connected
        dj_with_dc = 0
        dj_no_dc = 0
        for feat in dj_layer.getFeatures():
            dj_geom = feat.geometry()
            has_dc = any(dc_feat.geometry().distance(dj_geom) <= tolerance
                        for dc_feat in dc_layer.getFeatures())
            if has_dc:
                dj_with_dc += 1
            else:
                dj_no_dc += 1
        
        if dj_no_dc == 0:
            feedback.pushInfo(f"  PASS: All {dj_with_dc} DJs have DC connection")
            report_lines.append(f"  PASS: All {dj_with_dc} DJs have DC connection")
        else:
            feedback.reportError(f"  FAIL: {dj_no_dc} DJs have NO DC segments connected!")
            report_lines.append(f"  FAIL: {dj_no_dc} DJs have NO DC segments connected!")
            total_issues += dj_no_dc
            critical_issues += 1
        
        # ===== CHECK 12: FJ count == AG count (exact) =====
        feedback.pushInfo("")
        feedback.pushInfo("[Check 12] FJ count vs AG count...")
        report_lines.append("")
        report_lines.append("[12] FJ COUNT vs AG COUNT (must match exactly)")
        
        if fj_count == ag_count:
            feedback.pushInfo(f"  PASS: {fj_count} FJs == {ag_count} AGs (exact match)")
            report_lines.append(f"  PASS: {fj_count} FJs == {ag_count} AGs (exact match)")
        else:
            feedback.reportError(f"  FAIL: {fj_count} FJs != {ag_count} AGs! Must be equal (1 FJ per AG).")
            report_lines.append(f"  FAIL: {fj_count} FJs != {ag_count} AGs! Must be equal (1 FJ per AG).")
            total_issues += abs(ag_count - fj_count)
            critical_issues += 1
        
        # ===== SUMMARY =====
        feedback.setProgress(100)
        feedback.pushInfo("")
        feedback.pushInfo("=" * 60)
        feedback.pushInfo("VALIDATION SUMMARY")
        feedback.pushInfo("=" * 60)
        
        report_lines.append("")
        report_lines.append("=" * 60)
        report_lines.append("SUMMARY")
        report_lines.append("=" * 60)
        report_lines.append(f"Total issues found: {total_issues}")
        report_lines.append(f"Critical issues: {critical_issues}")
        
        if total_issues == 0:
            feedback.pushInfo("ALL CHECKS PASSED")
            feedback.pushInfo("You can safely run the FTTH Labeler.")
            report_lines.append("RESULT: ALL CHECKS PASSED — ready for FTTH Labeler")
        elif critical_issues == 0:
            feedback.pushInfo(f"{total_issues} warnings (non-critical)")
            feedback.pushInfo("You can run the FTTH Labeler, but review warnings.")
            report_lines.append(f"RESULT: {total_issues} warnings — labeler will run but review first")
        else:
            feedback.reportError(f"{critical_issues} CRITICAL issues found!")
            feedback.reportError("Fix critical issues before running FTTH Labeler.")
            report_lines.append(f"RESULT: {critical_issues} CRITICAL issues — fix before running labeler")
        
        # Write report
        os.makedirs(output_folder, exist_ok=True)
        try:
            report_path = os.path.join(output_folder, 'preflight_validation_report.txt')
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(report_lines))
            feedback.pushInfo(f"")
            feedback.pushInfo(f"Report saved: {report_path}")
        except Exception as e:
            feedback.pushInfo(f"  [ERROR] saving report: {e}")
        
        return {self.PARAM_OUTPUT: output_folder}



# =============================================================================
# FTTH SPLICING PLAN  v8.0  -- Excel output (direct .xlsx)
# =============================================================================
# CORRECTED APPROACH (v7.4):
#   Uses DJ Polygons as the PRIMARY method to match houses to DJs.
#   Each DJ Polygon defines the exact service area of one DJ.
#   Houses inside the polygon belong to that DJ -- exact, no ambiguity.
#
#   Input order:
#     1. DJ Polygons  (REQUIRED) -- one polygon per DJ, has block name attribute
#     2. DJ Points    (REQUIRED) -- labeled by Labeler (name, splitter, position, dc_name)
#     3. House Points (REQUIRED) -- premise/HP data with name field
#     4. Pole Points  (REQUIRED) -- labeled by Labeler with name column
#     5. Drop Cables  (OPTIONAL) -- fallback if DJ Polygons don't cover all houses
#     6. DC Lines     (OPTIONAL) -- for upstream FJ tracing
#     7. FJ Points    (OPTIONAL) -- for FJ pole lookup
#
#   OUTPUT: Formatted Excel (.xlsx) splicing diagram with merged headers,
#     styling, per-DJ cable routing (Feeder/Link, DC label/number/size,
#     splitter, DJ number, pole), and premise data (house names).
# =============================================================================

class FTTHSplicingPlanAlgorithm(QgsProcessingAlgorithm):
    """Generate splicing plan Excel (.xlsx) -- uses DJ Polygons as the PRIMARY
    method to match houses to DJs (point-in-polygon).  Produces ONE .xlsx
    file with a formatted splicing diagram worksheet.

    INPUT (all labeled by FTTH Labeler first):
      - DJ Polygons    : one polygon per DJ, carries block name attribute
      - DJ Points      : labeled with 'name', 'splitter', 'position', 'dc_name'
      - House Points   : premises / HP data with name field
      - Pole Points    : labeled with 'name' column
      - Drop Cable Lines (optional) : short lines from each house to its DJ
      - DC Lines         (optional) : for upstream FJ tracing
      - FJ Points        (optional) : for FJ pole lookup

    OUTPUT (one XLSX file):
      A formatted Excel splicing diagram with merged headers, styling,
      and per-DJ cable routing information including premise data.
    """

    IN_DJ_POLY = 'IN_DJ_POLY'  # DJ Polygons (REQUIRED)
    IN_DJ      = 'IN_DJ'       # DJ Points (REQUIRED)
    IN_HOUSES  = 'IN_HOUSES'   # House Points (REQUIRED)
    IN_POLES   = 'IN_POLES'    # Pole Points (REQUIRED)
    IN_DROPS   = 'IN_DROPS'    # Drop Cable Lines (OPTIONAL -- fallback)
    IN_DC      = 'IN_DC'       # DC Lines (OPTIONAL -- for FJ tracing)
    IN_FJ      = 'IN_FJ'       # FJ Points (OPTIONAL -- for FJ pole)

    PARAM_SNAP   = 'PARAM_SNAP'
    PARAM_OUTPUT = 'PARAM_OUTPUT'

    # ------------------------------------------------------------------
    # QGIS boilerplate
    # ------------------------------------------------------------------
    def name(self):
        return 'ftth_splicing_plan'

    def displayName(self):
        return 'FTTH Splicing Plan (Excel output)'

    def group(self):
        return 'FTTH Tools'

    def groupId(self):
        return 'ftth_tools'

    def createInstance(self):
        return FTTHSplicingPlanAlgorithm()

    def shortHelpString(self):
        return """
<h3>FTTH Splicing Plan (Excel output) v8.0</h3>
<p>Uses <b>DJ Polygons</b> as the PRIMARY method to match houses to DJs.
Each DJ Polygon defines the exact service area of one DJ. Houses inside
a polygon belong to that DJ -- exact, no ambiguity. ALL DJs appear in
the output (including those with 0 houses), ensuring 8 rows per DJ in
splicing diagrams.</p>
<p>Directly outputs a formatted <b>.xlsx</b> Excel file — no separate
script needed. openpyxl is bundled with QGIS.</p>

<h4>Required Inputs:</h4>
<ul>
<li><b>DJ Polygons</b> -- one polygon per DJ, must have block name attribute (auto-detected: block, Block, BLOCK, b_name, id, ID)</li>
<li><b>DJ Points</b> -- labeled DJ (run Labeler first; needs name, splitter, position, dc_name)</li>
<li><b>House Points</b> -- premises/HP (name field auto-detected)</li>
<li><b>Pole Points</b> -- labeled poles (run Labeler first)</li>
</ul>

<h4>Optional Inputs:</h4>
<ul>
<li><b>Drop Cable Lines</b> -- short lines from each house to its DJ. Used as FALLBACK for houses not covered by DJ Polygons.</li>
<li><b>DC Lines</b> -- labeled DC lines with from_node/to_node (traces upstream to find serving FJ)</li>
<li><b>FJ Points</b> -- labeled FJ points (looked up to find FJ pole name)</li>
</ul>

<h4>Output (one XLSX file):</h4>
<p>A formatted Excel splicing diagram with merged headers, styling,
per-DJ cable routing, and premise data.</p>
<p><b>fj_name</b> = serving FJ name (traced upstream via DC lines, requires optional DC + FJ layers)<br/>
<b>fj_pole</b> = pole where the FJ is mounted (requires optional DC + FJ layers)<br/>
<b>ag_name</b> = AG name extracted from dc_name (e.g. "GMGZ2_AG01", always populated)<br/>
<b>block</b> = block name from DJ Polygons (populated when DJ Polygons layer is provided)</p>
"""

    def initAlgorithm(self, config=None):
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_DJ_POLY, 'DJ Polygons -- one polygon per DJ (REQUIRED, must have block name attr)',
            [QgsProcessing.TypeVectorPolygon]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_DJ, 'DJ Points -- must be labeled (name, splitter, position, dc_name)',
            [QgsProcessing.TypeVectorPoint]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_HOUSES, 'House / Premise Points (HP)',
            [QgsProcessing.TypeVectorPoint]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_POLES, 'Pole Points -- must be labeled (name column)',
            [QgsProcessing.TypeVectorPoint]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_DROPS, 'Drop Cable Lines (optional -- fallback for houses not in DJ Polygons)',
            [QgsProcessing.TypeVectorLine], optional=True
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_DC, 'DC Lines (optional -- for FJ tracing)',
            [QgsProcessing.TypeVectorLine], optional=True
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_FJ, 'FJ Points (optional -- for FJ pole lookup)',
            [QgsProcessing.TypeVectorPoint], optional=True
        ))
        self.addParameter(QgsProcessingParameterNumber(
            self.PARAM_SNAP, 'Snap tolerance in meters',
            type=QgsProcessingParameterNumber.Double,
            defaultValue=5.0, minValue=0.1, maxValue=100.0
        ))
        self.addParameter(QgsProcessingParameterFileDestination(
            self.PARAM_OUTPUT, 'Output splicing diagram (XLSX)',
            fileFilter='Excel files (*.xlsx)'
        ))

    # ------------------------------------------------------------------
    # Excel generation -- integrated from generate_splicing_excel.py
    # ------------------------------------------------------------------
    @staticmethod
    def _build_splicing_excel(output_rows, output_path, feedback):
        """Build the splicing diagram Excel directly from output_rows.

        Column layout:
          A-C:   Fixed (Pole, FEEDER, AG)  — NOT repeated per DC
          D-M:   DC1 segment (10 cols)
          N-W:   DC2 segment (10 cols)
          X-AH:  DC3 segment (10 cols) ...
          last 3: slack_length, premise_data_name, Block
        """
        import re

        # --- Constants ---
        FIXED_COLS = 3       # Pole, FEEDER, AG
        COLS_PER_DC = 10     # DC_label, DC_number, Size, No., Splitter, Leg, Pole_num, DJ_num, DC_len, DC_slack
        TRAILING_COLS = 3    # slack_length, premise_data_name, Block

        HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        SUBHEADER_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        TITLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        BOLD_FONT = Font(bold=True, size=10)
        HEADER_FONT = Font(bold=True, size=11)
        TITLE_FONT = Font(bold=True, size=12)
        CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
        THIN_BORDER = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # --- Helper functions ---
        def _safe_val(v):
            if v is None:
                return ""
            v = str(v).strip()
            return "" if v.lower() in ("", "nan", "none", "null") else v

        def get_fc_from_fj(fj_name):
            if not fj_name:
                return "FC01"
            m = re.search(r"FJ(\d+)", str(fj_name))
            return f"FC{int(m.group(1)):02d}" if m else "FC01"

        def extract_sp_from_dj(dj_name):
            """GMGZ2_DJ205_1:8 -> SP_205"""
            m = re.search(r'DJ(\d+)', str(dj_name))
            return f"SP_{m.group(1)}" if m else ""

        def extract_dc_number(dc_name):
            """VTN_HHS_GMGZ2_AG13_DC053.4_... -> DC053.4"""
            m = re.search(r'DC(\d+\.\d+)', str(dc_name))
            return f"DC{m.group(1)}" if m else ""

        def extract_dc_info(dc_name):
            if not dc_name:
                return {"size": "", "length": "", "length_slack": ""}
            name = str(dc_name)
            size_m = re.search(r"_(\d+F)_", name)
            size = size_m.group(1) if size_m else ""
            len_m = re.search(r"\((\d+)m\s*-\s*(\d+)m\)", name)
            if len_m:
                return {"size": size, "length": int(len_m.group(1)), "length_slack": int(len_m.group(2))}
            len_m2 = re.search(r"\((\d+)m\)", name)
            if len_m2:
                d = int(len_m2.group(1))
                return {"size": size, "length": d, "length_slack": d + 10}
            return {"size": size, "length": "", "length_slack": ""}

        # --- Group rows by (ag_name -> block -> position) ---
        ag_groups = {}
        for row in output_rows:
            ag = _safe_val(row.get('ag_name', 'Unknown'))
            block = _safe_val(row.get('block', 'B001'))
            if not block:
                block = "B001"
            pos_str = str(row.get('position', '1'))
            try:
                pos = int(pos_str)
            except (ValueError, TypeError):
                pos = 1

            if ag not in ag_groups:
                ag_groups[ag] = {}
            if block not in ag_groups[ag]:
                ag_groups[ag][block] = {}

            if pos not in ag_groups[ag][block]:
                ag_groups[ag][block][pos] = {
                    'position': pos,
                    'original_dj_name': _safe_val(row.get('name_3', '')),
                    'original_dc_label': _safe_val(row.get('dc_name', '')),
                    'pole_number': _safe_val(row.get('name_2_2', '')),
                    'fj_name': _safe_val(row.get('fj_name', '')),
                    'fj_pole': _safe_val(row.get('fj_pole', '')),
                    'ag_name': ag,
                    'splitter': _safe_val(row.get('splitter', '1:8')),
                    'houses': [],
                }

            house = _safe_val(row.get('Name_2', ''))
            if house:
                ag_groups[ag][block][pos]['houses'].append(house)
                # Also store fields mapped by house name
                hf = row.get('house_fields', {})
                if hf:
                    if 'house_fields_map' not in ag_groups[ag][block][pos]:
                        ag_groups[ag][block][pos]['house_fields_map'] = {}
                    ag_groups[ag][block][pos]['house_fields_map'][house] = hf

        # Sort houses and convert to ordered DJ lists
        for ag in ag_groups:
            for block in ag_groups[ag]:
                for pos in ag_groups[ag][block]:
                    ag_groups[ag][block][pos]['houses'] = sorted(
                        set(ag_groups[ag][block][pos]['houses'])
                    )
                positions = sorted(ag_groups[ag][block].keys())
                ag_groups[ag][block] = [
                    ag_groups[ag][block][p] for p in positions
                ]

        # --- Build Excel ---
        wb = Workbook()
        ws = wb.active
        ws.title = "Splicing Diagram"

        max_djs_per_block = 0
        block_dj_counts = []
        for ag_name in sorted(ag_groups.keys()):
            for block_name in sorted(ag_groups[ag_name].keys()):
                num_djs = len(ag_groups[ag_name][block_name])
                block_dj_counts.append((ag_name, block_name, num_djs))
                if num_djs > max_djs_per_block:
                    max_djs_per_block = num_djs

        if max_djs_per_block == 0:
            feedback.pushInfo("  No data to write.")
            return

        # Collect all unique house field names (from first house in each group)
        house_field_names = []
        for ag_name in sorted(ag_groups.keys()):
            for block_name in sorted(ag_groups[ag_name].keys()):
                djs = ag_groups[ag_name][block_name]
                for dj in djs:
                    for h_name in dj.get('houses', []):
                        # houses are just strings now; get fields from output_rows
                        pass
        # Get field names from first output row that has house_fields
        house_field_names = []
        for row in output_rows:
            hf = row.get('house_fields', {})
            if hf:
                house_field_names = list(hf.keys())
                break
        num_house_fields = len(house_field_names)

        total_cols = FIXED_COLS + (max_djs_per_block * COLS_PER_DC) + TRAILING_COLS + num_house_fields
        trailing_base = FIXED_COLS + (max_djs_per_block * COLS_PER_DC) + 1
        house_fields_base = trailing_base + TRAILING_COLS  # start after trailing columns

        # Row 1: Title
        # Title 1: DC columns (1 to trailing_base - 1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=trailing_base - 1)
        c = ws.cell(row=1, column=1, value="Distribution Layer 1 (A - B)")
        c.font = TITLE_FONT; c.alignment = CENTER_ALIGNMENT; c.fill = HEADER_FILL

        # Title 2: Trailing columns (slack, premise name, Block)
        ws.merge_cells(start_row=1, start_column=trailing_base, end_row=1,
                       end_column=trailing_base + TRAILING_COLS - 1)
        c = ws.cell(row=1, column=trailing_base,
                    value="PREMIS DATA (MUST COMPLETE ALL COLUMS & DUPLICATE INFO REVIEWED)")
        c.font = TITLE_FONT; c.alignment = CENTER_ALIGNMENT; c.fill = TITLE_FILL

        # Title 3: House field columns (if any)
        if num_house_fields > 0:
            ws.merge_cells(start_row=1, start_column=house_fields_base, end_row=1,
                           end_column=total_cols)
            c = ws.cell(row=1, column=house_fields_base, value="PREMISE DATA FIELDS")
            c.font = TITLE_FONT; c.alignment = CENTER_ALIGNMENT; c.fill = TITLE_FILL

        for col in range(1, total_cols + 1):
            cell = ws.cell(row=1, column=col)
            if col < trailing_base:
                cell.fill = HEADER_FILL
            elif col < house_fields_base:
                cell.fill = TITLE_FILL
            else:
                cell.fill = TITLE_FILL
            cell.border = THIN_BORDER

        # Row 2: Category headers (merged groups per DC segment)
        for dc_idx in range(max_djs_per_block):
            base = FIXED_COLS + (dc_idx * COLS_PER_DC) + 1   # 1-based

            # Cable Route ID merges all 10 columns of this DC segment
            ws.merge_cells(
                start_row=2, start_column=base,
                end_row=2, end_column=base + COLS_PER_DC - 1
            )
            c = ws.cell(row=2, column=base,
                        value=f"Cable Route ID (AERIAL)_{dc_idx + 1}")
            c.font = BOLD_FONT; c.alignment = CENTER_ALIGNMENT
            c.fill = SUBHEADER_FILL; c.border = THIN_BORDER
            for o in range(COLS_PER_DC):
                ws.cell(row=2, column=base + o).border = THIN_BORDER
                ws.cell(row=2, column=base + o).fill = SUBHEADER_FILL

        # Trailing category headers
        ws.merge_cells(start_row=2, start_column=trailing_base, end_row=2, end_column=trailing_base + 1)
        c = ws.cell(row=2, column=trailing_base, value="+ Slack")
        c.font = BOLD_FONT; c.alignment = CENTER_ALIGNMENT; c.fill = SUBHEADER_FILL; c.border = THIN_BORDER
        ws.cell(row=2, column=trailing_base + 1).border = THIN_BORDER
        ws.cell(row=2, column=trailing_base + 1).fill = SUBHEADER_FILL

        c = ws.cell(row=2, column=trailing_base + 2, value="Block")
        c.font = BOLD_FONT; c.alignment = CENTER_ALIGNMENT; c.fill = SUBHEADER_FILL; c.border = THIN_BORDER

        # Row 3: Detailed headers
        # Fixed headers (cols 1-3)
        fixed_headers = ["Pole", "FEEDER", "AG"]
        for hi, hdr in enumerate(fixed_headers):
            c = ws.cell(row=3, column=1 + hi, value=hdr)
            c.font = BOLD_FONT; c.alignment = CENTER_ALIGNMENT
            c.fill = HEADER_FILL; c.border = THIN_BORDER

        # Per-DC detail headers (repeated for each DC segment)
        dc_detail_headers = [
            "DC label", "DC number", "Size", "No.",
            "splitter number", "Leg", "Pole_number",
            "DJ_number", "DC_length", "DC+slack",
        ]
        for dc_idx in range(max_djs_per_block):
            base = FIXED_COLS + (dc_idx * COLS_PER_DC) + 1
            for hi, hdr in enumerate(dc_detail_headers):
                c = ws.cell(row=3, column=base + hi, value=hdr)
                c.font = BOLD_FONT; c.alignment = CENTER_ALIGNMENT
                c.fill = HEADER_FILL; c.border = THIN_BORDER

        # Trailing detail headers
        for off, hdr in enumerate(["slack_length", "premise_data_name", "Block"]):
            c = ws.cell(row=3, column=trailing_base + off, value=hdr)
            c.font = BOLD_FONT; c.alignment = CENTER_ALIGNMENT
            c.fill = HEADER_FILL; c.border = THIN_BORDER

        # House field headers (dynamic, from premise data layer)
        for fi, fn in enumerate(house_field_names):
            c = ws.cell(row=3, column=house_fields_base + fi, value=fn)
            c.font = BOLD_FONT; c.alignment = CENTER_ALIGNMENT
            c.fill = HEADER_FILL; c.border = THIN_BORDER

        # --- Data Rows ---
        current_row = 4
        for ag_name, block_name, num_djs in block_dj_counts:
            djs = ag_groups[ag_name][block_name]

            # Precompute per-DJ info using ACTUAL labels from CSV
            dj_info_list = []
            for dj_entry in djs:
                dc_label = _safe_val(dj_entry.get('original_dc_label', ''))
                dc_info = extract_dc_info(dc_label)
                dj_name = _safe_val(dj_entry.get('original_dj_name', ''))

                info = {
                    'dc_label': dc_label,
                    'dc_number': extract_dc_number(dc_label),
                    'dc_size': dc_info['size'],
                    'dc_length': dc_info['length'],
                    'dc_slack': dc_info['length_slack'],
                    'splitter_number': extract_sp_from_dj(dj_name),
                    'pole': _safe_val(dj_entry.get('pole_number', '')),
                    'dj_number': dj_name,   # ACTUAL DJ name from CSV
                }
                dj_info_list.append(info)

            # Fixed column values come from the FIRST DJ in this block
            first_pole = _safe_val(djs[0].get('fj_pole', '')) if djs else ''
            first_fc = get_fc_from_fj(djs[0]['fj_name']) if djs else 'FC01'
            first_ag = _safe_val(djs[0]['ag_name']) if djs else ''

            for dj_idx in range(num_djs):
                dj_entry = djs[dj_idx]
                house_list = dj_entry['houses']

                # Alternating row color: even DJ = white, odd DJ = light blue
                dj_fill = None if (dj_idx % 2 == 0) else PatternFill(
                    start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")

                for leg in range(1, 9):
                    # 1. Write FIXED columns (A, B, C) — from FIRST DJ in block
                    ws.cell(row=current_row, column=1, value=first_pole)   # Pole
                    ws.cell(row=current_row, column=2, value=first_fc)     # FEEDER
                    ws.cell(row=current_row, column=3, value=first_ag)     # AG

                    # 2. Write per-DC segment columns
                    for seg_idx in range(num_djs):
                        base = FIXED_COLS + (seg_idx * COLS_PER_DC) + 1
                        source_info = dj_info_list[seg_idx]

                        if seg_idx < dj_idx:
                            show_leg = "9"          # link fiber
                        elif seg_idx == dj_idx:
                            show_leg = str(leg)      # actual house leg 1-8
                        else:
                            continue                  # skip (downstream of current DJ)

                        ws.cell(row=current_row, column=base + 0, value=source_info['dc_label'])
                        ws.cell(row=current_row, column=base + 1, value=source_info['dc_number'])
                        ws.cell(row=current_row, column=base + 2, value=source_info['dc_size'])
                        ws.cell(row=current_row, column=base + 3, value="1")
                        ws.cell(row=current_row, column=base + 4, value=source_info['splitter_number'])

                        # Leg column: yellow fill for "9", normal otherwise
                        leg_cell = ws.cell(row=current_row, column=base + 5, value=show_leg)
                        if show_leg == "9":
                            leg_cell.fill = PatternFill(
                                start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                        ws.cell(row=current_row, column=base + 6, value=source_info['pole'])
                        ws.cell(row=current_row, column=base + 7, value=source_info['dj_number'])
                        _dl = source_info['dc_length']
                        _ds = source_info['dc_slack']
                        ws.cell(row=current_row, column=base + 8,
                                value=f"{int(_dl):04d}" if _dl != '' else '')
                        ws.cell(row=current_row, column=base + 9,
                                value=f"{int(_ds):04d}" if _ds != '' else '')

                    # 3. Write TRAILING columns
                    if leg <= len(house_list):
                        house_name = house_list[leg - 1]
                    else:
                        house_name = ""

                    ws.cell(row=current_row, column=trailing_base + 0, value="")
                    ws.cell(row=current_row, column=trailing_base + 1, value=house_name)
                    ws.cell(row=current_row, column=trailing_base + 2, value=block_name)

                    # 4. Write HOUSE FIELD columns (all fields from premise data layer)
                    if house_name and num_house_fields > 0:
                        # Get fields from the current DJ's house_fields_map
                        fields_map = dj_entry.get('house_fields_map', {})
                        house_data = fields_map.get(house_name, {})
                        for fi, fn in enumerate(house_field_names):
                            ws.cell(row=current_row, column=house_fields_base + fi,
                                    value=house_data.get(fn, ''))

                    # Track Leg=9 yellow cells so we don't overwrite them
                    yellow_cols = set()
                    for seg_idx in range(num_djs):
                        base = FIXED_COLS + (seg_idx * COLS_PER_DC) + 1
                        source_info = dj_info_list[seg_idx]
                        if seg_idx < dj_idx:
                            yellow_cols.add(base + 5)  # Leg column for upstream DJ

                    # Borders + alignment + alternating row color for ALL columns
                    for col in range(1, total_cols + 1):
                        cell = ws.cell(row=current_row, column=col)
                        cell.border = THIN_BORDER
                        cell.alignment = CENTER_ALIGNMENT
                        # Apply alternating row color, but PRESERVE yellow Leg=9 fills
                        if dj_fill and col < trailing_base and col not in yellow_cols:
                            cell.fill = dj_fill

                    current_row += 1

                # If this is the LAST DJ in the block, highlight the last row in Block column (yellow)
                if dj_idx == num_djs - 1 and current_row > 4:
                    last_row = current_row - 1
                    block_cell = ws.cell(row=last_row, column=trailing_base + 2)
                    block_cell.fill = PatternFill(
                        start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # --- Column widths ---
        # Fixed columns
        ws.column_dimensions[get_column_letter(1)].width = 16   # Pole
        ws.column_dimensions[get_column_letter(2)].width = 10   # FEEDER
        ws.column_dimensions[get_column_letter(3)].width = 16   # AG

        # Per-DC columns
        for dc_idx in range(max_djs_per_block):
            base = FIXED_COLS + (dc_idx * COLS_PER_DC) + 1
            ws.column_dimensions[get_column_letter(base + 0)].width = 55   # DC label
            ws.column_dimensions[get_column_letter(base + 1)].width = 12   # DC number
            ws.column_dimensions[get_column_letter(base + 2)].width = 8    # Size
            ws.column_dimensions[get_column_letter(base + 3)].width = 6    # No.
            ws.column_dimensions[get_column_letter(base + 4)].width = 16   # splitter number
            ws.column_dimensions[get_column_letter(base + 5)].width = 8    # Leg
            ws.column_dimensions[get_column_letter(base + 6)].width = 16   # Pole_number
            ws.column_dimensions[get_column_letter(base + 7)].width = 25   # DJ_number
            ws.column_dimensions[get_column_letter(base + 8)].width = 12   # DC_length
            ws.column_dimensions[get_column_letter(base + 9)].width = 12   # DC+slack

        # Trailing columns
        ws.column_dimensions[get_column_letter(trailing_base + 0)].width = 14
        ws.column_dimensions[get_column_letter(trailing_base + 1)].width = 30
        ws.column_dimensions[get_column_letter(trailing_base + 2)].width = 10

        # House field columns (dynamic widths)
        for fi, fn in enumerate(house_field_names):
            ws.column_dimensions[get_column_letter(house_fields_base + fi)].width = max(12, len(fn) + 2)

        ws.freeze_panes = "A4"
        wb.save(output_path)
        feedback.pushInfo("  Data rows: {}".format(current_row - 4))

    # ------------------------------------------------------------------
    # Main processing
    # ------------------------------------------------------------------
    def processAlgorithm(self, parameters, context, feedback):
        snap_tol = self.parameterAsDouble(parameters, self.PARAM_SNAP, context)
        output_path = self.parameterAsString(parameters, self.PARAM_OUTPUT, context)

        dj_poly_layer = self.parameterAsVectorLayer(parameters, self.IN_DJ_POLY, context)
        dj_layer      = self.parameterAsVectorLayer(parameters, self.IN_DJ,      context)
        house_layer   = self.parameterAsVectorLayer(parameters, self.IN_HOUSES,  context)
        pole_layer    = self.parameterAsVectorLayer(parameters, self.IN_POLES,   context)
        drops_layer   = self.parameterAsVectorLayer(parameters, self.IN_DROPS,   context)
        dc_layer      = self.parameterAsVectorLayer(parameters, self.IN_DC,      context)
        fj_layer      = self.parameterAsVectorLayer(parameters, self.IN_FJ,      context)

        feedback.pushInfo("=" * 60)
        feedback.pushInfo("FTTH Splicing Plan v8.0 -- Excel output")
        feedback.pushInfo("=" * 60)
        feedback.pushInfo("Snap tolerance: {}m".format(snap_tol))
        feedback.pushInfo("")

        # ================================================================
        # Helper: case-insensitive field lookup
        # ================================================================
        def find_field(layer, names):
            """Find field index, case-insensitive."""
            all_names = [f.name() for f in layer.fields()]
            for fn in names:
                fn_lower = fn.lower()
                for actual_name in all_names:
                    if actual_name.lower() == fn_lower:
                        return layer.fields().indexFromName(actual_name)
            return -1

        # ================================================================
        # Helper: get start and end points of a line
        # ================================================================
        def get_line_endpoints(line_geom):
            """Get start and end QgsPointXY of a LineString/MultiLineString."""
            if line_geom.isMultipart():
                parts = line_geom.asMultiPolyline()
                if parts:
                    first_part = parts[0]
                    last_part = parts[-1]
                    if first_part and last_part:
                        return (QgsPointXY(first_part[0]), QgsPointXY(last_part[-1]))
            else:
                pts = line_geom.asPolyline()
                if pts and len(pts) >= 2:
                    return (QgsPointXY(pts[0]), QgsPointXY(pts[-1]))
            return (None, None)

        # ================================================================
        # Helper: find nearest node to a point
        # ================================================================
        def nearest_node(point, nodes, tolerance):
            """Find nearest node to a point within tolerance.
            nodes: dict of {fid: {'point': QgsPointXY, ...}}
            Returns: (fid, info) or (None, None)
            """
            best_fid = None
            best_info = None
            best_dist = float('inf')
            for fid, info in nodes.items():
                d = point.distance(info['point'])
                if d < best_dist and d <= tolerance:
                    best_dist = d
                    best_fid = fid
                    best_info = info
            return best_fid, best_info

        # ================================================================
        # Helper: find nearest pole to a point
        # ================================================================
        def nearest_pole(point, poles_list, tolerance):
            """Find nearest pole to a point within tolerance.
            poles_list: list of {'fid': int, 'point': QgsPointXY, 'name': str}
            Returns: (fid, name) or (None, '')
            """
            best = None
            best_dist = float('inf')
            for p in poles_list:
                d = point.distance(p['point'])
                if d < best_dist and d <= tolerance:
                    best_dist = d
                    best = p
            if best:
                return best['fid'], best['name']
            return None, ''

        # ================================================================
        # Helper: extract AG name from DC label
        # e.g. "VTN_HHS_GMGZ2_AG01_DC001.1_1F_ADSS_G.657.A1(152m-162m)" -> "GMGZ2_AG01"
        # ================================================================
        def extract_ag(dc_name):
            """Extract AG name from DC label. Returns '' if not found."""
            if not dc_name:
                return ''
            m = re.search(r'([A-Za-z0-9]+_AG\d+)', str(dc_name))
            return m.group(1) if m else ''

        # ================================================================
        # Helper: walk upstream from DJ to find FJ
        # ================================================================
        def find_fj_for_dj(dj_name, dc_by_to, fjs):
            """Trace upstream from DJ through DC lines to find FJ.
            Returns (fj_name, fj_point) or ('', None)
            """
            if not dj_name or not dc_by_to:
                return '', None
            visited = set()
            current = dj_name
            while current and current not in visited:
                visited.add(current)
                dc_info = dc_by_to.get(current)
                if not dc_info:
                    break
                upstream = dc_info['from']
                if upstream in fjs:
                    return upstream, fjs[upstream]['point']
                current = upstream
            return '', None

        # ================================================================
        # Step 1 -- Detect fields and validate
        # ================================================================
        feedback.pushInfo("[Step 1] Detecting fields...")

        # --- DJ Polygon block field (auto-detect) ---
        poly_block_idx = find_field(dj_poly_layer, [
            'block', 'Block', 'BLOCK',
            'b_name', 'B_NAME',
            'id', 'ID'
        ])

        # --- DJ fields (must be labeled by Labeler plugin) ---
        dj_name_idx = find_field(dj_layer, ['name', 'Name', 'NAME'])
        dj_splitter_idx = find_field(dj_layer, ['splitter', 'Splitter', 'SPLITTER'])
        dj_position_idx = find_field(dj_layer, ['position', 'Position', 'POSITION'])
        dj_dcname_idx = find_field(dj_layer, ['dc_name', 'DC_name', 'dcname', 'DCNAME'])

        # --- House name field (extensive auto-detection) ---
        house_name_idx = find_field(house_layer, [
            'name', 'Name', 'NAME',
            'premise', 'Premise', 'PREMISE',
            'house', 'House', 'HOUSE',
            'hp', 'HP',
            'street_no', 'Street_No', 'STREET_NO', 'streetno',
            'fid', 'FID'
        ])

        # --- Pole name field ---
        pole_name_idx = find_field(pole_layer, ['name', 'Name', 'NAME'])

        # --- Drop Cable name field (optional) ---
        drop_name_idx = -1
        if drops_layer and drops_layer.isValid():
            drop_name_idx = find_field(drops_layer, ['name', 'Name', 'NAME'])

        # Validate required layers
        if not dj_poly_layer or not dj_poly_layer.isValid():
            feedback.reportError("DJ Polygons layer is REQUIRED. Please provide the DJ Polygons layer.")
            return {self.PARAM_OUTPUT: output_path}
        if dj_name_idx < 0:
            feedback.reportError("DJ layer has no 'name' column. Run FTTH Labeler first!")
            return {self.PARAM_OUTPUT: output_path}

        feedback.pushInfo("  DJ Polygon block field: {}".format(
            'found' if poly_block_idx >= 0 else 'NOT FOUND'))
        feedback.pushInfo("  DJ name field:          found")
        feedback.pushInfo("  DJ splitter field:      {}".format(
            'found' if dj_splitter_idx >= 0 else 'NOT FOUND'))
        feedback.pushInfo("  DJ position field:      {}".format(
            'found' if dj_position_idx >= 0 else 'NOT FOUND'))
        feedback.pushInfo("  DJ dc_name field:       {}".format(
            'found' if dj_dcname_idx >= 0 else 'NOT FOUND'))
        feedback.pushInfo("  House name field:       {}".format(
            'found' if house_name_idx >= 0 else 'NOT FOUND (using HP_{fid})'))
        feedback.pushInfo("  Pole name field:        {}".format(
            'found' if pole_name_idx >= 0 else 'NOT FOUND'))
        feedback.pushInfo("  Drop Cable name field:  {}".format(
            'found' if drop_name_idx >= 0 else 'NOT FOUND (optional)'))

        # ================================================================
        # Step 2 -- Index all layers into memory
        # ================================================================
        feedback.pushInfo("")
        feedback.pushInfo("[Step 2] Indexing layers into memory...")

        # --- Index DJ Polygons: list of {'fid', 'geom', 'block'} ---
        dj_polygons = []
        for feat in dj_poly_layer.getFeatures():
            geom = feat.geometry()
            blk = ''
            if poly_block_idx >= 0:
                blk = str(feat[poly_block_idx] or '').strip()
            dj_polygons.append({
                'fid': feat.id(),
                'geom': geom,
                'block': blk
            })
        feedback.pushInfo("  Indexed {} DJ polygons".format(len(dj_polygons)))

        if len(dj_polygons) == 0:
            feedback.reportError("DJ Polygons layer has 0 features. Nothing to process.")
            return {self.PARAM_OUTPUT: output_path}

        # --- Index DJ Points: {name: {'fid', 'point', 'splitter', 'position', 'dc_name'}} ---
        djs = {}
        djs_by_fid = {}
        for feat in dj_layer.getFeatures():
            geom = feat.geometry()
            pt = geom.asPoint()
            dj_name = str(feat[dj_name_idx] or '')
            splitter = str(feat[dj_splitter_idx] or '') if dj_splitter_idx >= 0 else ''
            position = feat[dj_position_idx] if dj_position_idx >= 0 else ''
            dc_name = str(feat[dj_dcname_idx] or '') if dj_dcname_idx >= 0 else ''
            info = {
                'fid': feat.id(),
                'point': pt,
                'name': dj_name,
                'splitter': splitter,
                'position': position,
                'dc_name': dc_name
            }
            djs[dj_name] = info
            djs_by_fid[feat.id()] = info
        feedback.pushInfo("  Indexed {} DJ points".format(len(djs)))

        # --- Index Houses: {fid: {'point', 'name', 'fields'}} ---
        # Store ALL fields from the house layer for Excel output
        house_field_names = [f.name() for f in house_layer.fields()]
        houses = {}
        for feat in house_layer.getFeatures():
            geom = feat.geometry()
            pt = geom.asPoint()
            if house_name_idx >= 0:
                nm = str(feat[house_name_idx] or '').strip()
            else:
                nm = "HP_{}".format(feat.id())
            # Store all field values
            all_fields = {}
            for fn in house_field_names:
                all_fields[fn] = str(feat[fn] or '')
            houses[feat.id()] = {
                'point': pt,
                'name': nm,
                'fields': all_fields
            }
        feedback.pushInfo("  Indexed {} house points".format(len(houses)))

        # --- Index Poles: list of {'fid', 'point', 'name'} ---
        poles = []
        for feat in pole_layer.getFeatures():
            geom = feat.geometry()
            pt = geom.asPoint()
            nm = str(feat[pole_name_idx] or '') if pole_name_idx >= 0 else "Pole_{}".format(feat.id())
            poles.append({
                'fid': feat.id(),
                'point': pt,
                'name': nm
            })
        feedback.pushInfo("  Indexed {} poles".format(len(poles)))

        # --- Index DC lines (for upstream tracing) ---
        dc_by_to = {}   # {to_node_name: {'from': from_node_name, 'dc_name': str}}
        if dc_layer and dc_layer.isValid():
            dc_from_idx = find_field(dc_layer, ['from_node', 'from', 'From_Node', 'FROM_NODE'])
            dc_to_idx = find_field(dc_layer, ['to_node', 'to', 'To_Node', 'TO_NODE'])
            dc_name_idx2 = find_field(dc_layer, ['name', 'Name', 'NAME'])
            for feat in dc_layer.getFeatures():
                fn = str(feat[dc_from_idx] or '') if dc_from_idx >= 0 else ''
                tn = str(feat[dc_to_idx] or '') if dc_to_idx >= 0 else ''
                dn = str(feat[dc_name_idx2] or '') if dc_name_idx2 >= 0 else ''
                if tn and fn:
                    dc_by_to[tn] = {'from': fn, 'dc_name': dn}
            feedback.pushInfo("  Indexed {} DC lines".format(len(dc_by_to)))

        # --- Index FJ points (for FJ pole lookup) ---
        fjs = {}   # {fj_name: {'point': QgsPointXY}}
        if fj_layer and fj_layer.isValid():
            fj_name_idx2 = find_field(fj_layer, ['name', 'Name', 'NAME'])
            for feat in fj_layer.getFeatures():
                nm = str(feat[fj_name_idx2] or '') if fj_name_idx2 >= 0 else ''
                if nm:
                    fjs[nm] = {'point': feat.geometry().asPoint()}
            feedback.pushInfo("  Indexed {} FJ points".format(len(fjs)))

        # ================================================================
        # Step 3 -- Match DJ Polygons to DJ Points
        # ================================================================
        feedback.pushInfo("")
        feedback.pushInfo("[Step 3] Matching DJ Polygons to DJ Points...")

        # {polygon_fid: dj_name}
        poly_to_dj = {}
        unmatched_polys = []

        for poly in dj_polygons:
            poly_geom = poly['geom']
            best_dj = None
            best_dist = float('inf')

            # Try 1: point-in-polygon (DJ point inside polygon)
            for dj_name, dj_info in djs.items():
                if poly_geom.contains(dj_info['point']):
                    best_dj = dj_name
                    break

            # Try 2: nearest DJ within 10m of polygon centroid
            if not best_dj:
                centroid = poly_geom.centroid().asPoint()
                for dj_name, dj_info in djs.items():
                    d = centroid.distance(dj_info['point'])
                    if d < best_dist and d <= 10.0:
                        best_dist = d
                        best_dj = dj_name

            if best_dj:
                poly_to_dj[poly['fid']] = best_dj
            else:
                unmatched_polys.append(poly['fid'])

        feedback.pushInfo("  Matched {} of {} DJ Polygons to DJ Points".format(
            len(poly_to_dj), len(dj_polygons)))
        if unmatched_polys:
            feedback.pushInfo("  WARNING: {} DJ Polygons could not be matched to any DJ Point".format(
                len(unmatched_polys)))

        # ================================================================
        # Step 4 -- Match Houses to DJ Polygons (PRIMARY method)
        # ================================================================
        feedback.pushInfo("")
        feedback.pushInfo("[Step 4] Matching Houses to DJ Polygons (point-in-polygon)...")

        # {dj_name: [house_info, ...]}
        dj_to_houses = {}
        unmatched_houses = {}   # houses not matched by any DJ Polygon

        for house_fid, house_info in houses.items():
            matched = False
            for poly in dj_polygons:
                dj_name = poly_to_dj.get(poly['fid'])
                if not dj_name:
                    continue
                if poly['geom'].contains(house_info['point']):
                    if dj_name not in dj_to_houses:
                        dj_to_houses[dj_name] = []
                    dj_to_houses[dj_name].append({
                        'fid': house_fid,
                        'name': house_info['name'],
                        'fields': house_info.get('fields', {})
                    })
                    matched = True
                    break
            if not matched:
                unmatched_houses[house_fid] = house_info

        total_matched = sum(len(hlist) for hlist in dj_to_houses.values())
        feedback.pushInfo("  Matched {} houses to DJs via DJ Polygons".format(total_matched))
        if unmatched_houses:
            feedback.pushInfo("  {} houses not inside any DJ Polygon".format(len(unmatched_houses)))

        # ================================================================
        # Step 5 -- Fallback: match unmatched houses via Drop Cables
        # ================================================================
        if unmatched_houses and drops_layer and drops_layer.isValid():
            feedback.pushInfo("")
            feedback.pushInfo("[Step 5] Fallback: using Drop Cables for {} unmatched houses...".format(
                len(unmatched_houses)))

            # Build temporary DJ index by fid for nearest-node lookup
            djs_for_nearest = {}
            for dj_name, dj_info in djs.items():
                djs_for_nearest[dj_info['fid']] = {
                    'point': dj_info['point'],
                    'name': dj_name
                }

            fallback_matched = 0
            skipped_drops = 0

            for drop_feat in drops_layer.getFeatures():
                if feedback.isCanceled():
                    return {self.PARAM_OUTPUT: output_path}

                try:
                    drop_fid = drop_feat.id()
                    drop_geom = drop_feat.geometry()
                    pt_a, pt_b = get_line_endpoints(drop_geom)
                    if pt_a is None or pt_b is None:
                        skipped_drops += 1
                        continue

                    # Strategy: one end is an unmatched house, other end is a DJ
                    # Try both combinations
                    # Option 1: A=unmatched_house, B=DJ
                    h_fid_1, h_info_1 = nearest_node(pt_a, unmatched_houses, snap_tol)
                    d_fid_1, d_info_1 = nearest_node(pt_b, djs_for_nearest, snap_tol)
                    opt1_ok = (h_fid_1 is not None and d_fid_1 is not None)

                    # Option 2: A=DJ, B=unmatched_house
                    d_fid_2, d_info_2 = nearest_node(pt_a, djs_for_nearest, snap_tol)
                    h_fid_2, h_info_2 = nearest_node(pt_b, unmatched_houses, snap_tol)
                    opt2_ok = (d_fid_2 is not None and h_fid_2 is not None)

                    house_fid = None
                    dj_name = None

                    if opt1_ok and opt2_ok:
                        dist1 = pt_a.distance(h_info_1['point']) + pt_b.distance(d_info_1['point'])
                        dist2 = pt_a.distance(d_info_2['point']) + pt_b.distance(h_info_2['point'])
                        if dist1 <= dist2:
                            house_fid, dj_name = h_fid_1, d_info_1['name']
                        else:
                            house_fid, dj_name = h_fid_2, d_info_2['name']
                    elif opt1_ok:
                        house_fid, dj_name = h_fid_1, d_info_1['name']
                    elif opt2_ok:
                        house_fid, dj_name = h_fid_2, d_info_2['name']
                    else:
                        skipped_drops += 1
                        continue

                    if house_fid and house_fid in unmatched_houses:
                        house_info = unmatched_houses[house_fid]
                        if dj_name not in dj_to_houses:
                            dj_to_houses[dj_name] = []
                        dj_to_houses[dj_name].append({
                            'fid': house_fid,
                            'name': house_info['name'],
                            'fields': house_info.get('fields', {})
                        })
                        del unmatched_houses[house_fid]
                        fallback_matched += 1

                except Exception as e:
                    skipped_drops += 1
                    continue

            feedback.pushInfo("  Fallback matched {} additional houses via Drop Cables".format(
                fallback_matched))
            if skipped_drops > 0:
                feedback.pushInfo("  Skipped {} drop cables".format(skipped_drops))

        if unmatched_houses:
            feedback.pushInfo("  WARNING: {} houses remain unmatched (no DJ Polygon, no Drop Cable)".format(
                len(unmatched_houses)))

        # ================================================================
        # Step 6 -- Find nearest Pole for each DJ
        # ================================================================
        feedback.pushInfo("")
        feedback.pushInfo("[Step 6] Finding nearest poles for each DJ...")

        # {dj_name: {'pole_fid', 'pole_name'}}
        dj_to_pole = {}
        for poly in dj_polygons:
            dj_name = poly_to_dj.get(poly['fid'])
            if not dj_name or dj_name not in djs:
                continue
            dj_info = djs[dj_name]
            pole_fid, pole_name = nearest_pole(dj_info['point'], poles, snap_tol)
            dj_to_pole[dj_name] = {
                'pole_fid': pole_fid or '',
                'pole_name': pole_name
            }

        feedback.pushInfo("  Found poles for {} DJs".format(len(dj_to_pole)))

        # ================================================================
        # Step 7 -- Trace upstream to FJ (if DC and FJ provided)
        # ================================================================
        feedback.pushInfo("")
        feedback.pushInfo("[Step 7] Tracing upstream to FJ...")

        # {dj_name: {'fj_name', 'fj_pole'}}
        dj_to_fj = {}
        if dc_by_to and fjs:
            for poly in dj_polygons:
                dj_name = poly_to_dj.get(poly['fid'])
                if not dj_name:
                    continue
                fj_name, fj_point = find_fj_for_dj(dj_name, dc_by_to, fjs)
                fj_pole_name = ''
                if fj_name and fj_point:
                    _, fj_pole_name = nearest_pole(fj_point, poles, snap_tol)
                dj_to_fj[dj_name] = {
                    'fj_name': fj_name,
                    'fj_pole': fj_pole_name
                }
            feedback.pushInfo("  Traced FJ for {} DJs".format(len(dj_to_fj)))
        else:
            feedback.pushInfo("  Skipped (DC Lines and/or FJ Points not provided)")

        # ================================================================
        # Step 8 -- Build output rows
        # ================================================================
        feedback.pushInfo("")
        feedback.pushInfo("[Step 8] Building output rows...")

        output_rows = []
        row_counter = 0

        # Process ALL DJ Polygons in order (ensures ALL DJs appear, even with 0 houses)
        for poly in dj_polygons:
            dj_name = poly_to_dj.get(poly['fid'])
            if not dj_name:
                # Unmatched polygon -- skip (DJ point not found)
                continue

            dj_info = djs.get(dj_name)
            if not dj_info:
                continue

            pole_info = dj_to_pole.get(dj_name, {'pole_fid': '', 'pole_name': ''})
            fj_info = dj_to_fj.get(dj_name, {'fj_name': '', 'fj_pole': ''})
            ag_name = extract_ag(dj_info['dc_name'])
            block_name = poly.get('block', '')

            house_list = dj_to_houses.get(dj_name, [])

            if house_list:
                # Sort houses alphabetically
                house_list_sorted = sorted(house_list, key=lambda h: h['name'])
                for h in house_list_sorted:
                    row_counter += 1
                    output_rows.append({
                        'fid': row_counter,
                        'Name': '',
                        'fid_2': h['fid'],
                        'Name_2': h['name'],
                        'fid_3': dj_info['fid'],
                        'name_3': dj_name,
                        'splitter': dj_info['splitter'],
                        'position': dj_info['position'],
                        'dc_name': dj_info['dc_name'],
                        'fid_2_2': pole_info['pole_fid'],
                        'name_2_2': pole_info['pole_name'],
                        'fj_name': fj_info['fj_name'],
                        'fj_pole': fj_info['fj_pole'],
                        'ag_name': ag_name,
                        'block': block_name,
                        'house_fields': h.get('fields', {}),
                    })
            else:
                # DJ has no houses -- still output one row with empty Name_2
                row_counter += 1
                output_rows.append({
                    'fid': row_counter,
                    'Name': '',
                    'fid_2': '',
                    'Name_2': '',
                    'fid_3': dj_info['fid'],
                    'name_3': dj_name,
                    'splitter': dj_info['splitter'],
                    'position': dj_info['position'],
                    'dc_name': dj_info['dc_name'],
                    'fid_2_2': pole_info['pole_fid'],
                    'name_2_2': pole_info['pole_name'],
                    'fj_name': fj_info['fj_name'],
                    'fj_pole': fj_info['fj_pole'],
                    'ag_name': ag_name,
                    'block': block_name,
                    'house_fields': {},
                })

        feedback.pushInfo("  Built {} output rows for {} DJs".format(
            len(output_rows), len([p for p in dj_polygons if p['fid'] in poly_to_dj])))

        # ================================================================
        # Step 9 -- Build Excel splicing diagram
        # ================================================================
        feedback.pushInfo("")
        feedback.pushInfo("[Step 9] Building Excel splicing diagram...")

        if not HAS_OPENPYXL:
            feedback.reportError("  ERROR: openpyxl not available. Cannot generate Excel.")
            feedback.reportError("  Install: python -m pip install openpyxl")
            return {self.PARAM_OUTPUT: output_path}

        try:
            self._build_splicing_excel(output_rows, output_path, feedback)
            feedback.pushInfo("  [OK] Excel written: {}".format(output_path))
        except Exception as e:
            feedback.reportError("  [ERROR] Building Excel: {}".format(e))
            import traceback
            feedback.reportError(traceback.format_exc())
            return {self.PARAM_OUTPUT: output_path}

        feedback.pushInfo("")
        feedback.pushInfo("=" * 60)
        feedback.pushInfo("SPLICING PLAN COMPLETE")
        feedback.pushInfo("  DJ Polygons processed : {}".format(len(dj_polygons)))
        feedback.pushInfo("  DJs matched           : {}".format(len(poly_to_dj)))
        feedback.pushInfo("  Houses matched (poly) : {}".format(total_matched))
        feedback.pushInfo("  Output rows           : {}".format(len(output_rows)))
        feedback.pushInfo("  DJs with houses       : {}".format(
            len([d for d in dj_to_houses if dj_to_houses[d]])))
        feedback.pushInfo("  DJs with 0 houses     : {}".format(
            len([d for d in dj_to_houses if not dj_to_houses[d]])))
        feedback.pushInfo("  Output file           : {}".format(output_path))

        return {self.PARAM_OUTPUT: output_path}


# =============================================================================
# FTTH BOM ALGORITHM
# =============================================================================

class FTTHBOMAlgorithm(QgsProcessingAlgorithm):
    """Generate FTTH Bill of Materials (BOM) Excel.

    Calculates materials from DJ toward OLT:
    - Pre-conventional: DJ boxes, FJ closures, drop cables, pigtails, couplers
    - Conventional: feeder cable, dome joints, 1:2 splitters, core cable, ODF, patch cords
    - Pole hardware: tangent/dead-end clamps, hooks, brackets, bandit straps
    - Consumables: bunny clips, cable ties, tape, cleaning supplies
    """

    # Input layers
    IN_DJ_POLY = 'IN_DJ_POLY'
    IN_DJ = 'IN_DJ'
    IN_HOUSES = 'IN_HOUSES'
    IN_POLES = 'IN_POLES'
    IN_DC = 'IN_DC'
    IN_FC = 'IN_FC'
    IN_FJ = 'IN_FJ'

    # Parameters
    PARAM_AREA = 'PARAM_AREA'
    PARAM_ZONE = 'PARAM_ZONE'
    PARAM_CABLE_SIZE = 'PARAM_CABLE_SIZE'
    PARAM_FIBERS_PER_FJ = 'PARAM_FIBERS_PER_FJ'
    PARAM_SPLITTERS_PER_FJ = 'PARAM_SPLITTERS_PER_FJ'
    PARAM_POLE_TYPE = 'PARAM_POLE_TYPE'
    PARAM_SPLICING_PLAN = 'PARAM_SPLICING_PLAN'
    PARAM_OUTPUT = 'PARAM_OUTPUT'

    # Excel styles (reused from splicing plan)
    _HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    _HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
    _SUBHEADER_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    _SUBHEADER_FONT = Font(bold=True, size=10)
    _THIN_BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def name(self):
        return 'ftth_bom'

    def displayName(self):
        return 'FTTH BOM Generator'

    def group(self):
        return 'FTTH Tools'

    def groupId(self):
        return 'ftth_tools'

    def createInstance(self):
        return FTTHBOMAlgorithm()

    def shortHelpString(self):
        return """
        <h3>FTTH BOM Generator</h3>
        <p>Generates a Bill of Materials (BOM) Excel workbook from FTTH network layers.</p>

        <h4>Required Inputs:</h4>
        <ul>
        <li><b>DJ Polygons</b>: Polygon layer covering each DJ's distribution area</li>
        <li><b>DJ Points</b>: Distribution Joint points (with name, splitter, position, dc_name)</li>
        <li><b>House Points</b>: Customer premises points</li>
        <li><b>Pole Points</b>: Pole locations</li>
        <li><b>DC Lines</b>: Distribution cable lines (with from_node, to_node, length)</li>
        <li><b>FC Lines</b>: Feeder cable lines</li>
        <li><b>FJ Points</b>: Feeder Joint points</li>
        </ul>

        <h4>Parameters:</h4>
        <ul>
        <li><b>Cable Size</b>: Select feeder cable fibre count (12F to 288F)</li>
        <li><b>Splicing Plan CSV</b>: Optional — splicing plan CSV for accurate splice counts</li>
        <li><b>Fibers per FJ</b>: Fibres used per feeder joint (default 4)</li>
        <li><b>Splitters per FJ</b>: 1:2 splitters per FJ (default 2)</li>
        <li><b>Pole Type</b>: 6M / 7M / 9M gum poles</li>
        </ul>

        <h4>Output Sheets:</h4>
        <ul>
        <li><b>Pre-Conventional</b>: DJ boxes, FJ closures, drop cables (fixed length), pigtails, couplers</li>
        <li><b>Conventional</b>: Feeder cable, dome joints, splitters, splice protectors, patch cords</li>
        <li><b>Pole Hardware</b>: Tangent/dead-end clamps, hooks, brackets, mounting assemblies</li>
        <li><b>Consumables</b>: Alcohol, kim wipes, cable ties, tape, seals, bolts/nuts</li>
        <li><b>Other</b>: Poles, duct, GLAM nodes, tools</li>
        <li><b>Summary</b>: Totals per category</li>
        </ul>
        """

    def initAlgorithm(self, config=None):
        # Required vector layers
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_DJ_POLY, 'DJ Polygons (area coverage per DJ)',
            [QgsProcessing.TypeVectorPolygon]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_DJ, 'DJ Points (distribution joints)',
            [QgsProcessing.TypeVectorPoint]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_HOUSES, 'House Points (customer premises)',
            [QgsProcessing.TypeVectorPoint]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_POLES, 'Pole Points',
            [QgsProcessing.TypeVectorPoint]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_DC, 'DC Lines (distribution cables)',
            [QgsProcessing.TypeVectorLine]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_FC, 'FC Lines (feeder cables)',
            [QgsProcessing.TypeVectorLine]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_FJ, 'FJ Points (feeder joints)',
            [QgsProcessing.TypeVectorPoint]
        ))

        # String / numeric parameters
        self.addParameter(QgsProcessingParameterString(
            self.PARAM_AREA, 'Area Code (e.g., VTN_HHS_GMG)',
            defaultValue=''
        ))
        self.addParameter(QgsProcessingParameterString(
            self.PARAM_ZONE, 'Zone Code (optional, e.g., Z02)',
            defaultValue='', optional=True
        ))
        self.addParameter(QgsProcessingParameterEnum(
            self.PARAM_CABLE_SIZE, 'Feeder Cable Size',
            options=['12F', '24F', '48F', '72F', '96F', '144F', '288F'],
            defaultValue=5  # 144F default
        ))
        self.addParameter(QgsProcessingParameterNumber(
            self.PARAM_FIBERS_PER_FJ, 'Fibers per FJ',
            type=QgsProcessingParameterNumber.Integer,
            defaultValue=4, minValue=1, maxValue=24
        ))
        self.addParameter(QgsProcessingParameterNumber(
            self.PARAM_SPLITTERS_PER_FJ, '1:2 Splitters per FJ',
            type=QgsProcessingParameterNumber.Integer,
            defaultValue=2, minValue=0, maxValue=10
        ))
        self.addParameter(QgsProcessingParameterEnum(
            self.PARAM_POLE_TYPE, 'Pole Type',
            options=['6M', '7M', '9M'],
            defaultValue=1  # 7M
        ))
        self.addParameter(QgsProcessingParameterFile(
            self.PARAM_SPLICING_PLAN, 'Splicing Plan CSV (optional)',
            fileFilter='CSV files (*.csv)',
            optional=True
        ))

        # Output
        self.addParameter(QgsProcessingParameterFileDestination(
            self.PARAM_OUTPUT, 'Output BOM Excel (.xlsx)',
            fileFilter='Excel files (*.xlsx)'
        ))

    # ---- Static Helpers ----

    @staticmethod
    def _find_field(layer, field_names):
        """Find field index, case-insensitive. Returns -1 if not found."""
        all_names = [f.name() for f in layer.fields()]
        for fn in field_names:
            fn_lower = fn.lower()
            for actual_name in all_names:
                if actual_name.lower() == fn_lower:
                    return layer.fields().indexFromName(actual_name)
        return -1

    @staticmethod
    def _safe_val(feature, field_idx, default=''):
        """Extract string value from feature field safely."""
        if field_idx < 0:
            return default
        val = feature.attribute(field_idx)
        if val is None:
            return default
        try:
            from qgis.core import QgsVariantUtils
            if QgsVariantUtils.isNull(val):
                return default
        except Exception:
            pass
        return str(val).strip()

    @staticmethod
    def _safe_int(feature, field_idx, default=0):
        """Extract int value from feature field safely."""
        if field_idx < 0:
            return default
        val = feature.attribute(field_idx)
        if val is None:
            return default
        try:
            from qgis.core import QgsVariantUtils
            if QgsVariantUtils.isNull(val):
                return default
        except Exception:
            pass
        try:
            return int(float(str(val).strip()))
        except (ValueError, TypeError):
            return default

    @staticmethod
    def _safe_float(feature, field_idx, default=0.0):
        """Extract float value from feature field safely."""
        if field_idx < 0:
            return default
        val = feature.attribute(field_idx)
        if val is None:
            return default
        try:
            from qgis.core import QgsVariantUtils
            if QgsVariantUtils.isNull(val):
                return default
        except Exception:
            pass
        try:
            return float(str(val).strip())
        except (ValueError, TypeError):
            return default

    # ---- Excel Helpers ----

    def _write_sheet_header(self, ws, title, headers):
        """Write styled header row to a worksheet."""
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = self._HEADER_FILL
            cell.font = self._HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self._THIN_BORDER

    def _write_bom_sheet(self, ws, title, items, feedback):
        """Write a BOM category sheet with all line items."""
        ws.title = title
        headers = ['Item / Part No', 'Description', 'Qty', 'Unit', 'Notes']
        self._write_sheet_header(ws, title, headers)

        if not items:
            ws.append(['', 'No items', 0, '', ''])
            feedback.pushInfo("  Sheet '{}': no items".format(title))
            return

        for item in items:
            ws.append([
                item.get('part_no', ''),
                item.get('description', ''),
                item.get('qty', 0),
                item.get('unit', 'EACH'),
                item.get('notes', ''),
            ])

        # Auto-width columns
        for col_idx in range(1, len(headers) + 1):
            max_len = len(headers[col_idx - 1])
            for row in ws.iter_rows(min_row=2, max_col=len(headers), max_row=ws.max_row):
                cell_val = str(row[col_idx - 1].value) if row[col_idx - 1].value else ''
                max_len = max(max_len, len(cell_val))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)

        # Apply borders to data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = self._THIN_BORDER
                if cell.column == 3:  # Qty column
                    cell.alignment = Alignment(horizontal='center')

        feedback.pushInfo("  Sheet '{}': {} line items".format(title, len(items)))

    def _write_summary_sheet(self, ws, summary, feedback):
        """Write the summary sheet."""
        ws.title = 'Summary'
        headers = ['Category', 'Total Qty', 'Line Items', 'Notes']
        self._write_sheet_header(ws, 'Summary', headers)

        rows = [
            ['Pre-Conventional', summary.get('total_pre_conventional', 0),
             summary.get('line_items_pre_conventional', 0), 'DJ boxes, FJ closures, drop cables, pigtails, couplers'],
            ['Conventional', summary.get('total_conventional', 0),
             summary.get('line_items_conventional', 0), 'Feeder, dome joints, splitters, core cable, ODF, patch cords'],
            ['Pole Hardware', summary.get('total_pole_hardware', 0),
             summary.get('line_items_pole_hardware', 0), 'Clamps, hooks, brackets, bandit straps'],
            ['Consumables', summary.get('total_consumables', 0),
             summary.get('line_items_consumables', 0), 'Bunny clips, ties, tape, cleaning supplies'],
            ['', '', '', ''],
            ['GRAND TOTAL', summary.get('grand_total', 0),
             summary.get('grand_line_items', 0), 'All categories combined'],
        ]

        for row_data in rows:
            ws.append(row_data)

        # Style grand total row
        grand_total_row = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=grand_total_row, column=col)
            cell.fill = self._SUBHEADER_FILL
            cell.font = self._SUBHEADER_FONT
            cell.border = self._THIN_BORDER

        # Auto-width
        for col_idx in range(1, len(headers) + 1):
            max_len = len(headers[col_idx - 1])
            for row in ws.iter_rows(min_row=2, max_col=len(headers), max_row=ws.max_row):
                cell_val = str(row[col_idx - 1].value) if row[col_idx - 1].value else ''
                max_len = max(max_len, len(cell_val))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)

        # Apply borders
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, max_col=len(headers)):
            for cell in row:
                cell.border = self._THIN_BORDER
                if cell.column in (2, 3):
                    cell.alignment = Alignment(horizontal='center')

        feedback.pushInfo("  Sheet 'Summary': grand total = {} ({} line items)".format(
            summary.get('grand_total', 0), summary.get('grand_line_items', 0)))

    def _build_bom_excel(self, bom_result, output_path, area_code, zone, feedback):
        """Build the multi-sheet BOM Excel workbook."""
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Sheet 1: Pre-Conventional
        ws1 = wb.create_sheet()
        self._write_bom_sheet(ws1, 'Pre-Conventional',
                              bom_result.get('pre_conventional', []), feedback)

        # Sheet 2: Conventional
        ws2 = wb.create_sheet()
        self._write_bom_sheet(ws2, 'Conventional',
                              bom_result.get('conventional', []), feedback)

        # Sheet 3: Pole Hardware
        ws3 = wb.create_sheet()
        self._write_bom_sheet(ws3, 'Pole Hardware',
                              bom_result.get('pole_hardware', []), feedback)

        # Sheet 4: Consumables
        ws4 = wb.create_sheet()
        self._write_bom_sheet(ws4, 'Consumables',
                              bom_result.get('consumables', []), feedback)

        # Sheet 5: Other
        ws5 = wb.create_sheet()
        self._write_bom_sheet(ws5, 'Other',
                              bom_result.get('other', []), feedback)

        # Sheet 6: Summary
        ws6 = wb.create_sheet()
        self._write_summary_sheet(ws6, bom_result.get('summary', {}), feedback)

        # Add metadata to summary
        ws6.insert_rows(1)
        ws6.merge_cells('A1:D1')
        title_cell = ws6['A1']
        zone_str = "_{}".format(zone) if zone else ""
        title_cell.value = "FTTH Bill of Materials — {}{}".format(area_code, zone_str)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')

        wb.save(output_path)

    # ---- Main Processing ----

    def processAlgorithm(self, parameters, context, feedback):
        feedback.pushInfo("=" * 60)
        feedback.pushInfo("FTTH BOM Generator")
        feedback.pushInfo("=" * 60)

        # Read parameters
        area_code = self.parameterAsString(parameters, self.PARAM_AREA, context)
        zone = self.parameterAsString(parameters, self.PARAM_ZONE, context)
        cable_size_idx = self.parameterAsInt(parameters, self.PARAM_CABLE_SIZE, context)
        cable_sizes = ['12F', '24F', '48F', '72F', '96F', '144F', '288F']
        cable_size = cable_sizes[cable_size_idx] if 0 <= cable_size_idx < len(cable_sizes) else '144F'
        fibers_per_fj = self.parameterAsInt(parameters, self.PARAM_FIBERS_PER_FJ, context)
        splitters_per_fj = self.parameterAsInt(parameters, self.PARAM_SPLITTERS_PER_FJ, context)
        pole_type_idx = self.parameterAsInt(parameters, self.PARAM_POLE_TYPE, context)
        pole_types = ['6M', '7M', '9M']
        pole_type = pole_types[pole_type_idx] if 0 <= pole_type_idx < len(pole_types) else '7M'
        splicing_plan_csv = self.parameterAsString(parameters, self.PARAM_SPLICING_PLAN, context)
        output_path = self.parameterAsString(parameters, self.PARAM_OUTPUT, context)

        feedback.pushInfo("Area Code      : {}".format(area_code))
        feedback.pushInfo("Zone           : {}".format(zone or '(none)'))
        feedback.pushInfo("Cable Size     : {}".format(cable_size))
        feedback.pushInfo("Fibers per FJ  : {}".format(fibers_per_fj))
        feedback.pushInfo("Splitters/FJ   : {}".format(splitters_per_fj))
        feedback.pushInfo("Pole Type      : {}".format(pole_type))
        if splicing_plan_csv:
            feedback.pushInfo("Splicing Plan  : {}".format(splicing_plan_csv))
        feedback.pushInfo("")

        # Read layers
        dj_poly_layer = self.parameterAsVectorLayer(parameters, self.IN_DJ_POLY, context)
        dj_layer = self.parameterAsVectorLayer(parameters, self.IN_DJ, context)
        houses_layer = self.parameterAsVectorLayer(parameters, self.IN_HOUSES, context)
        poles_layer = self.parameterAsVectorLayer(parameters, self.IN_POLES, context)
        dc_layer = self.parameterAsVectorLayer(parameters, self.IN_DC, context)
        fc_layer = self.parameterAsVectorLayer(parameters, self.IN_FC, context)
        fj_layer = self.parameterAsVectorLayer(parameters, self.IN_FJ, context)

        # ================================================================
        # Step 1: Read all layers and index
        # ================================================================
        feedback.pushInfo("[Step 1] Reading input layers...")

        # -- DJ Points --
        dj_name_idx = self._find_field(dj_layer, ['name', 'Name', 'NAME'])
        dj_splitter_idx = self._find_field(dj_layer, ['splitter', 'Splitter', 'SPLITTER'])
        dj_position_idx = self._find_field(dj_layer, ['position', 'Position', 'POSITION'])
        dj_dc_name_idx = self._find_field(dj_layer, ['dc_name', 'dcName', 'DC_NAME'])
        dj_pole_idx = self._find_field(dj_layer, ['pole', 'pole_name', 'Pole'])
        dj_fj_idx = self._find_field(dj_layer, ['fj_name', 'fjName', 'FJ_NAME'])
        dj_block_idx = self._find_field(dj_layer, ['block', 'block_name', 'Block'])

        djs = {}
        for feat in dj_layer.getFeatures():
            name = self._safe_val(feat, dj_name_idx)
            if not name:
                continue
            splitter = self._safe_val(feat, dj_splitter_idx)
            # Auto-detect splitter from DJ name if field is empty
            if not splitter and '_1:8' in name:
                splitter = '1:8'
            elif not splitter and '_1:9' in name:
                splitter = '1:9'
            position = self._safe_int(feat, dj_position_idx)
            # Auto-detect position from DJ name if field is empty
            if not position:
                import re as _re
                m = _re.search(r'pos(\d)', name, _re.IGNORECASE)
                if m:
                    position = int(m.group(1))
            dc_name = self._safe_val(feat, dj_dc_name_idx)
            pole = self._safe_val(feat, dj_pole_idx)
            fj_name = self._safe_val(feat, dj_fj_idx)
            block = self._safe_val(feat, dj_block_idx)
            djs[name] = {
                'position': position,
                'splitter': splitter,
                'block': block,
                'pole': pole,
                'fj_name': fj_name,
                'dc_name': dc_name,
            }
        feedback.pushInfo("  DJ Points   : {} DJs read".format(len(djs)))

        # -- DJ Polygons --
        poly_name_idx = self._find_field(dj_poly_layer, ['name', 'Name', 'NAME'])
        poly_block_idx = self._find_field(dj_poly_layer, ['block', 'block_name', 'Block'])
        poly_dj_idx = self._find_field(dj_poly_layer, ['dj_name', 'djName', 'DJ_NAME'])

        dj_polygons = []
        poly_to_dj = {}  # fid -> dj_name
        for feat in dj_poly_layer.getFeatures():
            fid = feat.id()
            name = self._safe_val(feat, poly_name_idx)
            block = self._safe_val(feat, poly_block_idx)
            dj_name = self._safe_val(feat, poly_dj_idx)
            geom = feat.geometry()
            if geom and not geom.isEmpty():
                centroid = geom.centroid().asPoint()
            else:
                centroid = None
            dj_polygons.append({
                'fid': fid, 'name': name, 'block': block,
                'dj_name': dj_name, 'centroid': centroid,
            })
            if dj_name:
                poly_to_dj[fid] = dj_name
        feedback.pushInfo("  DJ Polygons : {} polygons read".format(len(dj_polygons)))

        # -- Houses --
        house_name_idx = self._find_field(houses_layer, ['name', 'Name', 'NAME'])
        houses = {}
        house_geoms = {}
        for feat in houses_layer.getFeatures():
            name = self._safe_val(feat, house_name_idx)
            if not name:
                continue
            geom = feat.geometry()
            if geom and not geom.isEmpty():
                pt = geom.asPoint()
                house_geoms[name] = pt
            houses[name] = name
        feedback.pushInfo("  Houses      : {} houses read".format(len(houses)))

        # -- Poles --
        pole_name_idx = self._find_field(poles_layer, ['name', 'Name', 'NAME'])
        poles = {}
        for feat in poles_layer.getFeatures():
            name = self._safe_val(feat, pole_name_idx)
            if not name:
                continue
            poles[name] = {'cables': [], 'has_dj': False, 'has_fj': False}
        feedback.pushInfo("  Poles       : {} poles read".format(len(poles)))

        # -- DC Lines --
        dc_name_idx = self._find_field(dc_layer, ['name', 'Name', 'NAME'])
        dc_from_idx = self._find_field(dc_layer, ['from_node', 'fromNode', 'from', 'from_name'])
        dc_to_idx = self._find_field(dc_layer, ['to_node', 'toNode', 'to', 'to_name'])
        dc_length_idx = self._find_field(dc_layer, ['length', 'Length', 'LENGTH'])

        dc_lines = []
        dc_lengths = {}
        for feat in dc_layer.getFeatures():
            name = self._safe_val(feat, dc_name_idx)
            if not name:
                continue
            from_node = self._safe_val(feat, dc_from_idx)
            to_node = self._safe_val(feat, dc_to_idx)
            length = self._safe_int(feat, dc_length_idx)
            if not length:
                # Compute from geometry
                geom = feat.geometry()
                if geom and not geom.isEmpty():
                    length = int(round(geom.length()))
                else:
                    length = 50
            dc_lines.append({
                'name': name, 'from_node': from_node,
                'to_node': to_node, 'length': length,
            })
            dc_lengths[name] = {'length': length, 'block': ''}
        feedback.pushInfo("  DC Lines    : {} segments read".format(len(dc_lines)))

        # -- FC Lines --
        fc_name_idx = self._find_field(fc_layer, ['name', 'Name', 'NAME'])
        fc_length_idx = self._find_field(fc_layer, ['length', 'Length', 'LENGTH'])
        fc_size_idx = self._find_field(fc_layer, ['size', 'cable_size', 'Size', 'fiber_count'])

        fc_lengths = {}
        for feat in fc_layer.getFeatures():
            name = self._safe_val(feat, fc_name_idx)
            if not name:
                continue
            length = self._safe_int(feat, fc_length_idx)
            if not length:
                geom = feat.geometry()
                if geom and not geom.isEmpty():
                    length = int(round(geom.length()))
                else:
                    length = 0
            size = self._safe_val(feat, fc_size_idx) or '144F'
            fc_lengths[name] = {'length': length, 'size': size}
        feedback.pushInfo("  FC Lines    : {} cables read".format(len(fc_lengths)))

        # -- FJ Points --
        fj_name_idx = self._find_field(fj_layer, ['name', 'Name', 'NAME'])
        fj_ag_idx = self._find_field(fj_layer, ['ag', 'ag_name', 'AG', 'AG_NAME'])
        fj_pole_idx = self._find_field(fj_layer, ['pole', 'pole_name', 'Pole'])

        fjs = {}
        for feat in fj_layer.getFeatures():
            name = self._safe_val(feat, fj_name_idx)
            if not name:
                continue
            ag = self._safe_val(feat, fj_ag_idx)
            pole = self._safe_val(feat, fj_pole_idx)
            # Parse AG from FJ name if field is empty
            if not ag:
                import re as _re
                parts = name.split('_')
                for p in parts:
                    if p.startswith('AG') or p.startswith('ag'):
                        ag = p.upper()
                        break
            fjs[name] = {'ag': ag, 'pole': pole, 'dj_names': []}
        feedback.pushInfo("  FJ Points   : {} FJs read".format(len(fjs)))

        # ================================================================
        # Step 2: Build topology and associate entities
        # ================================================================
        feedback.pushInfo("")
        feedback.pushInfo("[Step 2] Building topology...")

        # Associate houses to nearest DJ (spatial join)
        dj_centroids = {}
        for poly in dj_polygons:
            if poly['centroid'] and poly['dj_name']:
                dj_centroids[poly['dj_name']] = poly['centroid']
            elif poly['centroid']:
                dj_centroids[poly['name']] = poly['centroid']

        # Fallback: use DJ point geometries
        for feat in dj_layer.getFeatures():
            name = self._safe_val(feat, dj_name_idx)
            if name and name not in dj_centroids:
                geom = feat.geometry()
                if geom and not geom.isEmpty():
                    dj_centroids[name] = geom.asPoint()

        houses_by_dj = {}
        for dj_name in djs:
            houses_by_dj[dj_name] = []

        unmatched_houses = 0
        for h_name, h_pt in house_geoms.items():
            nearest_dj = None
            nearest_dist = float('inf')
            for dj_name, dj_pt in dj_centroids.items():
                dist = h_pt.sqrDist(dj_pt)
                if dist < nearest_dist:
                    nearest_dist = dist
                    nearest_dj = dj_name
            if nearest_dj and nearest_dj in houses_by_dj:
                houses_by_dj[nearest_dj].append(h_name)
            else:
                unmatched_houses += 1

        if unmatched_houses:
            feedback.pushInfo("  Houses not matched to DJ: {}".format(unmatched_houses))

        # Build block mapping
        blocks = {}
        for dj_name, dj_info in djs.items():
            block = dj_info.get('block', '')
            if not block:
                # Try to derive block from DC name
                dc_name = dj_info.get('dc_name', '')
                if dc_name:
                    import re as _re
                    m = _re.search(r'DC(\d{3})', dc_name)
                    if m:
                        block_num = int(m.group(1))
                        block = "B{:03d}".format(block_num)
            if not block:
                block = "UNKNOWN"
            dj_info['block'] = block
            if block not in blocks:
                # Derive FJ and AG from DJ info
                fj_name = dj_info.get('fj_name', '')
                ag = ''
                if fj_name and fj_name in fjs:
                    ag = fjs[fj_name].get('ag', '')
                elif block and block.startswith('B'):
                    # Try to match FJ by AG
                    pass
                blocks[block] = {'num_djs': 0, 'fj_name': fj_name, 'ag': ag}
            blocks[block]['num_djs'] += 1

        # Update FJ -> DJ associations
        for dj_name, dj_info in djs.items():
            fj_name = dj_info.get('fj_name', '')
            if fj_name and fj_name in fjs:
                if dj_name not in fjs[fj_name]['dj_names']:
                    fjs[fj_name]['dj_names'].append(dj_name)

        # Build AG mapping
        ags = {}
        for block_name, block_info in blocks.items():
            ag = block_info.get('ag', '')
            if not ag:
                ag = 'UNKNOWN'
            if ag not in ags:
                ags[ag] = {'num_blocks': 0, 'num_fjs': 0, 'num_djs': 0}
            ags[ag]['num_blocks'] += 1
            ags[ag]['num_djs'] += block_info['num_djs']

        for fj_name, fj_info in fjs.items():
            ag = fj_info.get('ag', '')
            if not ag:
                ag = 'UNKNOWN'
            if ag not in ags:
                ags[ag] = {'num_blocks': 0, 'num_fjs': 0, 'num_djs': 0}
            ags[ag]['num_fjs'] += 1

        # Update pole flags
        for dj_name, dj_info in djs.items():
            pole = dj_info.get('pole', '')
            if pole and pole in poles:
                poles[pole]['has_dj'] = True

        for fj_name, fj_info in fjs.items():
            pole = fj_info.get('pole', '')
            if pole and pole in poles:
                poles[pole]['has_fj'] = True

        # Derive DC -> block associations
        for dc_name, dc_info in dc_lengths.items():
            for dj_name, dj_info in djs.items():
                if dj_info.get('dc_name', '') == dc_name:
                    dc_info['block'] = dj_info.get('block', '')
                    dc_info['dj_name'] = dj_name
                    break

        # Populate pole cables from FC lines (simplified: each FC is a cable)
        for fc_name, fc_info in fc_lengths.items():
            # Find poles along FC route via spatial proximity (simplified)
            pass  # Will be handled by bom_engine's pole_hardware calc

        feedback.pushInfo("  Blocks      : {}".format(len(blocks)))
        feedback.pushInfo("  AGs         : {}".format(len(ags)))
        feedback.pushInfo("  DJs→Houses  : {} total associations".format(
            sum(len(v) for v in houses_by_dj.values())))

        # ================================================================
        # Step 3: Build data dict and call BOM engine
        # ================================================================
        feedback.pushInfo("")
        feedback.pushInfo("[Step 3] Calculating BOM...")

        # Load splicing plan data if provided
        splicing_data = []
        if splicing_plan_csv:
            try:
                splicing_data = bom_engine.load_splicing_plan(splicing_plan_csv)
                feedback.pushInfo("  Splicing plan: {} rows loaded".format(len(splicing_data)))
            except Exception as e:
                feedback.pushInfo("  WARNING: Could not load splicing plan: {}".format(str(e)))

        # Build DC segments list for fixed-length calculation
        dc_segments = []
        for dc_name, dc_info in dc_lengths.items():
            dc_segments.append({
                'dc_name': dc_name,
                'length_m': dc_info.get('length', 50),
                'block': dc_info.get('block', ''),
                'dj_name': dc_info.get('dj_name', ''),
            })

        # Build FC info list
        fc_info_list = []
        for fc_name, fc_info in fc_lengths.items():
            fc_info_list.append({
                'fc_name': fc_name,
                'length_m': fc_info.get('length', 0),
                'size': fc_info.get('size', cable_size),
            })

        data = {
            'djs': djs,
            'fjs': fjs,
            'blocks': blocks,
            'ags': ags,
            'poles': poles,
            'houses': houses_by_dj,
            'dc_lengths': dc_lengths,
            'dc_segments': dc_segments,
            'fc_info': fc_info_list,
            'fc_lengths': fc_lengths,
            'splicing_plan': splicing_data,
        }

        config = {
            'cable_size': cable_size,
            'fibers_per_fj': fibers_per_fj,
            'splitters_per_fj': splitters_per_fj,
            'pole_type': pole_type,
            'route_type': 'aerial',
            'core_length_per_ag_km': 2.0,
        }

        try:
            bom_result = calculate_bom(data, config)
        except Exception as e:
            feedback.reportError("  [ERROR] BOM calculation failed: {}".format(e))
            import traceback
            feedback.reportError(traceback.format_exc())
            return {self.PARAM_OUTPUT: output_path}

        feedback.pushInfo("  BOM calculated successfully")
        feedback.pushInfo("  Pre-conventional : {} items".format(
            len(bom_result.get('pre_conventional', []))))
        feedback.pushInfo("  Conventional     : {} items".format(
            len(bom_result.get('conventional', []))))
        feedback.pushInfo("  Pole hardware    : {} items".format(
            len(bom_result.get('pole_hardware', []))))
        feedback.pushInfo("  Consumables      : {} items".format(
            len(bom_result.get('consumables', []))))

        # ================================================================
        # Step 4: Generate Excel BOM
        # ================================================================
        feedback.pushInfo("")
        feedback.pushInfo("[Step 4] Generating Excel BOM...")

        if not HAS_OPENPYXL:
            feedback.reportError("  ERROR: openpyxl not available. Cannot generate Excel.")
            feedback.reportError("  Install: python -m pip install openpyxl")
            return {self.PARAM_OUTPUT: output_path}

        try:
            self._build_bom_excel(bom_result, output_path, area_code, zone, feedback)
            feedback.pushInfo("  [OK] BOM Excel written: {}".format(output_path))
        except Exception as e:
            feedback.reportError("  [ERROR] Building Excel: {}".format(e))
            import traceback
            feedback.reportError(traceback.format_exc())
            return {self.PARAM_OUTPUT: output_path}

        feedback.pushInfo("")
        feedback.pushInfo("=" * 60)
        feedback.pushInfo("BOM GENERATION COMPLETE")
        feedback.pushInfo("  DJs processed    : {}".format(len(djs)))
        feedback.pushInfo("  FJs processed    : {}".format(len(fjs)))
        feedback.pushInfo("  Blocks           : {}".format(len(blocks)))
        feedback.pushInfo("  AGs              : {}".format(len(ags)))
        feedback.pushInfo("  Houses           : {}".format(len(houses)))
        feedback.pushInfo("  Poles            : {}".format(len(poles)))
        feedback.pushInfo("  DC segments      : {}".format(len(dc_lines)))
        feedback.pushInfo("  FC cables        : {}".format(len(fc_lengths)))
        feedback.pushInfo("  Output file      : {}".format(output_path))

        return {self.PARAM_OUTPUT: output_path}


# =============================================================================
# END OF FILE
# =============================================================================