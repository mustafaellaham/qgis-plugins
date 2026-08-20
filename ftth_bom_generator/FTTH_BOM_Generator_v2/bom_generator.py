# -*- coding: utf-8 -*-
"""FTTH BOM Generator v2.7.1 — pole snap tolerance; Save-As dialog, template pool export, no-bracket warnings, CRS-safe."""

import os, json, math, re
from datetime import datetime
from collections import defaultdict
from qgis.PyQt.QtWidgets import QAction, QMessageBox
from qgis.core import (QgsProject, QgsGeometry, QgsPointXY,
                       QgsCoordinateTransform, QgsCoordinateReferenceSystem)
from qgis.utils import iface
import openpyxl

from .wizard_dialog import WizardDialogV2b
from .layer_mapping_dialog import LayerMappingDialog

PRECON_CABLES = [50, 80, 100, 120, 150, 180, 200, 250, 300, 350]
SLACK_VALUES = [10, 9, 8, 7]

# ── Helpers ──
def _pf(name):
    m = re.search(r'(\d+)F', name); return m.group(1) if m else '?'
def _height(desc):
    """Pole height in meters. Accepts HTML balloon ('<td>9m</td>'),
    plain number ('6' / '6m' / ' 7 '), or text containing '9m'."""
    if not desc: return 0
    m = re.search(r'<td>(\d+)m</td>', desc)
    if m: return int(m.group(1))
    s = str(desc).strip()
    m = re.fullmatch(r'(\d+(?:\.\d+)?)\s*m?', s)
    if m: return int(float(m.group(1)))
    m = re.search(r'(\d+)\s*m\b', s)
    return int(m.group(1)) if m else 0
def _bracket(od, table):
    for entry in table:
        lo, hi = entry['range'][0], entry['range'][1]
        if lo <= od <= hi: return f"{lo}-{hi}"
    return None
def _best_precon(length):
    best = None
    for sl in SLACK_VALUES:
        need = length + sl
        for c in PRECON_CABLES:
            if c >= need:
                w = c - need
                if best is None or w < best[2]: best = (sl, c, w)
                break
    return best if best else (None, None, None)

def _pick_work_crs(layers):
    """Return a metric CRS for all geometry math.
    Uses the project CRS when it is projected (meters); otherwise derives the
    correct UTM zone from the data extent center (works for EPSG:4326 layers)."""
    pcrs = QgsProject.instance().crs()
    if pcrs.isValid() and not pcrs.isGeographic():
        return pcrs
    wgs = QgsCoordinateReferenceSystem("EPSG:4326")
    lon = lat = None
    for l in layers:
        if not l: continue
        ext = l.extent()
        if ext.isEmpty(): continue
        try:
            xt = QgsCoordinateTransform(l.crs(), wgs, QgsProject.instance())
            c = xt.transform(ext.center())
            lon, lat = c.x(), c.y()
            break
        except Exception:
            continue
    if lon is None:
        return wgs  # no usable extent — fall back (legacy behavior)
    zone = int((lon + 180.0) / 6.0) + 1
    epsg = (32700 if lat < 0 else 32600) + zone
    return QgsCoordinateReferenceSystem(f"EPSG:{epsg}")


class BOMGeneratorV2b:
    def __init__(self, iface):
        self.iface = iface
        self.plugin_dir = os.path.dirname(__file__)
        self.config = self.load_config()
        self.layer_map = {}
        self.wiz = {}
        self.v = {}
        self.warnings = []
        self._work_crs = None
        self._xforms = {}
        self._nb_seen = set()
        self.pole_tol = float(self.config.get('_settings', {}).get('pole_snap_tolerance_m', 5.0))

    def load_config(self):
        with open(os.path.join(self.plugin_dir, 'bom_rules.json'), 'r') as f:
            return json.load(f)

    def initGui(self):
        self.action = QAction("Generate BOM", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.iface.addPluginToVectorMenu("FTTH BOM Generator", self.action)

    def unload(self):
        self.iface.removePluginVectorMenu("FTTH BOM Generator", self.action)

    def _L(self, role): return self.layer_map.get(role)
    def _FC(self, role):
        l = self._L(role); return l.featureCount() if l else 0

    # ════════════════════════════════════════════════════════
    #  CRS-SAFE GEOMETRY — all distance/intersect/length math
    #  is done in a metric working CRS (meters).
    # ════════════════════════════════════════════════════════
    def _init_geo(self):
        layers = [l for l in self.layer_map.values() if l]
        self._work_crs = _pick_work_crs(layers)
        self._xforms = {}
        pcrs = QgsProject.instance().crs()
        if pcrs.isGeographic() or any(l.crs() != self._work_crs for l in layers):
            self.warnings.append(
                f"CRS NOTE: geometry math reprojected to {self._work_crs.authid()} (meters)")

    def _g(self, layer, geom):
        """Transform a geometry into the metric working CRS."""
        if geom is None or geom.isNull(): return geom
        if not self._work_crs or not self._work_crs.isValid(): return geom
        crs = layer.crs()
        if crs == self._work_crs: return geom
        key = crs.authid()
        xt = self._xforms.get(key)
        if xt is None:
            xt = QgsCoordinateTransform(crs, self._work_crs, QgsProject.instance())
            self._xforms[key] = xt
        g = QgsGeometry(geom)
        g.transform(xt)
        return g

    # ════════════════════════════════════════════════════════
    #  NO-BRACKET WARNING — OD with no matching hardware bracket
    #  must surface as ITEM NOT FOUND, never skip silently.
    # ════════════════════════════════════════════════════════
    def _nb_warn(self, kind, fiber, od):
        key = (kind, fiber, od)
        if key in self._nb_seen: return
        self._nb_seen.add(key)
        self.warnings.append(f"ITEM NOT FOUND: no {kind} bracket for OD {od}mm ({fiber}F)")

    # ════════════════════════════════════════════════════════
    #  get_od — SINGLE SOURCE OF TRUTH
    #  Uses _cable_options list. Returns full option dict.
    # ════════════════════════════════════════════════════════
    def _get_od_full(self, cable_name, fiber):
        """Return the full cable option dict from _cable_options matching wizard selection."""
        opts = self.config.get('_cable_options', [])
        fiber_opts = [o for o in opts if o['fiber'] == fiber]
        if not fiber_opts: return {}

        # 1) Wizard selection (index stored in wiz['cable_option_indices'])
        indices = (self.wiz.get('cable_option_indices') or {})
        idx = indices.get(f"{fiber}||{cable_name}")
        if idx is not None and 0 <= idx < len(fiber_opts):
            return fiber_opts[idx]

        # 2) Auto-detect from cable name
        n = cable_name.lower()
        for i, o in enumerate(fiber_opts):
            if o['type'].lower() in n:
                return o

        # 3) Fallback: first option
        return fiber_opts[0]

    def _get_od(self, cable_name, fiber):
        """Return (type, od_mm) shortcut."""
        o = self._get_od_full(cable_name, fiber)
        return o.get('type','Slim'), o.get('od', 0)

    # ════════════════════════════════════════════════════════
    #  Cable length from name (YYYY) + 5%
    # ════════════════════════════════════════════════════════
    @staticmethod
    def _name_length(name):
        """Return YYYY from (XXXXm-YYYYm) pattern, or None."""
        m = re.search(r'\((\d+)m\s*-\s*(\d+)m\)', name)
        return float(m.group(2)) if m else None

    @staticmethod
    def _name_lengths(name):
        """Return (XXXX, YYYY) from (XXXXm-YYYYm) pattern, or (None, None)."""
        m = re.search(r'\((\d+)m\s*-\s*(\d+)m\)', name)
        return (float(m.group(1)), float(m.group(2))) if m else (None, None)

    # ════════════════════════════════════════════════════════
    #  MAIN PIPELINE (scan → wizard → hardware ALL)
    # ════════════════════════════════════════════════════════
    def run(self):
        self.warnings = []; self.v = {}

        dlg_map = LayerMappingDialog(self.iface)
        if not dlg_map.exec_(): return
        self.layer_map = dlg_map.get_mapping()
        self._init_geo()

        # Phase 1: base counts + cable scan (lengths only)
        self._base_counts()
        self._cable_scan()
        self._dj_parse()

        # Phase 2: Wizard
        dlg = WizardDialogV2b(self.iface, self, self.config)
        if not dlg.exec_(): return
        self.wiz = dlg.get_answers() or {}

        # Phase 3: ALL hardware AFTER wizard (uses _get_od)
        self._hardware_all()
        self._dist_accessories()
        self._pop_detect()

        # Phase 4: depended + precon
        self._depended()
        self.v['precon_counts'], _ = self._precon()

        summary = self._summary()
        if self.warnings:
            summary += "\n\nWARNINGS:\n" + "\n".join(f"  {w}" for w in self.warnings)
        reply = QMessageBox.question(self.iface.mainWindow(), "Summary", summary,
                                      QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            self._export()

    # ════════════════════════════════════════════════════════
    #  PHASE 1a: BASE COUNTS
    # ════════════════════════════════════════════════════════
    def _base_counts(self):
        v = self.v
        v['blocks'] = self._FC('Block polygons')
        v['core_aggs'] = self._FC('Core_Aggregation')
        v['feeder_joints'] = self._FC('Feeder_aggregation')
        v['link_joints'] = self._FC('Link_Aggregation')
        v['dj_count'] = self._FC('Distributions joints')
        v['premises'] = self._FC('Premise data')
        v['agg_polygons'] = self._FC('Aggregation_polygons')

        ph = defaultdict(int)
        l = self._L('Poles')
        if l:
            for f in l.getFeatures():
                h = _height(f['description'] or '')
                if h: ph[h] += 1
        v['poles_6m']=ph.get(6,0); v['poles_7m']=ph.get(7,0); v['poles_9m']=ph.get(9,0)
        v['total_poles']=sum(ph.values())
        v['bare_1_2']=math.ceil(v['blocks']/2)
        v['bandit_strap']=math.ceil(v['total_poles']*0.101)
        v['bandit_buckle']=math.ceil(v['total_poles']*0.061)
        v['item_52345294']=v['bare_1_2']

    # ════════════════════════════════════════════════════════
    #  PHASE 1b: CABLE SCAN (YYYY name-length, no hardware)
    # ════════════════════════════════════════════════════════
    def _cable_scan(self):
        cable_lengths_raw = defaultdict(float)
        cable_names = defaultdict(list)

        for role_key in ['core_cable','Feeder','links','Distribution']:
            l = self._L(role_key)
            if not l: continue
            for f in l.getFeatures():
                g = f.geometry()
                if not g or g.isNull(): continue
                name = f['Name'] or ''; fiber = _pf(name)
                if fiber in ('1','?'): continue
                yyyy = self._name_length(name)
                if yyyy:
                    cable_lengths_raw[fiber] += yyyy
                    cable_names[fiber].append((name, role_key, yyyy))
                else:
                    self.warnings.append(f"No name pattern: {name[:60]}")

        # Apply 5% margin: Qty = ROUNDUP(SUM * 1.05)
        cable_qty = {}
        for fiber, raw in cable_lengths_raw.items():
            cable_qty[fiber] = math.ceil(raw * 1.05)

        self.v['cable_lengths_raw'] = dict(cable_lengths_raw)
        self.v['cable_lengths'] = cable_qty
        self.v['cable_names_by_fiber'] = dict(cable_names)

    # ════════════════════════════════════════════════════════
    #  PHASE 1c: DJ PARSING
    # ════════════════════════════════════════════════════════
    def _dj_parse(self):
        v = self.v
        v['dj_1_8']=v['dj_1_9']=v['m8']=v['m16']=0
        l = self._L('Distributions joints')
        if l:
            has_name = l.fields().indexOf('name') != -1
            for f in l.getFeatures():
                desc = f['description'] or ''
                name = ''
                if has_name:
                    try: name = f['name'] or ''
                    except Exception: name = ''
                # --- 1:8 / 1:9 detection ---
                # Priority 1: HTML balloon <td>Type</td><td>8|9</td> in description
                # Priority 2: pattern 1:8 / 1:9 (also 1x8, 1X9, 1/8 ...) in name OR plain description
                t = None
                m = re.search(r'<td>Type</td>\s*<td>(\d+)</td>', desc)
                if m:
                    t = int(m.group(1))
                else:
                    m2 = (re.search(r'1\s*[:xX/]\s*([89])(?!\d)', name)
                          or re.search(r'1\s*[:xX/]\s*([89])(?!\d)', desc))
                    if m2:
                        t = int(m2.group(1))
                if t == 8: v['dj_1_8'] += 1
                elif t == 9: v['dj_1_9'] += 1
                # --- M8 / M16: check BOTH description and name ---
                d = (desc + ' ' + name).lower()
                if 'm8' in d: v['m8']+=1
                if 'm16' in d: v['m16']+=1

    # ════════════════════════════════════════════════════════
    #  PHASE 3: ALL HARDWARE (dead-ends, tangents, hooks, oval, gland)
    #  Uses _get_od_full() which respects wizard choices + oval/circular ranges.
    # ════════════════════════════════════════════════════════
    def _hardware_all(self):
        brackets = self.config.get('_od_brackets', {})

        # Load poles once (transformed to metric working CRS)
        pole_geoms = []
        l = self._L('Poles')
        if l:
            for f in l.getFeatures():
                g = f.geometry()
                if g and not g.isNull(): pole_geoms.append((f['Name'] or '', self._g(l, g)))

        # Collect conventional cables (transformed to metric working CRS)
        # OWNER RULE: Distribution layer is scanned too — any cable >1F
        # (12F/24F/48F...) is a conventional cable and follows the normal
        # calculation scenario (dead-ends, tangents, hooks, gland/oval, audit).
        conventional = []
        for role in ['core_cable','Feeder','links','Distribution']:
            l = self._L(role)
            if not l: continue
            for f in l.getFeatures():
                g = f.geometry()
                if not g or g.isNull(): continue
                name = f['Name'] or ''; fiber = _pf(name)
                if fiber in ('1','?'): continue
                opt = self._get_od_full(name, fiber)
                od = opt.get('od', 0)
                if od == 0:
                    self.warnings.append(f"ITEM NOT FOUND: no cable option for {name[:60]} ({fiber}F)")
                    continue
                conventional.append((name, fiber, opt, od, self._g(l, g), role))

        # ── Dead ends + tangents + hooks ──
        dead_ends = defaultdict(int)
        tangents = defaultdict(int)
        hook_count = 0
        tol = self.pole_tol
        self.v['pole_snap_tolerance_m'] = tol

        # feeder/link aggregation points for per-cable pass counting (audit table)
        agg_geoms = []
        for role_a in ['Feeder_aggregation', 'Link_Aggregation']:
            l_a = self._L(role_a)
            if not l_a: continue
            for f in l_a.getFeatures():
                ga = f.geometry()
                if ga and not ga.isNull():
                    ga = self._g(l_a, ga)
                    if ga.type() == 2: ga = ga.centroid()
                    agg_geoms.append(ga)

        cable_audit = []          # per-cable audit rows for Calculation_Audit sheet
        audit_by_name = {}

        for name, fiber, opt, od, g, role in conventional:
            xxxx, yyyy = self._name_lengths(name)
            # OWNER RULE (2026-08-20): 'atual length' = MEASURED geometry length
            # (meters, metric working CRS). Never trust the name — a planner
            # typo must not flow silently into the audit.
            measured = g.length()
            rec = {'name': name, 'fiber': fiber, 'od': od,
                   'actual': int(round(measured)),
                   'published': (int(yyyy * 1.05) if yyyy is not None else None),
                   'buffer': None,
                   'poles': 0, 'hooks': 0, 'agg_passes': 0, 'turns_gt15': 0,
                   'deadend_qty': 0, 'deadend_size': '',
                   'tangent_qty': 0, 'tangent_size': '',
                   'gland_oval': '', 'gland_oval_size': ''}
            # Typo guard: flag big disagreement between name XXXX and measurement
            if xxxx and measured > 0:
                diff = abs(measured - xxxx)
                if diff > 20 and diff / xxxx > 0.10:
                    self.warnings.append(
                        f"LENGTH MISMATCH: {name[:55]} — name says {int(xxxx)}m, "
                        f"measured {int(round(measured))}m (possible name typo)")
            if rec['actual'] is not None and rec['published'] is not None:
                rec['buffer'] = rec['published'] - rec['actual']
            cable_audit.append(rec); audit_by_name[name] = rec

            ordered = []
            for pn, pg in pole_geoms:
                if g.distance(pg) <= tol:  # snap tolerance: exact touch OR drawn-offset pass
                    nearest = g.nearestPoint(pg.centroid())
                    ordered.append((pn, pg, g.lineLocatePoint(nearest)))
            ordered.sort(key=lambda x: x[2]); n = len(ordered)
            rec['agg_passes'] = sum(1 for ag in agg_geoms if g.distance(ag) <= tol)
            if n == 0: continue
            hook_count += n
            rec['poles'] = n; rec['hooks'] = n

            # OWNER RULE (2026-08-20):
            #   turns   = middle poles where the cable turns (angle > 15°)
            #   tangent = 1 per middle pole where cable continues STRAIGHT (angle <= 15°)
            #   deadend = turns + 2  (start pole + end pole always get one, sized by cable OD)
            if n >= 3:
                for i in range(1, n-1):
                    pp = ordered[i-1][1].centroid().asPoint()
                    cp = ordered[i][1].centroid().asPoint()
                    np = ordered[i+1][1].centroid().asPoint()
                    vi = QgsPointXY(cp.x()-pp.x(), cp.y()-pp.y())
                    vo = QgsPointXY(np.x()-cp.x(), np.y()-cp.y())
                    dot = vi.x()*vo.x() + vi.y()*vo.y()
                    mi = math.hypot(vi.x(), vi.y())
                    mo = math.hypot(vo.x(), vo.y())
                    if mi < 0.001 or mo < 0.001: continue
                    ang = math.degrees(math.acos(max(-1, min(1, dot/(mi*mo)))))
                    if ang > 15.0:
                        rec['turns_gt15'] += 1
                    else:
                        b = _bracket(od, brackets.get('tangent',[]))
                        if b:
                            tangents[b] += 1
                            rec['tangent_qty'] += 1; rec['tangent_size'] = b
                        else: self._nb_warn('tangent', fiber, od)

            # Dead ends: 1 at start pole + 1 at end pole + 2 per turning pole
            # (a turn pole anchors the cable on BOTH sides; every path pole
            #  gets either dead-ends (turn) or a tangent (straight))
            de_total = 2 * rec['turns_gt15'] + 2
            b = _bracket(od, brackets.get('dead_end',[]))
            if b:
                dead_ends[b] += de_total
                rec['deadend_qty'] = de_total; rec['deadend_size'] = b
            else: self._nb_warn('dead_end', fiber, od)

        self.v['dead_ends'] = dict(dead_ends)
        self.v['tangents'] = dict(tangents)
        self.v['hook_count'] = hook_count + self.v['feeder_joints'] + self.v['link_joints']
        self.cable_audit = cable_audit
        self.audit_by_name = audit_by_name

        # ── OVAL + GLAND from wizard Express/Terminated or auto-compute ──
        express = self.wiz.get('express_cables', {})
        oval = defaultdict(int)
        gland = defaultdict(int)

        # OWNER RULE (2026-08-11): NO geometric express guessing.
        # Default = ALL cables Terminated → all get round glands (mech seals).
        # Ovals come ONLY from user-ticked Express checkboxes in Tab 4.
        # Each ticked Express cable = 1 oval kit sized by its OD.

        if express:
            # User ticked Express in Tab 4 for specific cables.
            for key, is_express in express.items():
                parts = key.split('||', 1)
                cable_name = parts[1] if len(parts) > 1 else key
                fiber = _pf(cable_name)
                if not fiber: continue
                opt = self._get_od_full(cable_name, fiber)
                od = opt.get('od', 0)
                if od == 0: continue

                if is_express:
                    b = _bracket(od, brackets.get('oval', []))
                    if b:
                        oval[b] += 1
                        rec = audit_by_name.get(cable_name)
                        if rec: rec['gland_oval'] = 'Oval'; rec['gland_oval_size'] = b
                    else: self._nb_warn('oval_port', fiber, od)

        # ALL cables (conventional, at manholes) = Terminated by default → glands.
        # Count one gland per cable-MH connection, UNLESS the cable was ticked Express.
        express_keys = set()
        for key in express:
            express_keys.add(key.split('||', 1)[-1])  # cable name only

        ca_layer = self._L('Core_Aggregation')
        if ca_layer:
            for f in ca_layer.getFeatures():
                mh_geom = f.geometry()
                if not mh_geom or mh_geom.isNull(): continue
                mh_geom = self._g(ca_layer, mh_geom)
                if mh_geom.type() == 2: mh_geom = mh_geom.centroid()
                for name, fiber, opt, od, g, role in conventional:
                    if g.distance(mh_geom) <= 10:
                        # Skip if this cable is Express (already counted as oval)
                        if name in express_keys:
                            continue
                        b = _bracket(od, brackets.get('mech_seal', []))
                        if b:
                            gland[b] += 1
                            rec = audit_by_name.get(name)
                            if rec: rec['gland_oval'] = 'Gland'; rec['gland_oval_size'] = b
                        else: self._nb_warn('mech_seal', fiber, od)

        self.v['oval_ports'] = dict(oval)
        self.v['mech_seals'] = dict(gland)

    # ════════════════════════════════════════════════════════
    #  PHASE 4a: DISTRIBUTION ACCESSORIES
    # ════════════════════════════════════════════════════════
    def _dist_accessories(self):
        pole_geoms = []
        l = self._L('Poles')
        if l:
            for f in l.getFeatures():
                g = f.geometry()
                if g and not g.isNull(): pole_geoms.append((f['Name'] or '', self._g(l, g)))

        l = self._L('Distribution')
        total_pass = dc_touch = 0; plum_set = set()
        tol = self.pole_tol
        if l:
            for f in l.getFeatures():
                g = f.geometry()
                if not g or g.isNull(): continue
                g = self._g(l, g)
                n = sum(1 for _, pg in pole_geoms if g.distance(pg) <= tol)
                if n > 0:
                    total_pass += n; dc_touch += 1
                    for pn, pg in pole_geoms:
                        if g.distance(pg) <= tol: plum_set.add(pg.asWkt())
        self.v['wedges'] = 2*total_pass - 2*dc_touch
        self.v['plum_hooks'] = len(plum_set)
        self.v['dist_pole_passes'] = total_pass  # audit Table 3, row 1

        # poles hosting a feeder/link aggregation box or a distribution joint
        # (audit Table 3, row 2) — unique poles within tolerance of any box/joint
        box_geoms = []
        for role_b in ['Feeder_aggregation', 'Link_Aggregation', 'Distributions joints']:
            l_b = self._L(role_b)
            if not l_b: continue
            for f in l_b.getFeatures():
                gb = f.geometry()
                if gb and not gb.isNull():
                    gb = self._g(l_b, gb)
                    if gb.type() == 2: gb = gb.centroid()
                    box_geoms.append(gb)
        box_poles = set()
        for pn, pg in pole_geoms:
            if any(pg.distance(gb) <= tol for gb in box_geoms):
                box_poles.add(pg.asWkt())
        self.v['poles_with_boxes'] = len(box_poles)

    # ════════════════════════════════════════════════════════
    #  PHASE 4b: POP DETECTION
    # ════════════════════════════════════════════════════════
    def _pop_detect(self):
        self.v['pop_288'] = self.v['pop_144'] = 0
        pop_layer = self._L('POP'); core_layer = self._L('core_cable')
        if not pop_layer or not core_layer: return
        pop_geom = None
        for f in pop_layer.getFeatures():
            g = f.geometry()
            if not g or g.isNull(): continue
            g = self._g(pop_layer, g)
            pop_geom = g.centroid() if g.type()==2 else g; break
        if not pop_geom: return
        for f in core_layer.getFeatures():
            g = f.geometry()
            if not g or g.isNull(): continue
            g = self._g(core_layer, g)
            fiber = _pf(f['Name'] or '')
            if fiber in ('288','144') and g.distance(pop_geom) <= 5:
                self.v[f'pop_{fiber}'] = self.v.get(f'pop_{fiber}',0) + 1

    # ════════════════════════════════════════════════════════
    #  PHASE 5: DEPENDED + PRECON
    # ════════════════════════════════════════════════════════
    def _depended(self):
        jtc = self.wiz.get('joint_type_counts', {})
        self.v['lmj_count'] = jtc.get('LMJ', 0)
        self.v['mmj_count'] = jtc.get('MMJ', 0)
        self.v['odc_fd4'] = jtc.get('ODC FD4', 0)
        self.v['odc_fd6'] = jtc.get('ODC FD6', 0)
        self.v['lmj_mmj_on_pole'] = self.v.get('lmj_count',0) + self.v.get('mmj_count',0)

    def _precon(self):
        l = self._L('Distribution')
        counts = {c:0 for c in PRECON_CABLES}
        w = []
        if not l: return counts, w
        for f in l.getFeatures():
            name = f['Name'] or ''
            if _pf(name) != '1': continue
            # Use name-pattern length (YYYY) for precon input
            yyyy = self._name_length(name)
            length = yyyy if yyyy else (self._g(l, f.geometry()).length() if f.geometry() and not f.geometry().isNull() else 0)
            if length <= 0: continue
            _, cable, _ = _best_precon(length)
            if cable: counts[cable] += 1
            else: w.append(f"{name[:60]} ({length:.0f}m)")
        return counts, w

    # ════════════════════════════════════════════════════════
    #  ITEM COMPUTATION
    # ════════════════════════════════════════════════════════
    def _compute(self, item):
        r = item.get('rule',''); v = self.v
        if r == 'blocks_count': return v['blocks']
        if r == 'core_aggs_count': return v['core_aggs']
        if r == 'dj_count': return v['dj_count']
        if r == 'agg_polygons_count': return v.get('agg_polygons',0)
        if r in ('poles_6m','poles_7m','poles_9m'): return v.get(r,0)
        if r == 'm8_count': return v.get('m8',0)
        if r == 'm16_count': return v.get('m16',0)
        if r == 'dj_1_8_count': return v.get('dj_1_8',0)
        if r == 'dj_1_9_count': return v.get('dj_1_9',0)
        if r == 'm8_plus_m16': return v.get('m8',0)+v.get('m16',0)
        if r == 'wedges': return v.get('wedges',0)
        if r == 'plum_hooks': return v.get('plum_hooks',0)
        if r == 'hooks': return v.get('hook_count',0)
        if r == 'manual': return 0
        if r == 'bare_splitter_1_2': return v['bare_1_2']
        if r == 'premises_div_52': return math.ceil(v['premises']/52)
        if r == 'blocks_div_4': return math.ceil(v['blocks']/4)
        if r == 'bandit_strap': return v['bandit_strap']
        if r == 'bandit_buckle': return v['bandit_buckle']
        if r == 'agg_polygons_div_24': return math.ceil(v.get('agg_polygons',0)/24)
        if r == 'item_52345294_div_2': return math.ceil(v['item_52345294']/2)
        if r == 'same_as_52345294': return v['item_52345294']
        if r == 'same_as_52345292': return v.get('m16',0)
        if r in ('wizard_lmj','lmj_mmj_on_pole','wizard_lmj_mmj_on_pole'): return v.get('lmj_mmj_on_pole',0)
        if r == 'wizard_mmj': return v.get('mmj_count',0)
        if r == 'pop_panel_288': return v.get('pop_288',0)
        if r == 'pop_panel_144': return v.get('pop_144',0)
        if r == 'pole_assembly': return v['dj_count']+2*v['feeder_joints']+2*v['link_joints']
        if r == 'cable_length': return v.get('cable_lengths',{}).get(item['fiber'],0)
        if r in ('dead_end','tangent'):
            od = item['od']; key = f"{od[0]}-{od[1]}"
            store = 'dead_ends' if r=='dead_end' else 'tangents'
            return v.get(store,{}).get(key,0)
        if r == 'oval_port':
            od = item['od']; key = f"{od[0]}-{od[1]}"
            return v.get('oval_ports',{}).get(key,0)
        if r == 'mech_seal_entry':
            od = item['od']; key = f"{od[0]}-{od[1]}"
            return v.get('mech_seals',{}).get(key,0)
        if r == 'precon': return v.get('precon_counts',{}).get(item['length'],0)
        self.warnings.append(f"Unknown rule: {r}"); return 0

    # ════════════════════════════════════════════════════════
    #  SUMMARY + EXPORT
    # ════════════════════════════════════════════════════════
    def _summary(self):
        v = self.v
        return "\n".join([
            f"Blocks={v['blocks']} | Agg_Polygons={v.get('agg_polygons',0)} | Premises={v.get('premises',0)}",
            f"Core Aggs={v['core_aggs']} | POP 288F={v.get('pop_288',0)} 144F={v.get('pop_144',0)}",
            f"Poles: 6m={v['poles_6m']} 7m={v['poles_7m']} 9m={v['poles_9m']} total={v['total_poles']}",
            f"DJ 1:8={v.get('dj_1_8',0)} 1:9={v.get('dj_1_9',0)} M8={v.get('m8',0)} M16={v.get('m16',0)}",
            f"Feeder joints={v['feeder_joints']} Link joints={v['link_joints']}",
            f"Hooks={v.get('hook_count',0)} | Dead ends={v.get('dead_ends',{})}",
            f"Tangents={v.get('tangents',{})}",
            f"Wedges={v.get('wedges',0)} | Plum hooks={v.get('plum_hooks',0)}",
            f"Oval ports={v.get('oval_ports',{})} | Mech seals={v.get('mech_seals',{})}",
            f"Cable lengths (Qty=ROUNDUP(SUM×1.05)): {dict((k,f'{v}m') for k,v in v.get('cable_lengths',{}).items())}",
        ])

    # ════════════════════════════════════════════════════════
    #  EXPORT — Save-As dialog + template pool export + crash-proof
    #  Template = master item pool (same rows every time, only Qty differs).
    #  Match by item CODE only. Qty 0 / unused → cell stays BLANK.
    # ════════════════════════════════════════════════════════
    def _export(self):
        from qgis.PyQt.QtWidgets import QFileDialog
        from qgis.PyQt.QtCore import QSettings
        proj = (self.wiz.get('project_name') or 'Project').replace(' ','_')
        dt = datetime.now().strftime('%Y%m%d')
        suggested = f"{proj}_BOM_{dt}.xlsx"
        last_dir = QSettings().value('ftth_bom/last_export_dir',
                                     os.path.join(os.path.expanduser('~'), 'Desktop'))
        out, _ = QFileDialog.getSaveFileName(self.iface.mainWindow(), "Save BOM As",
                                             os.path.join(last_dir, suggested),
                                             "Excel Workbook (*.xlsx)")
        if not out: return
        QSettings().setValue('ftth_bom/last_export_dir', os.path.dirname(out))
        try:
            tpl = os.path.join(self.plugin_dir, 'bom_template.xlsx')
            if os.path.exists(tpl):
                notes = self._export_from_template(tpl, out, dt)
            else:
                self._export_plain(out, dt)
                notes = ["Template bom_template.xlsx not found in plugin folder — used plain export."]
        except Exception as e:
            QMessageBox.critical(self.iface.mainWindow(), "Export failed",
                                 f"Could not save the BOM file:\n\n{e}")
            return
        msg = f"BOM saved to:\n{out}"
        if notes:
            msg += "\n\nNOTES:\n" + "\n".join(f"  {n}" for n in notes[:25])
        # Auto-open the saved workbook (Windows)
        try:
            if hasattr(os, 'startfile'): os.startfile(out)
        except Exception:
            pass
        QMessageBox.information(self.iface.mainWindow(), "BOM Export", msg)

    def _export_from_template(self, tpl, out, dt):
        """Fill the Vumatel template: header fields + Qty column, matched by item code.
        Every template row stays (pool); unused/zero items keep a BLANK Qty cell."""
        by_code = {}
        for it in self.config['bom_items']:
            by_code[str(it.get('code','')).strip()] = it

        wb = openpyxl.load_workbook(tpl)
        ws = wb['BOM'] if 'BOM' in wb.sheetnames else wb.active
        ws['B4'] = self.wiz.get('customer','')
        ws['B5'] = self.wiz.get('project_name','')
        ws['B6'] = self.wiz.get('project_code','')
        ws['B7'] = dt

        notes = []
        used_codes = set()
        for row in range(10, ws.max_row + 1):
            raw = ws.cell(row=row, column=1).value
            if raw is None: continue
            code = str(raw).strip()
            if code.endswith('.0'): code = code[:-2]
            if not code or not code[0].isdigit(): continue  # section header rows
            it = by_code.get(code)
            if it is None:
                notes.append(f"ITEM NOT FOUND: {code} — {str(ws.cell(row=row,column=2).value)[:45]}")
                continue
            used_codes.add(code)
            ws.cell(row=row, column=5, value=it.get('rule',''))
            if not it.get('enabled', True): continue
            qty = self._compute(it)
            if qty:
                ws.cell(row=row, column=3, value=qty)

        # Safety net: enabled JSON item with qty>0 but missing from the template pool
        for it in self.config['bom_items']:
            code = str(it.get('code','')).strip()
            if code in used_codes or not it.get('enabled', True): continue
            try: qty = self._compute(it)
            except Exception: qty = 0
            if qty:
                notes.append(f"NOT IN TEMPLATE: {code} — {it.get('desc','')[:40]} (qty={qty}) — add a row to bom_template.xlsx")

        # Refresh the audit sheet
        audit = wb['Calculation_Audit'] if 'Calculation_Audit' in wb.sheetnames else wb.create_sheet('Calculation_Audit')
        if audit.max_row > 1:
            audit.delete_rows(1, audit.max_row)
        audit['A1']='Variable'; audit['B1']='Value'
        for i,(k,val) in enumerate(sorted(self.v.items()),2):
            audit.cell(row=i,column=1,value=k); audit.cell(row=i,column=2,value=str(val))
        self._write_audit_tables(audit, len(self.v) + 3)

        wb.save(out)
        return notes

    def _export_plain(self, out, dt):
        """Legacy fallback export (no template present)."""
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "BOM"
        ws['A1']='Customer:'; ws['B1']=self.wiz.get('customer','')
        ws['A2']='Project:'; ws['B2']=self.wiz.get('project_name','')
        ws['A3']='Date:'; ws['B3']=dt
        for c,h in enumerate(['Item/Part No','Description','Qty','UOM','Rule'],1):
            ws.cell(row=5,column=c,value=h).font = openpyxl.styles.Font(bold=True)
        row = 6
        for item in self.config['bom_items']:
            if not item.get('enabled',True): continue
            qty = self._compute(item)
            ws.cell(row=row,column=1,value=item.get('code',''))
            ws.cell(row=row,column=2,value=item.get('desc',''))
            ws.cell(row=row,column=3,value=qty if qty else None)
            ws.cell(row=row,column=4,value=item.get('uom',''))
            ws.cell(row=row,column=5,value=item.get('rule',''))
            row += 1

        audit = wb.create_sheet("Calculation_Audit")
        audit['A1']='Variable'; audit['B1']='Value'
        for i,(k,val) in enumerate(sorted(self.v.items()),2):
            if isinstance(val,dict): val = str(val)
            audit.cell(row=i,column=1,value=k); audit.cell(row=i,column=2,value=str(val))
        self._write_audit_tables(audit, len(self.v) + 3)

        wb.save(out)

    # ════════════════════════════════════════════════════════
    #  AUDIT TABLES — per-cable table + key item QTY table
    #  (appended to Calculation_Audit below the variable dump)
    # ════════════════════════════════════════════════════════
    def _write_audit_tables(self, ws, header_row):
        from openpyxl.styles import Font
        bold = Font(bold=True)
        r = header_row

        # ── Table 1: per-cable audit ──
        headers = ['Name', 'atual length', 'pulished length', 'added buffer',
                   'number of passed  feeder or link aggregation ',
                   'number of passed poles', 'number of turns>15 degree',
                   'deadend QTY', 'Dead end size', 'tangent QTY', 'Tangent size',
                   'gland or Oval', 'gland or Oval size',
                   'BRACKET 3 WAY SHORT (HOOK) QTY']
        for c, h in enumerate(headers, 1):
            ws.cell(row=r, column=c, value=h).font = bold
        r += 1
        for rec in getattr(self, 'cable_audit', []):
            vals = [rec['name'], rec['actual'], rec['published'], rec['buffer'],
                    rec['agg_passes'], rec['poles'], rec['turns_gt15'],
                    rec['deadend_qty'] or None, rec['deadend_size'] or None,
                    rec['tangent_qty'] or None, rec['tangent_size'] or None,
                    rec['gland_oval'] or None, rec['gland_oval_size'] or None,
                    rec['hooks'] or None]
            for c, val in enumerate(vals, 1):
                ws.cell(row=r, column=c, value=val)
            r += 1

        # ── Table 2: key item QTY ──
        r += 1
        ws.cell(row=r, column=1, value='item').font = bold
        ws.cell(row=r, column=2, value='QTY').font = bold
        r += 1
        by_code = {str(it.get('code','')).strip(): it for it in self.config['bom_items']}
        def q(code):
            it = by_code.get(code)
            if not it: return None
            try: return self._compute(it)
            except Exception: return None
        rows2 = [
            ('aggregation polygon count', self.v.get('agg_polygons', 0)),
            ('V SHAPE SLACK BRACKET QTY', q('1524345')),
            ('Block  polygon count', self.v.get('blocks', 0)),
            ('(MK2 Fibre Tray and Face Plate) 1u Empty patch panel Quad faceplate 24 slots', q('52345313')),
            ('Quad flanged mid couplers LC/APC', q('52345311')),
            ('PLC 1 x 2 SPLITTER WITH LCAPC (PRE-CONNECTED)', q('52345310')),
            ('LC/APC to LC/APC patch cord 3m', q('52345308')),
            ('LC/APC to SC/UPC patch cord 3m', q('52345309')),
            ('SPLITTER BARE FIBRE 2 WAY', q('52345294')),
        ]
        for label, val in rows2:
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=val)
            r += 1

        # ── Table 3: pole-join summary ──
        r += 1
        ws.cell(row=r, column=1, value='case').font = bold
        ws.cell(row=r, column=2, value='QTY').font = bold
        r += 1
        rows3 = [
            ('sum of total numbrt of poles passed by each distribution 1F cable '
             '(join attributes by location, tolerance 5m, Distribution vs Poles, '
             'summed over all distribution cables)',
             self.v.get('dist_pole_passes', 0)),
            ('total number of poles that having feeder/Link aggregation box '
             'or distribution joint on it',
             self.v.get('poles_with_boxes', 0)),
        ]
        for label, val in rows3:
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=val)
            r += 1
        ws.column_dimensions['A'].width = 60
