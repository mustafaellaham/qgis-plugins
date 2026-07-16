#!/usr/bin/env python3
"""
bom_engine.py - FTTH Network Bill of Materials Engine

A standalone, QGIS-decoupled Python module for calculating Bill of Materials
in FTTH (Fibre-to-the-Home) network designs using Huawei QuickODN
pre-conventional architecture.

Network hierarchy: OLT -> Core Cable -> AGG MH -> Feeder Cable -> FJ -> DC -> DJ -> Houses

Author: FTTH Design Automation
License: MIT
"""

from __future__ import annotations

import ast
import csv
import os
from typing import Any, Dict, List, Optional, Tuple

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Standard drop-cable drum lengths (metres) used by the DC fixed-length formula.
DC_STANDARD_LENGTHS: Tuple[int, ...] = (50, 80, 100, 120, 150, 180, 200, 250, 300, 350)

# Mapping of cable size -> fibre count
FIBRE_COUNT_MAP: Dict[str, int] = {
    "12F": 12,
    "24F": 24,
    "48F": 48,
    "72F": 72,
    "96F": 96,
    "144F": 144,
    "288F": 288,
}

# Valid cable sizes
VALID_CABLE_SIZES: Tuple[str, ...] = tuple(FIBRE_COUNT_MAP.keys())

# Feeder cable part number by size
FEEDER_PART_MAP: Dict[str, str] = {
    "12F": "CAB00280",
    "24F": "CAB00281",
    "48F": "CAB00282",
    "72F": "CAB00344",
    "96F": "CAB00344",   # 96F uses 72F part as placeholder; adjust if a dedicated 96F part exists
    "144F": "CAB00346",
    "288F": "CAB00347",
}

# Tangent clamp part number by cable size
TANGENT_PART_MAP: Dict[str, str] = {
    "12F": "PAC00051",
    "24F": "PAC00051",
    "48F": "PAC00051",
    "72F": "PAC00051",
    "96F": "PAC00051",
    "144F": "PAC00051",
    "288F": "PAC00051",
}

# Dead-end clamp part number by cable size
DEADEND_PART_MAP: Dict[str, str] = {
    "12F": "PAC00073",
    "24F": "PAC00073",
    "48F": "PAC00073",
    "72F": "PAC00053",
    "96F": "PAC00063",   # 96F uses 96F/288F slim-line dead-end
    "144F": "PAC00053",
    "288F": "PAC00063",
}

# Drop-cable part number by fixed length
DROP_CABLE_PART_MAP: Dict[int, str] = {
    50:  "CAB00309",
    80:  "CAB00310",
    100: "CAB00311",
    120: "CAB00312",
    150: "CAB00313",
    180: "CAB00314",
    200: "CAB00315",
    250: "CAB00334",
    300: "CAB00335",
    350: "CAB00337",
}

# DJ box part by ratio (splitter ratio)
DJ_BOX_PART_MAP: Dict[str, str] = {
    "1:9":  "TRA00166",
    "1:8":  "TRA00167",
    "1x9":  "TRA00166",
    "1x8":  "TRA00167",
}


# ---------------------------------------------------------------------------
# Material Catalog
# ---------------------------------------------------------------------------

MATERIAL_CATALOG: Dict[str, Dict[str, str]] = {
    # === PRE-CONVENTIONAL ===
    "TRA00166": {"description": "FastConnect fiber access terminal-Termination 9 cores-Wall/Pole/Aerial mount-220*187*77mm-Plastic-RAL7035-FAT-SSC2811-SM-9U-Mechanical sealing-with valve-with 1 pcs 30/70 Uneven 1:9 SPL9105", "unit": "EACH", "category": "pre_conventional", "sub_category": "dj_box_1x9"},
    "TRA00167": {"description": "FastConnect fiber access terminal-Termination 8 cores-Wall/Pole mounting/aerial mounting-220*187*77mm-Plastic-RAL7035-FAT-SSC2811-SM-8-Mechanical sealing-Including 1pcs 1:8 SPL9105", "unit": "EACH", "category": "pre_conventional", "sub_category": "dj_box_1x8"},
    "CLO00172": {"description": "FastConnect closure-Termination 8cores&Splicing 48cores-Wall/Pole/Aerial mount-367*209*127mm-Plastic-Black-SSC2802-TM-8-SC/APC-Mechanical sealing-with valve-with 1 FP-S01,2 FP-B01", "unit": "EACH", "category": "pre_conventional", "sub_category": "fj_closure"},
    "CLO00174": {"description": "M8 MICROLOOP 24F JOINT (Entry 6-9mm)3-4.5mm Drop", "unit": "EACH", "category": "pre_conventional", "sub_category": "fj_closure_microloop"},
    "CAB00309": {"description": "Pre-terminated cable,Fastconnect SC/APC-Fastconnect SC/APC,GYFXTY-1B6a2,50m,Non-metal,Half Dry core,Black HDPE,1core,5mm", "unit": "EACH", "category": "pre_conventional", "sub_category": "drop_cable_50m"},
    "CAB00310": {"description": "Pre-terminated cable,Fastconnect SC/APC-Fastconnect SC/APC,GYFXTY-1B6a2,80m,Non-metal,Half Dry core,Black HDPE,1core,5mm", "unit": "EACH", "category": "pre_conventional", "sub_category": "drop_cable_80m"},
    "CAB00311": {"description": "Pre-terminated cable,Fastconnect SC/APC-Fastconnect SC/APC,GYFXTY-1B6a2,100m,Non-metal,Half Dry core,Black HDPE,1core,5mm", "unit": "EACH", "category": "pre_conventional", "sub_category": "drop_cable_100m"},
    "CAB00312": {"description": "Pre-terminated cable,Fastconnect SC/APC-Fastconnect SC/APC,GYFXTY-1B6a2,120m,Non-metal,Half Dry core,Black HDPE,1core,5mm", "unit": "EACH", "category": "pre_conventional", "sub_category": "drop_cable_120m"},
    "CAB00313": {"description": "Pre-terminated cable,Fastconnect SC/APC-Fastconnect SC/APC,GYFXTY-1B6a2,150m,Non-metal,Half Dry core,Black HDPE,1core,5mm", "unit": "EACH", "category": "pre_conventional", "sub_category": "drop_cable_150m"},
    "CAB00314": {"description": "Pre-terminated cable,Fastconnect SC/APC-Fastconnect SC/APC,GYFXTY-1B6a2,180m,Non-metal,Half Dry core,Black HDPE,1core,5mm", "unit": "EACH", "category": "pre_conventional", "sub_category": "drop_cable_180m"},
    "CAB00315": {"description": "Pre-terminated cable,Fastconnect SC/APC-Fastconnect SC/APC,GYFXTY-1B6a2,200m,Non-metal,Half Dry core,Black HDPE,1core,5mm", "unit": "EACH", "category": "pre_conventional", "sub_category": "drop_cable_200m"},
    "CAB00334": {"description": "Pre-terminated cable,Fastconnect SC/APC-Fastconnect SC/APC,GYFXTY-1B6a2,250m,Non-metal,Half Dry core,Black HDPE,1core,5mm", "unit": "EACH", "category": "pre_conventional", "sub_category": "drop_cable_250m"},
    "CAB00335": {"description": "Pre-terminated cable,Fastconnect SC/APC-Fastconnect SC/APC,GYFXTY-1B6a2,300m,Non-metal,Half Dry core,Black HDPE,1core,5mm", "unit": "EACH", "category": "pre_conventional", "sub_category": "drop_cable_300m"},
    "CAB00337": {"description": "Pre-terminated cable,Fastconnect SC/APC-Fastconnect SC/APC,GYFXTY-1B6a2,350m,Non-metal,Half Dry core,Black HDPE,1core,5mm", "unit": "EACH", "category": "pre_conventional", "sub_category": "drop_cable_350m"},
    "CAB00326": {"description": "Pigtail,SC/APC-/,Single mode,G.652D,0.9mm,1.5m,LSZH,with 2.0mm protection tube,ODN", "unit": "EACH", "category": "pre_conventional", "sub_category": "pigtail"},
    "CAB00175": {"description": "VUMA - UNJACKED PIGTAILS LC/APC - 1M (9/125)", "unit": "EACH", "category": "pre_conventional", "sub_category": "pigtail_lc"},
    "TRA00115": {"description": "VUMA - MID COUPLER LC/APC", "unit": "EACH", "category": "pre_conventional", "sub_category": "mid_coupler"},
    "CLO00137": {"description": "SPLICE PROTECTOR 1.3MM", "unit": "EACH", "category": "pre_conventional", "sub_category": "splice_protector_1.3"},

    # === CONVENTIONAL ===
    "CAB00280": {"description": "MINI ADSS CABLE 12 CORE", "unit": "km", "category": "conventional", "sub_category": "feeder_cable"},
    "CAB00281": {"description": "MINI ADSS CABLE 24 CORE", "unit": "km", "category": "conventional", "sub_category": "feeder_cable"},
    "CAB00282": {"description": "MINI ADSS CABLE 48 CORE", "unit": "km", "category": "conventional", "sub_category": "feeder_cable"},
    "CAB00344": {"description": "CABLE 72 CORE Slim-line ADSS", "unit": "km", "category": "conventional", "sub_category": "feeder_cable"},
    "CAB00346": {"description": "CABLE 144 CORE Slim-line ADSS", "unit": "km", "category": "conventional", "sub_category": "feeder_cable"},
    "CAB00347": {"description": "CABLE 288 CORE Slim-line ADSS", "unit": "km", "category": "conventional", "sub_category": "feeder_cable"},
    "CLO00110": {"description": "DOME JOINT LMJ SHORT Cap  + 24XSE2.2MM 12F TRAYS", "unit": "EACH", "category": "conventional", "sub_category": "dome_joint"},
    "CLO00138": {"description": "LMJ SPLITTER TRAY (GREEN TRAYS)", "unit": "EACH", "category": "conventional", "sub_category": "splitter_tray"},
    "TRA00142": {"description": "PLC 1 x 2 SPLITTER WITH LCAPC (PRE-CONNECTED)", "unit": "EACH", "category": "conventional", "sub_category": "splitter_1x2"},
    "TRA00077": {"description": "PLC 1 x 8 SPLITTER WITH LCAPC (PRE-CONNECTED)", "unit": "EACH", "category": "conventional", "sub_category": "splitter_1x8"},
    "TRA00117": {"description": "SPLITTER BARE FIBRE 2 WAY", "unit": "EACH", "category": "conventional", "sub_category": "splitter_bare"},
    "CLO00016": {"description": "2.2mm x 45mm splice protectors", "unit": "EACH", "category": "conventional", "sub_category": "splice_protector"},
    "TRA00122": {"description": "Quad flanged mid couplers LC/APC", "unit": "EACH", "category": "conventional", "sub_category": "mid_coupler_quad"},
    "CAB00188": {"description": "LC/APC to SC/UPC patch cord 3m", "unit": "EACH", "category": "conventional", "sub_category": "patch_cord"},
    "CAB00189": {"description": "LC/APC to LC/APC patch cord 3m", "unit": "EACH", "category": "conventional", "sub_category": "patch_cord"},
    "TRA00031": {"description": "(MK2 Fibre Tray and Face Plate) 1u Empty patch panel Quad faceplate 24 slots", "unit": "EACH", "category": "conventional", "sub_category": "patch_panel"},
    "TRA00035": {"description": "PANEL MFPS-IXD P-SILD-288-(ZAV2) Right", "unit": "EACH", "category": "conventional", "sub_category": "odf_panel"},

    # === POLE HARDWARE ===
    "PAC00051": {"description": "TANGENT ADSS 12F", "unit": "EACH", "category": "pole_hardware", "sub_category": "tangent"},
    "PAC00053": {"description": "DEAD-END ADSS 144F", "unit": "EACH", "category": "pole_hardware", "sub_category": "dead_end"},
    "PAC00054": {"description": "V SHAPE SLACK BRACKET", "unit": "EACH", "category": "pole_hardware", "sub_category": "v_shape"},
    "PAC00063": {"description": "DEAD-END ADSS 96F(288F Slim-line)", "unit": "EACH", "category": "pole_hardware", "sub_category": "dead_end_288"},
    "PAC00064": {"description": "BRACKET 3 WAY SHORT (HOOK)", "unit": "EACH", "category": "pole_hardware", "sub_category": "3way_hook"},
    "PAC00068": {"description": "AERIAL SLACK STORAGE BRACKET (BLK)", "unit": "EACH", "category": "pole_hardware", "sub_category": "slack_storage"},
    "PAC00073": {"description": "DEAD-END MINI ADSS 12/24", "unit": "EACH", "category": "pole_hardware", "sub_category": "dead_end_mini"},
    "PAC00077": {"description": "HUAWEI Pole mounting assembly-For diameter 114~381mm", "unit": "EACH", "category": "pole_hardware", "sub_category": "pole_mount"},
    "PAC00078": {"description": "HUAWEI AERIAL SLACK STORAGE BRACKET (BLK)", "unit": "EACH", "category": "pole_hardware", "sub_category": "huawei_slack"},
    "CLO00177": {"description": "M 8/M16 UNIVERSAL BRACKET M-RANGE - BLACK", "unit": "EACH", "category": "pole_hardware", "sub_category": "universal_bracket"},

    # === CONSUMABLES ===
    "CLE00004": {"description": "ALCOHOL SPRAY PUMP ACTION BOTTLE 200ml", "unit": "EACH", "category": "consumables", "sub_category": "alcohol"},
    "CLE00006": {"description": "KIM WIPES (BOX)", "unit": "EACH", "category": "consumables", "sub_category": "kim_wipes"},
    "GEN00007": {"description": "CABLE TIES 305 X 4.7MM BLACK LARGE", "unit": "EACH", "category": "consumables", "sub_category": "cable_tie_large"},
    "GEN00116": {"description": "CABLE TIES BLACK 100 X 2.5MM SML", "unit": "EACH", "category": "consumables", "sub_category": "cable_tie_small"},
    "GEN00035": {"description": "NITTO TAPE (BLACK)", "unit": "EACH", "category": "consumables", "sub_category": "nitto_tape"},
    "GEN00130": {"description": "VELCRO BLACK 25M X 19MM", "unit": "EACH", "category": "consumables", "sub_category": "velcro"},
    "GEN00064": {"description": "BANDIT S/STEEL BUCKLES 19mm", "unit": "EACH", "category": "consumables", "sub_category": "bandit_buckle"},
    "GEN00202": {"description": "BANDIT S/STEEL STRAP 19mm (30M)", "unit": "EACH", "category": "consumables", "sub_category": "bandit_strap"},
    "GEN00087": {"description": "BUNNY CLIPS", "unit": "EACH", "category": "consumables", "sub_category": "bunny_clips"},
    "GEN00134": {"description": "DARTAG SLEEVE 10MM X 100MM", "unit": "EACH", "category": "consumables", "sub_category": "dartag"},
    "GEN00170": {"description": "VINYL CARTRIDGE (BLACK ON YELLOW) 9.53MM", "unit": "EACH", "category": "consumables", "sub_category": "vinyl"},
    "GEN00191": {"description": "CAUTION FIBER TAPE ( VUMA TRENCH TAPE)", "unit": "EACH", "category": "consumables", "sub_category": "caution_tape"},
    "CLO00015": {"description": "LMJ OVAL PORT KIT 12.1MM-14MM", "unit": "EACH", "category": "consumables", "sub_category": "oval_port"},
    "CLO00035": {"description": "MECH SEAL-MEDIUM ENTRY GLAND 7-20MM (ENTRY PORTS)", "unit": "EACH", "category": "consumables", "sub_category": "mech_seal"},
    "CLO00039": {"description": "MECH SEAL CIRCULAR PORT QUAD (4-6MM)(4 CABLES)", "unit": "EACH", "category": "consumables", "sub_category": "mech_seal_quad"},
    "BOL00002": {"description": "CAGE NUT & SCREW M6", "unit": "EACH", "category": "consumables", "sub_category": "cage_nut"},
    "BOL00056": {"description": "NYLOC NUT HEX 6mm", "unit": "EACH", "category": "consumables", "sub_category": "nyloc_nut"},
    "BOL00059": {"description": "BOLT HEX SCREW SET 6X16MM", "unit": "EACH", "category": "consumables", "sub_category": "bolt"},

    # === OTHER ===
    "GLA00032": {"description": "RHI-NODE 1000 MD-STD GLAM (VUMA)", "unit": "EACH", "category": "other", "sub_category": "glam"},
    "POL00013": {"description": "TRANS GUM POLES (CCA) 7M (120/140)", "unit": "EACH", "category": "other", "sub_category": "pole_7m"},
    "POL00015": {"description": "TRANS GUM POLES (CCA) 9M (120/140)", "unit": "EACH", "category": "other", "sub_category": "pole_9m"},
    "POL00017": {"description": "TRANS GUM POLES (CCA) 6M (100/120)", "unit": "EACH", "category": "other", "sub_category": "pole_6m"},
    "TUB00026": {"description": "DUCT RIPPLE 110MM - (6M LENGTH)", "unit": "EACH", "category": "other", "sub_category": "duct_ripple"},
    "TUB00028": {"description": "114MM X 2MM HALF ROUND PIPE 3M", "unit": "EACH", "category": "other", "sub_category": "half_round"},
    "TUB00128": {"description": "SUB DUCT 40/34MM BLACK", "unit": "EACH", "category": "other", "sub_category": "sub_duct"},
    "S001219": {"description": "SMART LOCKS MANHOLE LID", "unit": "EACH", "category": "other", "sub_category": "smart_lock"},
    "TOO00280": {"description": "Plastic wedge clamping tool,ITC3102-A1,Installation accessory of optical", "unit": "EACH", "category": "other", "sub_category": "tool_wedge"},
    "TOO00281": {"description": "Plum ring hook,without metal hoop,ITC3301-P1", "unit": "EACH", "category": "other", "sub_category": "tool_hook"},
}


# ---------------------------------------------------------------------------
# Helper / Lookup Functions
# ---------------------------------------------------------------------------

def calculate_dc_fixed_length(length_m: Optional[float]) -> int:
    """Apply the user's DC fixed-length formula.

    Take the DC segment length (metres), add a 10 m buffer, and round **up**
    to the nearest standard drum length from ``DC_STANDARD_LENGTHS``.

    Excel equivalent::

        =IF(OR(A239=0, A239=""), 0,
            INDEX({50,80,100,120,150,180,200,250,300,350},
                  MATCH(TRUE, {50,80,100,120,150,180,200,250,300,350}>=A239+10, 0)))

    Args:
        length_m: Measured drop-cable segment length in metres. ``None`` or
            ``0`` means no cable is required.

    Returns:
        The standard drum length (e.g. 50, 80, 100 … 350) or ``0``.
    """
    if length_m is None:
        return 0
    try:
        length_val = float(length_m)
    except (TypeError, ValueError):
        return 0

    if length_val <= 0:
        return 0

    buffered = length_val + 10.0

    # Find the first standard length >= buffered (round UP)
    for std in DC_STANDARD_LENGTHS:
        if std >= buffered:
            return std

    # If buffered exceeds all standard lengths, return the largest
    return DC_STANDARD_LENGTHS[-1]


def get_fibre_count(cable_size: str) -> int:
    """Return the fibre count for a given cable size string.

    Args:
        cable_size: One of the keys in ``FIBRE_COUNT_MAP`` (e.g. ``"48F"``).

    Returns:
        Number of fibres.

    Raises:
        ValueError: If *cable_size* is not recognised.
    """
    if cable_size not in FIBRE_COUNT_MAP:
        raise ValueError(
            f"Invalid cable_size '{cable_size}'. Expected one of {VALID_CABLE_SIZES}"
        )
    return FIBRE_COUNT_MAP[cable_size]


def get_feeder_part_by_size(cable_size: str) -> str:
    """Map cable size (e.g. ``"48F"``) to the feeder-cable part number.

    Args:
        cable_size: Cable size designation.

    Returns:
        Part number string.

    Raises:
        ValueError: If *cable_size* is not recognised.
    """
    if cable_size not in FEEDER_PART_MAP:
        raise ValueError(
            f"Invalid cable_size '{cable_size}'. Expected one of {VALID_CABLE_SIZES}"
        )
    return FEEDER_PART_MAP[cable_size]


def get_tangent_part_by_size(cable_size: str) -> str:
    """Map cable size to tangent-clamp part number.

    Args:
        cable_size: Cable size designation.

    Returns:
        Part number string.
    """
    if cable_size not in TANGENT_PART_MAP:
        raise ValueError(
            f"Invalid cable_size '{cable_size}'. Expected one of {VALID_CABLE_SIZES}"
        )
    return TANGENT_PART_MAP[cable_size]


def get_deadend_part_by_size(cable_size: str) -> str:
    """Map cable size to dead-end clamp part number.

    Args:
        cable_size: Cable size designation.

    Returns:
        Part number string.
    """
    if cable_size not in DEADEND_PART_MAP:
        raise ValueError(
            f"Invalid cable_size '{cable_size}'. Expected one of {VALID_CABLE_SIZES}"
        )
    return DEADEND_PART_MAP[cable_size]


def get_drop_cable_part(fixed_length: int) -> Optional[str]:
    """Return the part number for a drop cable of a given fixed length.

    Args:
        fixed_length: One of the standard lengths (50, 80, … 350).

    Returns:
        Part number or ``None`` if the length is not standard.
    """
    return DROP_CABLE_PART_MAP.get(fixed_length)


def get_dj_box_part(ratio: str) -> Optional[str]:
    """Return the DJ box part number for a given splitter ratio.

    Args:
        ratio: Splitter ratio string such as ``"1:8"``, ``"1:9"``,
            ``"1x8"``, ``"1x9"``.

    Returns:
        Part number or ``None``.
    """
    return DJ_BOX_PART_MAP.get(ratio)


def _get_part_info(part_no: str) -> Dict[str, str]:
    """Return catalog entry for a part number.

    Args:
        part_no: Material part number.

    Returns:
        Dictionary with ``description``, ``unit``, ``category``,
        ``sub_category``.

    Raises:
        KeyError: If the part number is not in the catalog.
    """
    if part_no not in MATERIAL_CATALOG:
        raise KeyError(f"Part number '{part_no}' not found in MATERIAL_CATALOG")
    return MATERIAL_CATALOG[part_no]


# ---------------------------------------------------------------------------
# Splice Counting
# ---------------------------------------------------------------------------

def count_splices_from_plan(splicing_data: List[Dict[str, Any]], cable_size: str) -> Dict[str, Any]:
    """Count fusion splices required from a splicing plan.

    Rules:
    * Every feeder-cable (FC) segment that enters or exits a closure
      (FJ/AGG) requires fusion splices equal to the fibre count.
    * The pre-conventional side (DJ -> FJ) uses mechanical / quick-connect;
      **no** fusion splices are counted there.
    * The conventional side (FJ -> OLT) requires fusion splices.

    The splicing plan rows are expected to contain at least:
    ``ag_name``, ``block_name``, ``dj_name``, ``dc_name``, ``from_node``,
    ``to_node``, ``cable_type`` (``"FC"`` or ``"DC"``).

    Args:
        splicing_data: List of row dictionaries from the splicing-plan CSV.
        cable_size: Feeder-cable size (determines fibres per splice).

    Returns:
        Dictionary with keys:

        * ``total_splices`` – overall fusion-splice count.
        * ``fc_splices`` – splices attributed to feeder-cable segments.
        * ``core_splices`` – splices at the core-cable / AGG level.
        * ``per_ag`` – dict ``{ag_name: splice_count}``.
    """
    fibres = get_fibre_count(cable_size)

    fc_splices = 0
    core_splices = 0
    per_ag: Dict[str, int] = {}

    if not splicing_data:
        return {
            "total_splices": 0,
            "fc_splices": 0,
            "core_splices": 0,
            "per_ag": {},
        }

    for row in splicing_data:
        cable_type = str(row.get("cable_type", "")).strip().upper()
        ag_name = str(row.get("ag_name", "unknown")).strip()

        if ag_name not in per_ag:
            per_ag[ag_name] = 0

        if cable_type == "FC":
            # Each FC segment entering/exiting a closure needs fibres-per-segment splices
            fc_splices += fibres
            per_ag[ag_name] += fibres
        elif cable_type == "CORE":
            # Core cable splices at AGG level
            core_splices += fibres
            per_ag[ag_name] += fibres
        # DC cables on pre-conventional side use mechanical quick-connect: no fusion splices

    total_splices = fc_splices + core_splices

    return {
        "total_splices": total_splices,
        "fc_splices": fc_splices,
        "core_splices": core_splices,
        "per_ag": per_ag,
    }


# ---------------------------------------------------------------------------
# BOM Line-Item Builder
# ---------------------------------------------------------------------------

def _line_item(part_no: str, qty: float, notes: str = "") -> Dict[str, Any]:
    """Build a BOM line-item dictionary.

    Args:
        part_no: Material part number.
        qty: Required quantity (will be rounded up to whole units for
            ``unit == "EACH"``).
        notes: Optional note string.

    Returns:
        Dictionary with keys ``part_no``, ``description``, ``qty``, ``unit``,
        ``notes``.
    """
    info = _get_part_info(part_no)
    unit = info["unit"]

    # Round up to whole units for discrete items
    if unit == "EACH":
        qty = int(qty) if qty == int(qty) else int(qty) + 1
        qty = max(0, qty)
    else:
        # For km / metre lengths keep as float rounded to 3 dp
        qty = round(float(qty), 3)

    return {
        "part_no": part_no,
        "description": info["description"],
        "qty": qty,
        "unit": unit,
        "notes": notes,
    }


# ---------------------------------------------------------------------------
# Category Calculation Helpers
# ---------------------------------------------------------------------------

def _calc_pre_conventional(data: Dict[str, Any], config: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Calculate pre-conventional (DJ-side) material line items.

    Args:
        data: Network topology data (see :func:`calculate_bom`).
        config: Design configuration (see :func:`calculate_bom`).

    Returns:
        List of line-item dictionaries.
    """
    items: List[Dict[str, Any]] = []

    djs = data.get("djs", {})
    fjs = data.get("fjs", {})
    houses = data.get("houses", {})
    dc_segments = data.get("dc_segments", [])

    # --- DJ Boxes ---
    dj_1x9_count = 0
    dj_1x8_count = 0

    for dj_name, dj_info in djs.items():
        ratio = str(dj_info.get("ratio", "1:8")).strip()
        # Normalize ratio string
        ratio_clean = ratio.replace("x", ":").replace("X", ":")
        if ratio_clean in ("1:9",):
            dj_1x9_count += 1
        else:
            # Default to 1:8
            dj_1x8_count += 1

    if dj_1x9_count > 0:
        items.append(_line_item("TRA00166", dj_1x9_count, f"DJ boxes 1:9 for {dj_1x9_count} DJs"))
    if dj_1x8_count > 0:
        items.append(_line_item("TRA00167", dj_1x8_count, f"DJ boxes 1:8 for {dj_1x8_count} DJs"))

    # --- FJ Closures ---
    num_fjs = len(fjs)
    if num_fjs > 0:
        items.append(_line_item("CLO00172", num_fjs, f"FJ closures for {num_fjs} FJs"))

    # --- Drop Cables ---
    # Apply fixed-length formula to each DC segment, then bin by length
    dc_length_counts: Dict[int, int] = {}
    for seg in dc_segments:
        length_m = seg.get("length_m", 0)
        fixed_len = calculate_dc_fixed_length(length_m)
        if fixed_len > 0:
            dc_length_counts[fixed_len] = dc_length_counts.get(fixed_len, 0) + 1

    for length, count in sorted(dc_length_counts.items()):
        part = get_drop_cable_part(length)
        if part:
            items.append(_line_item(part, count, f"Drop cable {length}m x{count}"))

    total_drop_cables = sum(dc_length_counts.values())

    # --- Pigtails (SC/APC) ---
    # One pigtail per drop cable at the FJ side + 10% spare
    if total_drop_cables > 0:
        pigtail_qty = int(total_drop_cables * 1.1) + 1
        items.append(_line_item("CAB00326", pigtail_qty, f"Pigtails for {total_drop_cables} drop cables + 10% spare"))

    # --- Mid Couplers ---
    if total_drop_cables > 0:
        coupler_qty = total_drop_cables  # One mid coupler per drop cable
        items.append(_line_item("TRA00115", coupler_qty, f"Mid couplers for {total_drop_cables} drop cables"))

    # --- Splice Protectors 1.3mm ---
    # These are used on the conventional side for FC splices, not pre-conventional
    # (pre-conventional uses mechanical sealing). Include a minimal quantity
    # for any conventional-side work inside FJ closures.
    splice_info = data.get("_splice_info", {})
    total_splices = splice_info.get("total_splices", 0)
    if total_splices > 0:
        items.append(_line_item("CLO00137", 0, "Splice protectors 1.3mm - see Conventional sheet"))

    return items


def _calc_conventional(data: Dict[str, Any], config: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Calculate conventional (OLT-side) material line items.

    Args:
        data: Network topology data.
        config: Design configuration.

    Returns:
        List of line-item dictionaries.
    """
    items: List[Dict[str, Any]] = []

    cable_size = config.get("cable_size", "48F")
    fibres_per_fj = config.get("fibers_per_fj", 4)
    fc_info = data.get("fc_info", [])
    ags = data.get("ags", {})
    fjs = data.get("fjs", {})
    splicing_plan = data.get("splicing_plan", [])

    # --- Feeder Cable ---
    total_fc_length_km = 0.0
    for fc in fc_info:
        length_m = float(fc.get("length_m", 0))
        total_fc_length_km += length_m / 1000.0

    # Add core cable length (from AG to OLT aggregation point)
    core_length_per_ag_km = config.get("core_length_per_ag_km", 2.0)
    num_ags = len(ags)
    total_core_length_km = num_ags * core_length_per_ag_km

    total_feeder_km = total_fc_length_km + total_core_length_km

    if total_feeder_km > 0:
        feeder_part = get_feeder_part_by_size(cable_size)
        items.append(_line_item(feeder_part, round(total_feeder_km, 3),
                                f"Total feeder {total_fc_length_km:.3f}km + core {total_core_length_km:.3f}km"))

    # --- Dome Joints ---
    # One dome joint per AG (aggregation manhole)
    if num_ags > 0:
        items.append(_line_item("CLO00110", num_ags, f"Dome joints for {num_ags} AGs"))

    # --- Splitter Trays ---
    num_fjs = len(fjs)
    if num_fjs > 0:
        # Typically 1 tray per FJ for splitter accommodation
        items.append(_line_item("CLO00138", num_fjs, f"Splitter trays for {num_fjs} FJs"))

    # --- Splitters ---
    # 1x8 splitters: one per FJ (each FJ serves multiple DJs via 1x8)
    if num_fjs > 0:
        items.append(_line_item("TRA00077", num_fjs, f"1x8 splitters for {num_fjs} FJs"))

    # 1x2 splitters: for cascading / sub-splitting if needed
    # Estimate: one per AG
    if num_ags > 0:
        items.append(_line_item("TRA00142", num_ags, f"1x2 splitters for {num_ags} AGs"))

    # --- Splice Protectors 2.2mm ---
    splice_counts = count_splices_from_plan(splicing_plan, cable_size)
    total_splices = splice_counts["total_splices"]
    # Store splice info for other categories
    data["_splice_info"] = splice_counts

    if total_splices > 0:
        # Add 10% spare
        protector_qty = int(total_splices * 1.1) + 1
        items.append(_line_item("CLO00016", protector_qty,
                                f"Splice protectors for {total_splices} splices + 10% spare"))

    # --- Patch Cords ---
    # LC/APC to SC/UPC: for ODF connections
    fibre_count = get_fibre_count(cable_size)
    if fibre_count > 0:
        patch_qty = max(4, fibre_count // 12)  # Minimum 4, or 1 per 12 fibres
        items.append(_line_item("CAB00188", patch_qty, f"LC/SC patch cords for ODF"))
        items.append(_line_item("CAB00189", patch_qty, f"LC/LC patch cords for ODF"))

    # --- Patch Panels ---
    # One patch panel per AG, minimum 1
    if num_ags > 0:
        items.append(_line_item("TRA00031", num_ags, f"Patch panels for {num_ags} AGs"))

    # --- ODF Panels ---
    # One ODF panel per AG for fibre management
    if num_ags > 0:
        items.append(_line_item("TRA00035", num_ags, f"ODF panels for {num_ags} AGs"))

    # --- Quad Mid Couplers ---
    if num_fjs > 0:
        items.append(_line_item("TRA00122", num_fjs, f"Quad mid couplers for {num_fjs} FJs"))

    return items


def _calc_pole_hardware(data: Dict[str, Any], config: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Calculate pole hardware material line items.

    Args:
        data: Network topology data.
        config: Design configuration.

    Returns:
        List of line-item dictionaries.
    """
    items: List[Dict[str, Any]] = []

    cable_size = config.get("cable_size", "48F")
    poles = data.get("poles", [])
    fjs = data.get("fjs", {})
    djs = data.get("djs", {})
    fc_info = data.get("fc_info", [])

    if not poles:
        return items

    # Count poles by type
    total_poles = len(poles)

    # Poles with feeder cable (FC) running through them need tangent clamps
    # Assume all poles with "fc" or "aerial" type have FC
    poles_with_fc = sum(1 for p in poles.values() if p.get("pole_type", "").lower() in ("fc", "aerial", "mixed"))

    # Dead-end poles: poles where FC terminates or turns
    # Estimate: ~20% of poles are dead-ends
    dead_end_poles = max(1, int(poles_with_fc * 0.2)) if poles_with_fc > 0 else 0

    # Tangent clamps: 1 per FC pole (holding the cable in tangent position)
    if poles_with_fc > 0:
        tangent_part = get_tangent_part_by_size(cable_size)
        items.append(_line_item(tangent_part, poles_with_fc,
                                f"Tangent clamps for {poles_with_fc} FC poles"))

    # Dead-end clamps: for poles where FC terminates
    if dead_end_poles > 0:
        deadend_part = get_deadend_part_by_size(cable_size)
        items.append(_line_item(deadend_part, dead_end_poles,
                                f"Dead-end clamps for {dead_end_poles} dead-end poles"))

    # V-shape slack brackets: for slack storage points (~1 per 5 poles)
    v_shape_qty = max(1, total_poles // 5) if total_poles > 0 else 0
    if v_shape_qty > 0:
        items.append(_line_item("PAC00054", v_shape_qty,
                                f"V-shape slack brackets (~1 per 5 poles)"))

    # 3-way hook brackets: for poles with multiple cable directions
    # Estimate: ~30% of poles
    hook_qty = max(1, int(total_poles * 0.3)) if total_poles > 0 else 0
    if hook_qty > 0:
        items.append(_line_item("PAC00064", hook_qty,
                                f"3-way hook brackets for {hook_qty} poles"))

    # Slack storage brackets: for aerial slack management
    slack_qty = max(1, total_poles // 10) if total_poles > 0 else 0
    if slack_qty > 0:
        items.append(_line_item("PAC00068", slack_qty,
                                f"Aerial slack storage brackets"))
        items.append(_line_item("PAC00078", slack_qty,
                                f"Huawei aerial slack storage brackets"))

    # Pole mounting assemblies: for poles with equipment (DJ or FJ)
    # Each DJ and FJ needs a mounting assembly
    equip_on_poles = len(djs) + len(fjs)
    if equip_on_poles > 0:
        items.append(_line_item("PAC00077", equip_on_poles,
                                f"Pole mounting assemblies for {equip_on_poles} equipment mounts"))

    # Universal brackets: for closures on poles
    num_fjs = len(fjs)
    if num_fjs > 0:
        items.append(_line_item("CLO00177", num_fjs,
                                f"Universal brackets for {num_fjs} FJs"))

    return items


def _calc_consumables(data: Dict[str, Any], config: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Calculate consumables material line items.

    Args:
        data: Network topology data.
        config: Design configuration.

    Returns:
        List of line-item dictionaries.
    """
    items: List[Dict[str, Any]] = []

    poles = data.get("poles", [])
    fjs = data.get("fjs", {})
    ags = data.get("ags", {})
    djs = data.get("djs", {})
    dc_segments = data.get("dc_segments", [])
    fc_info = data.get("fc_info", [])
    splicing_plan = data.get("splicing_plan", [])
    cable_size = config.get("cable_size", "48F")

    total_poles = len(poles)
    num_fjs = len(fjs)
    num_ags = len(ags)
    num_djs = len(djs)
    total_fc_segments = len(fc_info)
    total_dc_segments = len(dc_segments)

    # Get splice counts
    splice_counts = count_splices_from_plan(splicing_plan, cable_size)
    total_splices = splice_counts["total_splices"]

    # --- Alcohol Spray ---
    # ~1 bottle per 50 splices, minimum 1
    if total_splices > 0:
        alcohol_qty = max(1, total_splices // 50 + 1)
        items.append(_line_item("CLE00004", alcohol_qty,
                                f"Alcohol spray for {total_splices} splices"))

    # --- Kim Wipes ---
    # ~1 box per 100 splices, minimum 1
    if total_splices > 0:
        wipes_qty = max(1, total_splices // 100 + 1)
        items.append(_line_item("CLE00006", wipes_qty,
                                f"Kim wipes for {total_splices} splices"))

    # --- Cable Ties (large) ---
    # ~10 per pole + 5 per FC segment
    if total_poles > 0 or total_fc_segments > 0:
        ties_large = total_poles * 10 + total_fc_segments * 5
        ties_large = max(10, ties_large)
        items.append(_line_item("GEN00007", ties_large,
                                f"Large cable ties for {total_poles} poles + {total_fc_segments} FC segments"))

    # --- Cable Ties (small) ---
    # ~5 per DJ + 2 per DC segment
    if num_djs > 0 or total_dc_segments > 0:
        ties_small = num_djs * 5 + total_dc_segments * 2
        ties_small = max(10, ties_small)
        items.append(_line_item("GEN00116", ties_small,
                                f"Small cable ties for {num_djs} DJs + {total_dc_segments} DC segments"))

    # --- Nitto Tape ---
    # ~1 roll per 5 closures (FJ + AG dome joints)
    total_closures = num_fjs + num_ags
    if total_closures > 0:
        nitto_qty = max(1, total_closures // 5 + 1)
        items.append(_line_item("GEN00035", nitto_qty,
                                f"Nitto tape for {total_closures} closures"))

    # --- Velcro ---
    # ~1 roll per AG for cable bundling
    if num_ags > 0:
        items.append(_line_item("GEN00130", num_ags,
                                f"Velcro for {num_ags} AGs"))

    # --- Bandit Buckles ---
    # ~2 per pole
    if total_poles > 0:
        buckle_qty = total_poles * 2
        items.append(_line_item("GEN00064", buckle_qty,
                                f"Bandit buckles for {total_poles} poles"))

    # --- Bandit Straps ---
    # ~1 roll per 10 poles (30m roll)
    if total_poles > 0:
        strap_qty = max(1, total_poles // 10 + 1)
        items.append(_line_item("GEN00202", strap_qty,
                                f"Bandit straps for {total_poles} poles"))

    # --- Bunny Clips ---
    # ~2 per DC segment + 1 per DJ
    if total_dc_segments > 0 or num_djs > 0:
        bunny_qty = total_dc_segments * 2 + num_djs
        bunny_qty = max(10, bunny_qty)
        items.append(_line_item("GEN00087", bunny_qty,
                                f"Bunny clips for {total_dc_segments} DC + {num_djs} DJ"))

    # --- Dartag Sleeves ---
    # ~1 per 2 poles for cable identification
    if total_poles > 0:
        dartag_qty = max(10, total_poles // 2 + 1)
        items.append(_line_item("GEN00134", dartag_qty,
                                f"Dartag sleeves for {total_poles} poles"))

    # --- Vinyl Cartridges ---
    # ~1 per AG for labeling
    if num_ags > 0:
        items.append(_line_item("GEN00170", num_ags,
                                f"Vinyl cartridges for {num_ags} AGs"))

    # --- Caution Tape ---
    # ~1 per AG for trenching
    route_type = config.get("route_type", "aerial")
    if route_type in ("underground", "mixed") and num_ags > 0:
        items.append(_line_item("GEN00191", num_ags,
                                f"Caution tape for {route_type} route"))

    # --- Oval Port Kits ---
    # For dome joints: 1 per AG
    if num_ags > 0:
        items.append(_line_item("CLO00015", num_ags,
                                f"Oval port kits for {num_ags} dome joints"))

    # --- Mechanical Seals (medium entry) ---
    # For FC entry into closures: 2 per FJ + 2 per AG
    mech_seal_qty = num_fjs * 2 + num_ags * 2
    if mech_seal_qty > 0:
        items.append(_line_item("CLO00035", mech_seal_qty,
                                f"Mech seals for FC entry"))

    # --- Mechanical Seals (quad) ---
    # For DC entry into FJs: 1 per FJ (holds up to 4 DCs)
    if num_fjs > 0:
        quad_seal_qty = max(1, num_fjs)
        items.append(_line_item("CLO00039", quad_seal_qty,
                                f"Quad mech seals for DC entry"))

    # --- Cage Nuts & Screws ---
    # For rack-mounted equipment: 4 per patch panel
    if num_ags > 0:
        cage_nut_qty = num_ags * 8  # 2 panels per AG x 4 nuts
        items.append(_line_item("BOL00002", cage_nut_qty,
                                f"Cage nuts for rack equipment"))

    # --- Nyloc Nuts ---
    # For pole hardware: 2 per pole
    if total_poles > 0:
        nyloc_qty = total_poles * 2
        items.append(_line_item("BOL00056", nyloc_qty,
                                f"Nyloc nuts for pole hardware"))

    # --- Bolts ---
    # For pole hardware: 2 per pole
    if total_poles > 0:
        bolt_qty = total_poles * 2
        items.append(_line_item("BOL00059", bolt_qty,
                                f"Bolts for pole hardware"))

    return items


def _calc_other(data: Dict[str, Any], config: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Calculate 'other' category material line items.

    Args:
        data: Network topology data.
        config: Design configuration.

    Returns:
        List of line-item dictionaries.
    """
    items: List[Dict[str, Any]] = []

    poles = data.get("poles", [])
    ags = data.get("ags", {})

    total_poles = len(poles)
    num_ags = len(ags)

    # --- GLAM (VUMA Node) ---
    # One per AG / aggregation point
    if num_ags > 0:
        items.append(_line_item("GLA00032", num_ags,
                                f"GLAM nodes for {num_ags} AGs"))

    # --- Poles ---
    # Count by pole length from pole_type
    poles_6m = sum(1 for p in poles.values() if "6" in str(p.get("pole_type", "")))
    poles_7m = sum(1 for p in poles.values() if "7" in str(p.get("pole_type", "")))
    poles_9m = sum(1 for p in poles.values() if "9" in str(p.get("pole_type", "")))

    # Default distribution if no specific lengths given
    if poles_6m == 0 and poles_7m == 0 and poles_9m == 0 and total_poles > 0:
        # Default: 70% 7m, 20% 9m, 10% 6m
        poles_7m = int(total_poles * 0.7)
        poles_9m = int(total_poles * 0.2)
        poles_6m = total_poles - poles_7m - poles_9m

    if poles_6m > 0:
        items.append(_line_item("POL00017", poles_6m, f"6m poles"))
    if poles_7m > 0:
        items.append(_line_item("POL00013", poles_7m, f"7m poles"))
    if poles_9m > 0:
        items.append(_line_item("POL00015", poles_9m, f"9m poles"))

    # --- Ducting (for underground / mixed routes) ---
    route_type = config.get("route_type", "aerial")
    if route_type in ("underground", "mixed") and num_ags > 0:
        # Ripple duct: ~2 lengths per AG
        items.append(_line_item("TUB00026", num_ags * 2,
                                f"Ripple duct for {route_type}"))
        # Half-round pipe: ~1 per AG
        items.append(_line_item("TUB00028", num_ags,
                                f"Half-round pipe for {route_type}"))
        # Sub-duct: ~2 per AG
        items.append(_line_item("TUB00128", num_ags * 2,
                                f"Sub-duct for {route_type}"))

    # --- Smart Locks ---
    # One per AG manhole
    if num_ags > 0:
        items.append(_line_item("S001219", num_ags,
                                f"Smart locks for {num_ags} AG manholes"))

    # --- Tools ---
    # One-time tooling per project (not per unit)
    items.append(_line_item("TOO00280", 1, "Wedge clamping tool (project)"))
    items.append(_line_item("TOO00281", 1, "Plum ring hook tool (project)"))

    return items


# ---------------------------------------------------------------------------
# Main BOM Calculation
# ---------------------------------------------------------------------------

def calculate_bom(data: Dict[str, Any], config: Dict[str, Any]) -> Dict[str, Any]:
    """Calculate the complete Bill of Materials for an FTTH network design.

    This is the main entry point. It delegates to category-specific helpers
    and returns a structured BOM dictionary ready for Excel export.

    Args:
        data: Network topology and design data. Expected keys:

            - ``djs`` – dict ``{dj_name: {feat_id, position, ratio, block_name, ag_name, dc_name}}``
            - ``blocks`` – dict ``{block_name: {block_id, ag_name, num_djs}}``
            - ``ags`` – dict ``{ag_name: {num_fjs, num_blocks}}``
            - ``fjs`` – dict ``{fj_name: {ag_name}}``
            - ``poles`` – list of pole info ``[{feat_id, pole_type}]``
            - ``fc_info`` – list of feeder-cable info ``[{fc_name, length_m, from_ag, to_ag}]``
            - ``dc_segments`` – list of DC segment info ``[{dc_name, from_node, to_node, length_m, block_id, ag_name}]``
            - ``houses`` – dict ``{dj_name: [house_names]}``
            - ``splicing_plan`` – list of splicing-plan row dicts

        config: Design configuration. Expected keys:

            - ``cable_size`` – str (e.g. ``"48F"``, ``"144F"``, ``"288F"``)
            - ``fibers_per_fj`` – int (default ``4``)
            - ``core_length_per_ag_km`` – float (default ``2.0``)
            - ``route_type`` – str (``"aerial"``, ``"underground"``, ``"mixed"``)

    Returns:
        Dictionary with keys:

        - ``pre_conventional`` – list of line items
        - ``conventional`` – list of line items
        - ``pole_hardware`` – list of line items
        - ``consumables`` – list of line items
        - ``other`` – list of line items
        - ``summary`` – dict with category totals and grand total
    """
    # Validate cable_size
    cable_size = config.get("cable_size", "48F")
    if cable_size not in FIBRE_COUNT_MAP:
        raise ValueError(
            f"Invalid cable_size '{cable_size}'. Expected one of {VALID_CABLE_SIZES}"
        )

    # Calculate each category
    pre_conv_items = _calc_pre_conventional(data, config)
    conv_items = _calc_conventional(data, config)
    pole_items = _calc_pole_hardware(data, config)
    consumable_items = _calc_consumables(data, config)
    other_items = _calc_other(data, config)

    # Build summary
    def _category_total(items: List[Dict[str, Any]]) -> int:
        return sum(item["qty"] for item in items if isinstance(item["qty"], int))

    def _category_total_float(items: List[Dict[str, Any]]) -> float:
        return sum(
            item["qty"] for item in items if isinstance(item["qty"], (int, float))
        )

    summary = {
        "pre_conventional_count": len(pre_conv_items),
        "pre_conventional_qty": _category_total_float(pre_conv_items),
        "conventional_count": len(conv_items),
        "conventional_qty": _category_total_float(conv_items),
        "pole_hardware_count": len(pole_items),
        "pole_hardware_qty": _category_total_float(pole_items),
        "consumables_count": len(consumable_items),
        "consumables_qty": _category_total_float(consumable_items),
        "other_count": len(other_items),
        "other_qty": _category_total_float(other_items),
        "grand_total_line_items": (
            len(pre_conv_items) + len(conv_items) + len(pole_items)
            + len(consumable_items) + len(other_items)
        ),
        "cable_size": cable_size,
        "total_djs": len(data.get("djs", {})),
        "total_fjs": len(data.get("fjs", {})),
        "total_ags": len(data.get("ags", {})),
        "total_poles": len(data.get("poles", [])),
        "total_houses": sum(
            len(h) for h in data.get("houses", {}).values()
        ),
    }

    # Add splice summary if available
    splice_info = data.get("_splice_info", {})
    if splice_info:
        summary["total_splices"] = splice_info.get("total_splices", 0)
        summary["fc_splices"] = splice_info.get("fc_splices", 0)
        summary["core_splices"] = splice_info.get("core_splices", 0)

    return {
        "pre_conventional": pre_conv_items,
        "conventional": conv_items,
        "pole_hardware": pole_items,
        "consumables": consumable_items,
        "other": other_items,
        "summary": summary,
    }


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------

def generate_bom_excel(
    bom: Dict[str, Any],
    output_path: str,
    area_code: str = "",
    zone: str = "",
) -> str:
    """Generate an Excel workbook from BOM data.

    Creates a workbook with 5 sheets (one per category) plus a summary sheet.
    Each sheet contains: row number, part number, description, quantity, unit,
    and notes.  The header row is bold with a light-grey fill, auto-filter is
    enabled, the top row is frozen, and columns are auto-sized.

    Args:
        bom: Output dictionary from :func:`calculate_bom`.
        output_path: Filesystem path for the ``.xlsx`` file.
        area_code: Optional area code string for the title/header.
        zone: Optional zone string for the title/header.

    Returns:
        Absolute path to the written Excel file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for Excel export. Install it via: pip install openpyxl"
        ) from exc

    wb = openpyxl.Workbook()

    # Remove default sheet; we'll create our own
    wb.remove(wb.active)

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    categories = [
        ("pre_conventional", "Pre-Conventional"),
        ("conventional", "Conventional"),
        ("pole_hardware", "Pole Hardware"),
        ("consumables", "Consumables"),
        ("other", "Other"),
    ]

    headers = ["#", "Part No", "Description", "Qty", "Unit", "Notes"]

    for cat_key, cat_title in categories:
        ws = wb.create_sheet(title=cat_title)
        items = bom.get(cat_key, [])

        # Title row
        title_text = f"BOM - {cat_title}"
        if area_code:
            title_text += f" | Area: {area_code}"
        if zone:
            title_text += f" | Zone: {zone}"

        ws.merge_cells("A1:F1")
        ws["A1"] = title_text
        ws["A1"].font = Font(bold=True, size=13)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        # Header row (row 2)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        for row_idx, item in enumerate(items, 3):
            ws.cell(row=row_idx, column=1, value=row_idx - 2).border = thin_border
            ws.cell(row=row_idx, column=2, value=item.get("part_no", "")).border = thin_border
            ws.cell(row=row_idx, column=3, value=item.get("description", "")).border = thin_border
            ws.cell(row=row_idx, column=4, value=item.get("qty", 0)).border = thin_border
            ws.cell(row=row_idx, column=5, value=item.get("unit", "")).border = thin_border
            ws.cell(row=row_idx, column=6, value=item.get("notes", "")).border = thin_border

        # Auto-filter
        if len(items) > 0:
            ws.auto_filter.ref = f"A2:F{len(items) + 2}"
        else:
            ws.auto_filter.ref = "A2:F2"

        # Freeze top 2 rows (title + header)
        ws.freeze_panes = "A3"

        # Auto-width columns (approximate)
        col_widths = [6, 14, 70, 10, 8, 35]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Quantity column right-aligned
        for row_idx in range(3, len(items) + 3):
            ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="right")

    # --- Summary Sheet ---
    summary_ws = wb.create_sheet(title="Summary", index=0)
    summary = bom.get("summary", {})

    title_text = "BOM Summary"
    if area_code:
        title_text += f" | Area: {area_code}"
    if zone:
        title_text += f" | Zone: {zone}"

    summary_ws.merge_cells("A1:D1")
    summary_ws["A1"] = title_text
    summary_ws["A1"].font = Font(bold=True, size=14)
    summary_ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    summary_ws.row_dimensions[1].height = 28

    # Summary headers
    sum_headers = ["Category", "Line Items", "Total Qty", "Notes"]
    for col_idx, header in enumerate(sum_headers, 1):
        cell = summary_ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    summary_rows = [
        ("Pre-Conventional", summary.get("pre_conventional_count", 0), summary.get("pre_conventional_qty", 0), "DJ boxes, drop cables, pigtails, couplers"),
        ("Conventional", summary.get("conventional_count", 0), summary.get("conventional_qty", 0), "Feeder cable, joints, splitters, patch cords"),
        ("Pole Hardware", summary.get("pole_hardware_count", 0), summary.get("pole_hardware_qty", 0), "Clamps, brackets, mounting assemblies"),
        ("Consumables", summary.get("consumables_count", 0), summary.get("consumables_qty", 0), "Ties, tape, seals, cleaning supplies"),
        ("Other", summary.get("other_count", 0), summary.get("other_qty", 0), "Poles, duct, tools, GLAM nodes"),
    ]

    for row_idx, (cat, count, qty, note) in enumerate(summary_rows, 3):
        summary_ws.cell(row=row_idx, column=1, value=cat).border = thin_border
        summary_ws.cell(row=row_idx, column=2, value=count).border = thin_border
        summary_ws.cell(row=row_idx, column=3, value=qty).border = thin_border
        summary_ws.cell(row=row_idx, column=4, value=note).border = thin_border

    # Totals row
    total_row = len(summary_rows) + 3
    summary_ws.cell(row=total_row, column=1, value="GRAND TOTAL").border = thin_border
    summary_ws.cell(row=total_row, column=1).font = Font(bold=True)
    summary_ws.cell(row=total_row, column=2, value=summary.get("grand_total_line_items", 0)).border = thin_border
    summary_ws.cell(row=total_row, column=2).font = Font(bold=True)
    summary_ws.cell(row=total_row, column=3, value="").border = thin_border
    summary_ws.cell(row=total_row, column=4, value="").border = thin_border

    # Network stats section
    stats_start = total_row + 2
    summary_ws.cell(row=stats_start, column=1, value="Network Statistics").font = Font(bold=True, size=12)
    summary_ws.merge_cells(start_row=stats_start, start_column=1, end_row=stats_start, end_column=4)

    stats = [
        ("Cable Size", summary.get("cable_size", "N/A")),
        ("Total D Js", summary.get("total_djs", 0)),
        ("Total F Js", summary.get("total_fjs", 0)),
        ("Total AGs", summary.get("total_ags", 0)),
        ("Total Poles", summary.get("total_poles", 0)),
        ("Total Houses", summary.get("total_houses", 0)),
        ("Total Splices", summary.get("total_splices", "N/A")),
        ("FC Splices", summary.get("fc_splices", "N/A")),
        ("Core Splices", summary.get("core_splices", "N/A")),
    ]

    for row_idx, (label, value) in enumerate(stats, stats_start + 1):
        summary_ws.cell(row=row_idx, column=1, value=label).border = thin_border
        summary_ws.cell(row=row_idx, column=1).font = Font(bold=True)
        summary_ws.cell(row=row_idx, column=2, value=value).border = thin_border
        summary_ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=4)

    # Summary sheet formatting
    summary_ws.column_dimensions["A"].width = 22
    summary_ws.column_dimensions["B"].width = 14
    summary_ws.column_dimensions["C"].width = 14
    summary_ws.column_dimensions["D"].width = 50
    summary_ws.freeze_panes = "A3"
    summary_ws.auto_filter.ref = f"A2:D{total_row}"

    # Save
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)

    return os.path.abspath(output_path)


# ---------------------------------------------------------------------------
# CSV Splicing Plan Loader
# ---------------------------------------------------------------------------

def load_splicing_plan(csv_path: str) -> List[Dict[str, Any]]:
    """Load a splicing-plan CSV file into a list of row dictionaries.

    Expected columns (case-insensitive matching):
    ``ag_name``, ``block_name``, ``dj_name``, ``dc_name``, ``from_node``,
    ``to_node``, ``cable_type``.

    Args:
        csv_path: Path to the CSV file.

    Returns:
        List of row dictionaries.
    """
    rows: List[Dict[str, Any]] = []
    column_map = {
        "ag_name": ["ag_name", "ag", "aggregation", "ag_name"],
        "block_name": ["block_name", "block", "block_id"],
        "dj_name": ["dj_name", "dj", "distribution_joint"],
        "dc_name": ["dc_name", "dc", "drop_cable"],
        "from_node": ["from_node", "from", "source"],
        "to_node": ["to_node", "to", "destination"],
        "cable_type": ["cable_type", "type", "cable"],
    }

    with open(csv_path, "r", newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        if not reader.fieldnames:
            return rows

        # Build reverse lookup: lowercase header -> canonical key
        header_lookup: Dict[str, str] = {}
        for canonical, aliases in column_map.items():
            for alias in aliases:
                header_lookup[alias.lower()] = canonical

        field_lookup: Dict[str, str] = {}
        for field in reader.fieldnames:
            field_lower = field.strip().lower()
            if field_lower in header_lookup:
                field_lookup[field] = header_lookup[field_lower]

        for raw_row in reader:
            row: Dict[str, Any] = {}
            for raw_field, canonical in field_lookup.items():
                row[canonical] = raw_row.get(raw_field, "").strip()
            rows.append(row)

    return rows


# ---------------------------------------------------------------------------
# Quick Self-Test
# ---------------------------------------------------------------------------

def _self_test() -> None:
    """Run a quick sanity check on the module's core functions."""
    # DC fixed length tests
    assert calculate_dc_fixed_length(0) == 0
    assert calculate_dc_fixed_length(None) == 0
    assert calculate_dc_fixed_length(35) == 50   # 35+10=45 -> round up to 50
    assert calculate_dc_fixed_length(40) == 50   # 40+10=50 -> 50
    assert calculate_dc_fixed_length(71) == 100  # 71+10=81 -> round up to 100 (next after 80)
    assert calculate_dc_fixed_length(340) == 350 # 340+10=350 -> 350
    assert calculate_dc_fixed_length(341) == 350 # 341+10=351 -> exceeds, return 350

    # Part lookups
    assert get_feeder_part_by_size("48F") == "CAB00282"
    assert get_tangent_part_by_size("144F") == "PAC00051"
    assert get_deadend_part_by_size("288F") == "PAC00063"
    assert get_drop_cable_part(100) == "CAB00311"

    # Splice counting
    sample_splicing = [
        {"ag_name": "AG01", "cable_type": "FC", "from_node": "FJ01", "to_node": "AG01"},
        {"ag_name": "AG01", "cable_type": "FC", "from_node": "FJ02", "to_node": "AG01"},
        {"ag_name": "AG02", "cable_type": "CORE", "from_node": "AGG", "to_node": "AG02"},
    ]
    splice_result = count_splices_from_plan(sample_splicing, "48F")
    assert splice_result["fc_splices"] == 96   # 2 FC segments x 48 fibres
    assert splice_result["core_splices"] == 48 # 1 CORE segment x 48 fibres
    assert splice_result["total_splices"] == 144

    # BOM calculation smoke test
    test_data = {
        "djs": {
            "DJ01": {"feat_id": 1, "ratio": "1:8", "block_name": "B1", "ag_name": "AG01", "dc_name": "DC01"},
            "DJ02": {"feat_id": 2, "ratio": "1:8", "block_name": "B1", "ag_name": "AG01", "dc_name": "DC02"},
        },
        "blocks": {"B1": {"block_id": 1, "ag_name": "AG01", "num_djs": 2}},
        "ags": {"AG01": {"num_fjs": 1, "num_blocks": 1}},
        "fjs": {"FJ01": {"ag_name": "AG01"}},
        "poles": [
            {"feat_id": 1, "pole_type": "aerial"},
            {"feat_id": 2, "pole_type": "aerial"},
            {"feat_id": 3, "pole_type": "dead_end"},
        ],
        "fc_info": [
            {"fc_name": "FC01", "length_m": 500, "from_ag": "AG01", "to_ag": "OLT"},
        ],
        "dc_segments": [
            {"dc_name": "DC01", "from_node": "FJ01", "to_node": "DJ01", "length_m": 45, "block_id": 1, "ag_name": "AG01"},
            {"dc_name": "DC02", "from_node": "FJ01", "to_node": "DJ02", "length_m": 60, "block_id": 1, "ag_name": "AG01"},
        ],
        "houses": {"DJ01": ["H1", "H2", "H3", "H4", "H5", "H6", "H7", "H8"]},
        "splicing_plan": sample_splicing,
    }
    test_config = {
        "cable_size": "48F",
        "fibers_per_fj": 4,
        "core_length_per_ag_km": 2.0,
        "route_type": "aerial",
    }

    bom = calculate_bom(test_data, test_config)
    assert "pre_conventional" in bom
    assert "conventional" in bom
    assert "pole_hardware" in bom
    assert "consumables" in bom
    assert "other" in bom
    assert "summary" in bom
    assert bom["summary"]["total_djs"] == 2
    assert bom["summary"]["total_fjs"] == 1

    print("All self-tests passed!")


if __name__ == "__main__":
    _self_test()
