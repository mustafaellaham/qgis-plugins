# -*- coding: utf-8 -*-
"""
FTTH Auto Splicing Plan v8.6.22 -- QGIS Processing Plugin
Copyright (c) Mustafa M M Ellaham. All rights reserved.

Generates FTTH splicing plan Excel (.xlsx) from labeled network data.
Uses DJ Polygons as the PRIMARY method to match houses to DJs
(point-in-polygon). Produces ONE .xlsx file with a formatted splicing
diagram worksheet.

By Mustafa M M Ellaham
"""

import os
import re

from qgis.PyQt.QtWidgets import QAction, QMenu, QMessageBox
from qgis.PyQt.QtGui import QIcon
from qgis.core import (
    QgsApplication, QgsProcessing,
    QgsProject, QgsVectorLayer, QgsFeature, QgsGeometry, QgsPointXY,
    QgsField, QgsFields, QgsWkbTypes,
    QgsProcessingAlgorithm, QgsProcessingParameterVectorLayer,
    QgsProcessingParameterString, QgsProcessingParameterBoolean,
    QgsProcessingParameterNumber, QgsProcessingParameterFolderDestination,
    QgsProcessingParameterFileDestination, QgsProcessingParameterFile,
    QgsProcessingParameterEnum,
    QgsProcessingOutputFolder,
    QgsProcessingProvider,
    QgsVectorFileWriter, QgsCoordinateTransformContext,
    QgsCoordinateTransform, QgsCoordinateReferenceSystem
)
from qgis import processing

# openpyxl for Excel output (bundled with QGIS)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# =============================================================================
# Utility functions
# =============================================================================

def find_field(layer, field_names):
    """Find field index in a QGIS vector layer, case-insensitive.
    Returns -1 if not found.
    """
    if not layer or not layer.isValid():
        return -1
    all_names = [f.name() for f in layer.fields()]
    for fn in field_names:
        fn_lower = fn.lower()
        for actual_name in all_names:
            if actual_name.lower() == fn_lower:
                return layer.fields().indexFromName(actual_name)
    return -1


def _safe_val(v, default=''):
    """Safely convert a value to a cleaned string.
    Returns default for None, NaN, 'none', 'null', etc.
    """
    if v is None:
        return default
    v = str(v).strip()
    return default if v.lower() in ("", "nan", "none", "null") else v


# =============================================================================
# FTTH SPLICING PLAN ALGORITHM  v8.6.22
# =============================================================================

class FTTHSplicingPlanAlgorithm(QgsProcessingAlgorithm):
    """Generate splicing plan Excel (.xlsx) -- uses DJ Polygons as the PRIMARY
    method to match houses to DJs (point-in-polygon). Produces ONE .xlsx
    file with a formatted splicing diagram worksheet.

    INPUT (all labeled by FTTH Labeler first):
      - DJ Polygons    : one polygon per DJ, carries block name attribute
      - DJ Points      : labeled with 'name', 'splitter', 'position', 'dc_name'
      - House Points   : premises / HP data with name field
      - Pole Points    : labeled with 'name' column
      - Drop Cable Lines (optional) : short lines from each house to its DJ
      - DC Lines         (optional) : for upstream FJ tracing
      - FJ Points        (optional) : for FJ pole lookup

    OUTPUT (one XLSX file):
      A formatted Excel splicing diagram with merged headers, styling,
      and per-DJ cable routing information including premise data.
    """

    IN_DJ_POLY = 'IN_DJ_POLY'
    IN_DJ      = 'IN_DJ'
    IN_HOUSES  = 'IN_HOUSES'
    IN_POLES   = 'IN_POLES'
    IN_DROPS   = 'IN_DROPS'
    IN_DC      = 'IN_DC'
    IN_FJ      = 'IN_FJ'

    PARAM_SNAP   = 'PARAM_SNAP'
    PARAM_OUTPUT = 'PARAM_OUTPUT'

    # ------------------------------------------------------------------
    # QGIS boilerplate
    # ------------------------------------------------------------------
    def name(self):
        return 'ftth_splicing_plan'

    def displayName(self):
        return 'FTTH Splicing Plan (Excel output)'

    def group(self):
        return 'FTTH Tools'

    def groupId(self):
        return 'ftth_tools'

    def createInstance(self):
        return FTTHSplicingPlanAlgorithm()

    def shortHelpString(self):
        return """
<h3>FTTH Splicing Plan (Excel output) v8.6.22</h3>
<p>Uses <b>DJ Polygons</b> as the PRIMARY method to match houses to DJs.
Each DJ Polygon defines the exact service area of one DJ. Houses inside
a polygon belong to that DJ -- exact, no ambiguity. ALL DJs appear in
the output (including those with 0 houses), ensuring 8 rows per DJ in
splicing diagrams.</p>
<p>Directly outputs a formatted <b>.xlsx</b> Excel file -- no separate
script needed. openpyxl is bundled with QGIS.</p>

<h4>Required Inputs:</h4>
<ul>
<li><b>DJ Polygons</b> -- one polygon per DJ, must have block name attribute (auto-detected: block, Block, BLOCK, b_name, id, ID)</li>
<li><b>DJ Points</b> -- labeled DJ (run Labeler first; needs name, splitter, position, dc_name)</li>
<li><b>House Points</b> -- premises/HP (name field auto-detected)</li>
<li><b>Pole Points</b> -- labeled poles (run Labeler first)</li>
</ul>

<h4>Optional Inputs:</h4>
<ul>
<li><b>Drop Cable Lines</b> -- short lines from each house to its DJ. Used as FALLBACK for houses not covered by DJ Polygons.</li>
<li><b>DC Lines</b> -- labeled DC lines with from_node/to_node (traces upstream to find serving FJ)</li>
<li><b>FJ Points</b> -- labeled FJ points (looked up to find FJ pole name)</li>
</ul>

<h4>Output (one XLSX file):</h4>
<p>A formatted Excel splicing diagram with merged headers, styling,
per-DJ cable routing, and premise data.</p>
<p><b>fj_name</b> = serving FJ name (traced upstream via DC lines, requires optional DC + FJ layers)<br/>
<b>fj_pole</b> = pole where the FJ is mounted (requires optional DC + FJ layers)<br/>
<b>ag_name</b> = AG name extracted from dc_name (e.g. "GMGZ2_AG01", always populated)<br/>
<b>block</b> = block name from DJ Polygons (populated when DJ Polygons layer is provided)</p>
"""

    def initAlgorithm(self, config=None):
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_DJ_POLY, 'DJ Polygons -- one polygon per DJ (REQUIRED, must have block name attr)',
            [QgsProcessing.TypeVectorPolygon]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_DJ, 'DJ Points -- must be labeled (name, splitter, position, dc_name)',
            [QgsProcessing.TypeVectorPoint]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_HOUSES, 'House / Premise Points (HP)',
            [QgsProcessing.TypeVectorPoint]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_POLES, 'Pole Points -- must be labeled (name column)',
            [QgsProcessing.TypeVectorPoint]
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_DROPS, 'Drop Cable Lines (optional -- fallback for houses not in DJ Polygons)',
            [QgsProcessing.TypeVectorLine], optional=True
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_DC, 'DC Lines (optional -- for FJ tracing)',
            [QgsProcessing.TypeVectorLine], optional=True
        ))
        self.addParameter(QgsProcessingParameterVectorLayer(
            self.IN_FJ, 'FJ Points (optional -- for FJ pole lookup)',
            [QgsProcessing.TypeVectorPoint], optional=True
        ))
        self.addParameter(QgsProcessingParameterNumber(
            self.PARAM_SNAP, 'Snap tolerance in meters',
            type=QgsProcessingParameterNumber.Double,
            defaultValue=5.0, minValue=0.1, maxValue=100.0
        ))
        self.addParameter(QgsProcessingParameterFileDestination(
            self.PARAM_OUTPUT, 'Output splicing diagram (XLSX)',
            fileFilter='Excel files (*.xlsx)'
        ))


    # ------------------------------------------------------------------
    # Excel generation -- integrated from generate_splicing_excel.py
    # ------------------------------------------------------------------
    @staticmethod
    def _build_splicing_excel(output_rows, output_path, feedback):
        """Build the splicing diagram Excel directly from output_rows.

        Column layout:
          A-C:   Fixed (Pole, FEEDER, AG)  -- NOT repeated per DC
          D-M:   DC1 segment (10 cols)
          N-W:   DC2 segment (10 cols)
          X-AH:  DC3 segment (10 cols) ...
          last 3: slack_length, premise_data_name, Block
        """
        import re

        # --- Constants ---
        FIXED_COLS = 3       # Pole, FEEDER, AG
        COLS_PER_DC = 10     # DC_label, DC_number, Size, No., Splitter, Leg, Pole_num, DJ_num, DC_len, DC_slack
        TRAILING_COLS = 3    # slack_length, premise_data_name, Block
        COORD_COLS = 2       # LAT, LONG (EPSG:4326 decimal degrees)

        # No Plus Code -- LAT/LONG in decimal degrees (EPSG:4326) only

        HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        SUBHEADER_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        TITLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        BOLD_FONT = Font(bold=True, size=10)
        HEADER_FONT = Font(bold=True, size=11)
        TITLE_FONT = Font(bold=True, size=12)
        CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
        THIN_BORDER = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # --- Helper functions ---
        def _safe_val(v):
            if v is None:
                return ""
            v = str(v).strip()
            return "" if v.lower() in ("", "nan", "none", "null") else v

        def get_fc_from_fj(fj_name):
            if not fj_name:
                return "FC01"
            m = re.search(r"FJ(\d+)", str(fj_name))
            return f"FC{int(m.group(1)):02d}" if m else "FC01"

        def extract_sp_from_dj(dj_name):
            """GMGZ2_DJ205_1:8 -> SP_205"""
            m = re.search(r'DJ(\d+)', str(dj_name))
            return f"SP_{m.group(1)}" if m else ""

        def extract_dc_number(dc_name):
            """VTN_HHS_GMGZ2_AG13_DC053.4_... -> DC053.4"""
            m = re.search(r'DC(\d+\.\d+)', str(dc_name))
            return f"DC{m.group(1)}" if m else ""

        def extract_dc_info(dc_name):
            if not dc_name:
                return {"size": "", "length": "", "length_slack": ""}
            name = str(dc_name)
            size_m = re.search(r"_(\d+F)_", name)
            size = size_m.group(1) if size_m else ""
            len_m = re.search(r"\((\d+)m\s*-\s*(\d+)m\)", name)
            if len_m:
                return {"size": size, "length": int(len_m.group(1)), "length_slack": int(len_m.group(2))}
            len_m2 = re.search(r"\((\d+)m\)", name)
            if len_m2:
                d = int(len_m2.group(1))
                return {"size": size, "length": d, "length_slack": d + 10}
            return {"size": size, "length": "", "length_slack": ""}

        # --- Group rows by (ag_name -> block -> position) ---
        ag_groups = {}
        for row in output_rows:
            ag = _safe_val(row.get('ag_name', 'Unknown'))
            block = _safe_val(row.get('block', 'B001'))
            if not block:
                block = "B001"
            pos_str = str(row.get('position', '1'))
            try:
                pos = int(float(pos_str))  # Handles both Integer (1) and Real (1.0) fields
            except (ValueError, TypeError):
                pos = 1

            if ag not in ag_groups:
                ag_groups[ag] = {}
            if block not in ag_groups[ag]:
                ag_groups[ag][block] = {}

            if pos not in ag_groups[ag][block]:
                ag_groups[ag][block][pos] = {
                    'position': pos,
                    'original_dj_name': _safe_val(row.get('name_3', '')),
                    'original_dc_label': _safe_val(row.get('dc_name', '')),
                    'pole_number': _safe_val(row.get('name_2_2', '')),
                    'fj_name': _safe_val(row.get('fj_name', '')),
                    'fj_pole': _safe_val(row.get('fj_pole', '')),
                    'ag_name': ag,
                    'splitter': _safe_val(row.get('splitter', '1:8')),
                    'houses': [],
                }

            house = _safe_val(row.get('Name_2', ''))
            if house:
                ag_groups[ag][block][pos]['houses'].append(house)
                # Also store fields mapped by house name
                hf = row.get('house_fields', {})
                if hf:
                    if 'house_fields_map' not in ag_groups[ag][block][pos]:
                        ag_groups[ag][block][pos]['house_fields_map'] = {}
                    ag_groups[ag][block][pos]['house_fields_map'][house] = hf

        # Sort houses and convert to ordered DJ lists
        # v8.6.22: Simple list conversion (no gap-splitting)
        for ag in ag_groups:
            for block in ag_groups[ag]:
                for pos in ag_groups[ag][block]:
                    ag_groups[ag][block][pos]['houses'] = sorted(
                        set(ag_groups[ag][block][pos]['houses'])
                    )
                positions = sorted(ag_groups[ag][block].keys())
                ag_groups[ag][block] = [
                    ag_groups[ag][block][p] for p in positions
                ]

        # --- Build Excel ---
        wb = Workbook()
        ws = wb.active
        ws.title = "Splicing Diagram"

        max_djs_per_block = 0
        block_dj_counts = []
        for ag_name in sorted(ag_groups.keys()):
            for block_name in sorted(ag_groups[ag_name].keys()):
                num_djs = len(ag_groups[ag_name][block_name])
                block_dj_counts.append((ag_name, block_name, num_djs))
                if num_djs > max_djs_per_block:
                    max_djs_per_block = num_djs

        if max_djs_per_block == 0:
            feedback.pushInfo("  No data to write.")
            return

        # Collect all unique house field names (from first house in each group)
        house_field_names = []
        for ag_name in sorted(ag_groups.keys()):
            for block_name in sorted(ag_groups[ag_name].keys()):
                djs = ag_groups[ag_name][block_name]
                for dj in djs:
                    for h_name in dj.get('houses', []):
                        pass
        # Get field names from first output row that has house_fields
        house_field_names = []
        for row in output_rows:
            hf = row.get('house_fields', {})
            if hf:
                house_field_names = list(hf.keys())
                break
        num_house_fields = len(house_field_names)

        total_cols = FIXED_COLS + (max_djs_per_block * COLS_PER_DC) + TRAILING_COLS + num_house_fields + COORD_COLS
        trailing_base = FIXED_COLS + (max_djs_per_block * COLS_PER_DC) + 1
        house_fields_base = trailing_base + TRAILING_COLS  # start after trailing columns
        coord_base = house_fields_base + num_house_fields  # start after house fields

        # Row 1: Title
        # Title 1: DC columns (1 to trailing_base - 1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=trailing_base - 1)
        c = ws.cell(row=1, column=1, value="Distribution Layer 1 (A - B)")
        c.font = TITLE_FONT; c.alignment = CENTER_ALIGNMENT; c.fill = HEADER_FILL

        # Title 2: Trailing columns (slack, premise name, Block)
        ws.merge_cells(start_row=1, start_column=trailing_base, end_row=1,
                       end_column=trailing_base + TRAILING_COLS - 1)
        c = ws.cell(row=1, column=trailing_base,
                    value="PREMIS DATA (MUST COMPLETE ALL COLUMS & DUPLICATE INFO REVIEWED)")
        c.font = TITLE_FONT; c.alignment = CENTER_ALIGNMENT; c.fill = TITLE_FILL

        # Title 3: House field columns (if any)
        if num_house_fields > 0:
            ws.merge_cells(start_row=1, start_column=house_fields_base, end_row=1,
                           end_column=coord_base - 1)
            c = ws.cell(row=1, column=house_fields_base, value="PREMISE DATA FIELDS")
            c.font = TITLE_FONT; c.alignment = CENTER_ALIGNMENT; c.fill = TITLE_FILL

        # Title 4: Coordinate columns (LAT, LONG)
        ws.merge_cells(start_row=1, start_column=coord_base, end_row=1, end_column=total_cols)
        c = ws.cell(row=1, column=coord_base, value="COORDINATES (EPSG:4326)")
        c.font = TITLE_FONT; c.alignment = CENTER_ALIGNMENT; c.fill = TITLE_FILL

        for col in range(1, total_cols + 1):
            cell = ws.cell(row=1, column=col)
            if col < trailing_base:
                cell.fill = HEADER_FILL
            elif col < house_fields_base:
                cell.fill = TITLE_FILL
            elif col < coord_base:
                cell.fill = TITLE_FILL
            else:
                cell.fill = TITLE_FILL
            cell.border = THIN_BORDER

        # Row 2: Category headers (merged groups per DC segment)
        for dc_idx in range(max_djs_per_block):
            base = FIXED_COLS + (dc_idx * COLS_PER_DC) + 1   # 1-based

            # Cable Route ID merges all 10 columns of this DC segment
            ws.merge_cells(
                start_row=2, start_column=base,
                end_row=2, end_column=base + COLS_PER_DC - 1
            )
            c = ws.cell(row=2, column=base,
                        value=f"Cable Route ID (AERIAL)_{dc_idx + 1}")
            c.font = BOLD_FONT; c.alignment = CENTER_ALIGNMENT
            c.fill = SUBHEADER_FILL; c.border = THIN_BORDER
            for o in range(COLS_PER_DC):
                ws.cell(row=2, column=base + o).border = THIN_BORDER
                ws.cell(row=2, column=base + o).fill = SUBHEADER_FILL

        # Trailing category headers
        ws.merge_cells(start_row=2, start_column=trailing_base, end_row=2, end_column=trailing_base + 1)
        c = ws.cell(row=2, column=trailing_base, value="+ Slack")
        c.font = BOLD_FONT; c.alignment = CENTER_ALIGNMENT; c.fill = SUBHEADER_FILL; c.border = THIN_BORDER
        ws.cell(row=2, column=trailing_base + 1).border = THIN_BORDER
        ws.cell(row=2, column=trailing_base + 1).fill = SUBHEADER_FILL

        c = ws.cell(row=2, column=trailing_base + 2, value="Block")
        c.font = BOLD_FONT; c.alignment = CENTER_ALIGNMENT; c.fill = SUBHEADER_FILL; c.border = THIN_BORDER

        # Row 3: Detailed headers
        # Fixed headers (cols 1-3)
        fixed_headers = ["Pole", "FEEDER", "AG"]
        for hi, hdr in enumerate(fixed_headers):
            c = ws.cell(row=3, column=1 + hi, value=hdr)
            c.font = BOLD_FONT; c.alignment = CENTER_ALIGNMENT
            c.fill = HEADER_FILL; c.border = THIN_BORDER

        # Per-DC detail headers (repeated for each DC segment)
        dc_detail_headers = [
            "DC label", "DC number", "Size", "No.",
            "splitter number", "Leg", "Pole_number",
            "DJ_number", "DC_length", "DC+slack",
        ]
        for dc_idx in range(max_djs_per_block):
            base = FIXED_COLS + (dc_idx * COLS_PER_DC) + 1
            for hi, hdr in enumerate(dc_detail_headers):
                c = ws.cell(row=3, column=base + hi, value=hdr)
                c.font = BOLD_FONT; c.alignment = CENTER_ALIGNMENT
                c.fill = HEADER_FILL; c.border = THIN_BORDER

        # Trailing detail headers
        for off, hdr in enumerate(["slack_length", "premise_data_name", "Block"]):
            c = ws.cell(row=3, column=trailing_base + off, value=hdr)
            c.font = BOLD_FONT; c.alignment = CENTER_ALIGNMENT
            c.fill = HEADER_FILL; c.border = THIN_BORDER

        # House field headers (dynamic, from premise data layer)
        for fi, fn in enumerate(house_field_names):
            c = ws.cell(row=3, column=house_fields_base + fi, value=fn)
            c.font = BOLD_FONT; c.alignment = CENTER_ALIGNMENT
            c.fill = HEADER_FILL; c.border = THIN_BORDER

        # Coordinate headers (LAT, LONG in EPSG:4326)
        coord_headers = ["LAT", "LONG"]
        for ci, ch in enumerate(coord_headers):
            c = ws.cell(row=3, column=coord_base + ci, value=ch)
            c.font = BOLD_FONT; c.alignment = CENTER_ALIGNMENT
            c.fill = HEADER_FILL; c.border = THIN_BORDER

        # Build house coordinate lookup from output_rows
        house_coord_lookup = {}
        for row in output_rows:
            h_name = row.get('Name_2', '')
            if h_name:
                house_coord_lookup[h_name] = (row.get('house_lat', ''), row.get('house_lon', ''))

        # --- Data Rows ---
        current_row = 4
        for ag_name, block_name, num_djs in block_dj_counts:
            djs = ag_groups[ag_name][block_name]

            # Precompute per-DJ info using ACTUAL labels from CSV
            dj_info_list = []
            for dj_entry in djs:
                dc_label = _safe_val(dj_entry.get('original_dc_label', ''))
                dc_info = extract_dc_info(dc_label)
                dj_name = _safe_val(dj_entry.get('original_dj_name', ''))

                info = {
                    'dc_label': dc_label,
                    'dc_number': extract_dc_number(dc_label),
                    'dc_size': dc_info['size'],
                    'dc_length': dc_info['length'],
                    'dc_slack': dc_info['length_slack'],
                    'splitter_number': extract_sp_from_dj(dj_name),
                    'pole': _safe_val(dj_entry.get('pole_number', '')),
                    'dj_number': dj_name,   # ACTUAL DJ name from CSV
                }
                dj_info_list.append(info)

            # Fixed column values come from the FIRST DJ in this block
            first_pole = _safe_val(djs[0].get('fj_pole', '')) if djs else ''
            first_fc = get_fc_from_fj(djs[0]['fj_name']) if djs else 'FC01'
            first_ag = _safe_val(djs[0]['ag_name']) if djs else ''

            for dj_idx in range(num_djs):
                dj_entry = djs[dj_idx]
                house_list = dj_entry['houses']

                # Alternating row color: even DJ = white, odd DJ = light blue
                dj_fill = None if (dj_idx % 2 == 0) else PatternFill(
                    start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")

                for leg in range(1, 9):
                    # 1. Write FIXED columns (A, B, C) -- from FIRST DJ in block
                    ws.cell(row=current_row, column=1, value=first_pole)   # Pole
                    ws.cell(row=current_row, column=2, value=first_fc)     # FEEDER
                    ws.cell(row=current_row, column=3, value=first_ag)     # AG

                    # 2. Write per-DC segment columns
                    for seg_idx in range(num_djs):
                        base = FIXED_COLS + (seg_idx * COLS_PER_DC) + 1
                        source_info = dj_info_list[seg_idx]

                        if seg_idx < dj_idx:
                            show_leg = "9"          # link fiber
                        elif seg_idx == dj_idx:
                            show_leg = str(leg)      # actual house leg 1-8
                        else:
                            continue                  # skip (downstream of current DJ)

                        ws.cell(row=current_row, column=base + 0, value=source_info['dc_label'])
                        ws.cell(row=current_row, column=base + 1, value=source_info['dc_number'])
                        ws.cell(row=current_row, column=base + 2, value=source_info['dc_size'])
                        ws.cell(row=current_row, column=base + 3, value="1")
                        ws.cell(row=current_row, column=base + 4, value=source_info['splitter_number'])

                        # Leg column: yellow fill for "9", normal otherwise
                        leg_cell = ws.cell(row=current_row, column=base + 5, value=show_leg)
                        if show_leg == "9":
                            leg_cell.fill = PatternFill(
                                start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                        ws.cell(row=current_row, column=base + 6, value=source_info['pole'])
                        ws.cell(row=current_row, column=base + 7, value=source_info['dj_number'])
                        _dl = source_info['dc_length']
                        _ds = source_info['dc_slack']
                        ws.cell(row=current_row, column=base + 8,
                                value=f"{int(_dl):04d}" if _dl != '' else '')
                        ws.cell(row=current_row, column=base + 9,
                                value=f"{int(_ds):04d}" if _ds != '' else '')

                    # 3. Write TRAILING columns
                    if leg <= len(house_list):
                        house_name = house_list[leg - 1]
                    else:
                        house_name = ""

                    ws.cell(row=current_row, column=trailing_base + 0, value="")
                    ws.cell(row=current_row, column=trailing_base + 1, value=house_name)
                    ws.cell(row=current_row, column=trailing_base + 2, value=block_name)

                    # 4. Write HOUSE FIELD columns (all fields from premise data layer)
                    if house_name and num_house_fields > 0:
                        # Get fields from the current DJ's house_fields_map
                        fields_map = dj_entry.get('house_fields_map', {})
                        house_data = fields_map.get(house_name, {})
                        for fi, fn in enumerate(house_field_names):
                            ws.cell(row=current_row, column=house_fields_base + fi,
                                    value=house_data.get(fn, ''))

                    # 5. Write COORDINATE columns (LAT, LONG in EPSG:4326)
                    h_lat, h_lon = house_coord_lookup.get(house_name, ('', ''))
                    ws.cell(row=current_row, column=coord_base + 0, value=h_lat)
                    ws.cell(row=current_row, column=coord_base + 1, value=h_lon)

                    # Track Leg=9 yellow cells so we don't overwrite them
                    yellow_cols = set()
                    for seg_idx in range(num_djs):
                        base = FIXED_COLS + (seg_idx * COLS_PER_DC) + 1
                        source_info = dj_info_list[seg_idx]
                        if seg_idx < dj_idx:
                            yellow_cols.add(base + 5)  # Leg column for upstream DJ

                    # Borders + alignment + alternating row color for ALL columns
                    for col in range(1, total_cols + 1):
                        cell = ws.cell(row=current_row, column=col)
                        cell.border = THIN_BORDER
                        cell.alignment = CENTER_ALIGNMENT
                        # Apply alternating row color, but PRESERVE yellow Leg=9 fills
                        if dj_fill and col < trailing_base and col not in yellow_cols:
                            cell.fill = dj_fill

                    current_row += 1

                # If this is the LAST DJ in the block, highlight the last row in Block column (yellow)
                if dj_idx == num_djs - 1 and current_row > 4:
                    last_row = current_row - 1
                    block_cell = ws.cell(row=last_row, column=trailing_base + 2)
                    block_cell.fill = PatternFill(
                        start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # --- Column widths ---
        # Fixed columns
        ws.column_dimensions[get_column_letter(1)].width = 16   # Pole
        ws.column_dimensions[get_column_letter(2)].width = 10   # FEEDER
        ws.column_dimensions[get_column_letter(3)].width = 16   # AG

        # Per-DC columns
        for dc_idx in range(max_djs_per_block):
            base = FIXED_COLS + (dc_idx * COLS_PER_DC) + 1
            ws.column_dimensions[get_column_letter(base + 0)].width = 55   # DC label
            ws.column_dimensions[get_column_letter(base + 1)].width = 12   # DC number
            ws.column_dimensions[get_column_letter(base + 2)].width = 8    # Size
            ws.column_dimensions[get_column_letter(base + 3)].width = 6    # No.
            ws.column_dimensions[get_column_letter(base + 4)].width = 16   # splitter number
            ws.column_dimensions[get_column_letter(base + 5)].width = 8    # Leg
            ws.column_dimensions[get_column_letter(base + 6)].width = 16   # Pole_number
            ws.column_dimensions[get_column_letter(base + 7)].width = 25   # DJ_number
            ws.column_dimensions[get_column_letter(base + 8)].width = 12   # DC_length
            ws.column_dimensions[get_column_letter(base + 9)].width = 12   # DC+slack

        # Trailing columns
        ws.column_dimensions[get_column_letter(trailing_base + 0)].width = 14
        ws.column_dimensions[get_column_letter(trailing_base + 1)].width = 30
        ws.column_dimensions[get_column_letter(trailing_base + 2)].width = 10

        # House field columns (dynamic widths)
        for fi, fn in enumerate(house_field_names):
            ws.column_dimensions[get_column_letter(house_fields_base + fi)].width = max(12, len(fn) + 2)

        # Coordinate columns (LAT, LONG)
        ws.column_dimensions[get_column_letter(coord_base + 0)].width = 14   # LAT
        ws.column_dimensions[get_column_letter(coord_base + 1)].width = 14   # LONG

        ws.freeze_panes = "A4"
        wb.save(output_path)
        feedback.pushInfo("  Data rows: {}".format(current_row - 4))


    # ------------------------------------------------------------------
    # Main processing
    # ------------------------------------------------------------------
    def processAlgorithm(self, parameters, context, feedback):
        snap_tol = self.parameterAsDouble(parameters, self.PARAM_SNAP, context)
        output_path = self.parameterAsString(parameters, self.PARAM_OUTPUT, context)

        dj_poly_layer = self.parameterAsVectorLayer(parameters, self.IN_DJ_POLY, context)
        dj_layer      = self.parameterAsVectorLayer(parameters, self.IN_DJ,      context)
        house_layer   = self.parameterAsVectorLayer(parameters, self.IN_HOUSES,  context)
        pole_layer    = self.parameterAsVectorLayer(parameters, self.IN_POLES,   context)
        drops_layer   = self.parameterAsVectorLayer(parameters, self.IN_DROPS,   context)
        dc_layer      = self.parameterAsVectorLayer(parameters, self.IN_DC,      context)
        fj_layer      = self.parameterAsVectorLayer(parameters, self.IN_FJ,      context)

        feedback.pushInfo("=" * 60)
        feedback.pushInfo("FTTH Splicing Plan v8.6.22 -- Excel output")
        feedback.pushInfo("=" * 60)
        feedback.pushInfo("Snap tolerance: {}m".format(snap_tol))
        feedback.pushInfo("")

        # ================================================================
        # Helper: case-insensitive field lookup
        # ================================================================
        def find_field(layer, names):
            """Find field index, case-insensitive."""
            all_names = [f.name() for f in layer.fields()]
            for fn in names:
                fn_lower = fn.lower()
                for actual_name in all_names:
                    if actual_name.lower() == fn_lower:
                        return layer.fields().indexFromName(actual_name)
            return -1

        # ================================================================
        # Helper: get start and end points of a line
        # ================================================================
        def get_line_endpoints(line_geom):
            """Get start and end QgsPointXY of a LineString/MultiLineString."""
            if line_geom.isMultipart():
                parts = line_geom.asMultiPolyline()
                if parts:
                    first_part = parts[0]
                    last_part = parts[-1]
                    if first_part and last_part:
                        return (QgsPointXY(first_part[0]), QgsPointXY(last_part[-1]))
            else:
                pts = line_geom.asPolyline()
                if pts and len(pts) >= 2:
                    return (QgsPointXY(pts[0]), QgsPointXY(pts[-1]))
            return (None, None)

        # ================================================================
        # Helper: find nearest node to a point
        # ================================================================
        def nearest_node(point, nodes, tolerance):
            """Find nearest node to a point within tolerance.
            nodes: dict of {fid: {'point': QgsPointXY, ...}}
            Returns: (fid, info) or (None, None)
            """
            best_fid = None
            best_info = None
            best_dist = float('inf')
            for fid, info in nodes.items():
                d = point.distance(info['point'])
                if d < best_dist and d <= tolerance:
                    best_dist = d
                    best_fid = fid
                    best_info = info
            return best_fid, best_info

        # ================================================================
        # Helper: find nearest pole to a point
        # ================================================================
        def nearest_pole(point, poles_list, tolerance):
            """Find nearest pole to a point within tolerance.
            poles_list: list of {'fid': int, 'point': QgsPointXY, 'name': str}
            Returns: (fid, name) or (None, '')
            """
            best = None
            best_dist = float('inf')
            for p in poles_list:
                d = point.distance(p['point'])
                if d < best_dist and d <= tolerance:
                    best_dist = d
                    best = p
            if best:
                return best['fid'], best['name']
            return None, ''

        # ================================================================
        # Helper: extract AG name from DC label
        # e.g. "VTN_HHS_GMGZ2_AG01_DC001.1_1F_ADSS_G.657.A1(152m-162m)" -> "GMGZ2_AG01"
        # ================================================================
        def extract_ag(dc_name):
            """Extract AG name from DC label. Returns '' if not found."""
            if not dc_name:
                return ''
            m = re.search(r'([A-Za-z0-9]+_AG\d+)', str(dc_name))
            return m.group(1) if m else ''

        # ================================================================
        # Helper: walk upstream from DJ to find FJ
        # ================================================================
        def find_fj_for_dj(dj_name, dc_by_to, fjs):
            """Trace upstream from DJ through DC lines to find FJ.
            Returns (fj_name, fj_point) or ('', None)
            """
            if not dj_name or not dc_by_to:
                return '', None
            visited = set()
            current = dj_name
            while current and current not in visited:
                visited.add(current)
                dc_info = dc_by_to.get(current)
                if not dc_info:
                    break
                upstream = dc_info['from']
                if upstream in fjs:
                    return upstream, fjs[upstream]['point']
                current = upstream
            return '', None

        # ================================================================
        # Step 1 -- Detect fields and validate
        # ================================================================
        feedback.pushInfo("[Step 1] Detecting fields...")

        # --- DJ Polygon block field (auto-detect) ---
        poly_block_idx = find_field(dj_poly_layer, [
            'block', 'Block', 'BLOCK',
            'b_name', 'B_NAME',
            'id', 'ID'
        ])

        # --- DJ fields (must be labeled by Labeler plugin) ---
        dj_name_idx = find_field(dj_layer, ['name', 'Name', 'NAME'])
        dj_splitter_idx = find_field(dj_layer, ['splitter', 'Splitter', 'SPLITTER'])
        dj_position_idx = find_field(dj_layer, ['position', 'Position', 'POSITION'])
        dj_dcname_idx = find_field(dj_layer, ['dc_name', 'DC_name', 'dcname', 'DCNAME'])

        # --- House name field (extensive auto-detection) ---
        house_name_idx = find_field(house_layer, [
            'name', 'Name', 'NAME',
            'premise', 'Premise', 'PREMISE',
            'house', 'House', 'HOUSE',
            'hp', 'HP',
            'street_no', 'Street_No', 'STREET_NO', 'streetno',
            'fid', 'FID'
        ])

        # --- Pole name field ---
        pole_name_idx = find_field(pole_layer, ['name', 'Name', 'NAME'])

        # --- Drop Cable name field (optional) ---
        drop_name_idx = -1
        if drops_layer and drops_layer.isValid():
            drop_name_idx = find_field(drops_layer, ['name', 'Name', 'NAME'])

        # Validate required layers
        if not dj_poly_layer or not dj_poly_layer.isValid():
            feedback.reportError("DJ Polygons layer is REQUIRED. Please provide the DJ Polygons layer.")
            return {self.PARAM_OUTPUT: output_path}
        if dj_name_idx < 0:
            feedback.reportError("DJ layer has no 'name' column. Run FTTH Labeler first!")
            return {self.PARAM_OUTPUT: output_path}

        feedback.pushInfo("  DJ Polygon block field: {}".format(
            'found' if poly_block_idx >= 0 else 'NOT FOUND'))
        feedback.pushInfo("  DJ name field:          found")
        feedback.pushInfo("  DJ splitter field:      {}".format(
            'found' if dj_splitter_idx >= 0 else 'NOT FOUND'))
        feedback.pushInfo("  DJ position field:      {}".format(
            'found' if dj_position_idx >= 0 else 'NOT FOUND'))
        feedback.pushInfo("  DJ dc_name field:       {}".format(
            'found' if dj_dcname_idx >= 0 else 'NOT FOUND'))
        feedback.pushInfo("  House name field:       {}".format(
            'found' if house_name_idx >= 0 else 'NOT FOUND (using HP_{fid})'))
        feedback.pushInfo("  Pole name field:        {}".format(
            'found' if pole_name_idx >= 0 else 'NOT FOUND'))
        feedback.pushInfo("  Drop Cable name field:  {}".format(
            'found' if drop_name_idx >= 0 else 'NOT FOUND (optional)'))

        # ================================================================
        # Step 2 -- Index all layers into memory
        # ================================================================
        feedback.pushInfo("")
        feedback.pushInfo("[Step 2] Indexing layers into memory...")

        # --- Index DJ Polygons: list of {'fid', 'geom', 'block'} ---
        dj_polygons = []
        for feat in dj_poly_layer.getFeatures():
            geom = feat.geometry()
            blk = ''
            if poly_block_idx >= 0:
                blk = str(feat[poly_block_idx] or '').strip()
            dj_polygons.append({
                'fid': feat.id(),
                'geom': geom,
                'block': blk
            })
        feedback.pushInfo("  Indexed {} DJ polygons".format(len(dj_polygons)))

        if len(dj_polygons) == 0:
            feedback.reportError("DJ Polygons layer has 0 features. Nothing to process.")
            return {self.PARAM_OUTPUT: output_path}

        # --- Index DJ Points: {name: {'fid', 'point', 'splitter', 'position', 'dc_name'}} ---
        djs = {}
        djs_by_fid = {}
        skipped_null_djs = 0
        for feat in dj_layer.getFeatures():
            geom = feat.geometry()
            if geom.isNull() or geom.isEmpty():
                skipped_null_djs += 1
                continue
            pt = geom.asPoint()
            dj_name = str(feat[dj_name_idx] or '')
            splitter = str(feat[dj_splitter_idx] or '') if dj_splitter_idx >= 0 else ''
            position = feat[dj_position_idx] if dj_position_idx >= 0 else ''
            dc_name = str(feat[dj_dcname_idx] or '') if dj_dcname_idx >= 0 else ''
            info = {
                'fid': feat.id(),
                'point': pt,
                'name': dj_name,
                'splitter': splitter,
                'position': position,
                'dc_name': dc_name
            }
            djs[dj_name] = info
            djs_by_fid[feat.id()] = info
        feedback.pushInfo("  Indexed {} DJ points".format(len(djs)))
        if skipped_null_djs > 0:
            feedback.pushInfo("  Skipped {} DJ features with null/empty geometry".format(skipped_null_djs))

        # --- Index Houses: {fid: {'point', 'name', 'fields'}} ---
        # Store ALL fields from the house layer for Excel output
        house_field_names = [f.name() for f in house_layer.fields()]
        houses = {}
        skipped_null_houses = 0
        for feat in house_layer.getFeatures():
            geom = feat.geometry()
            if geom.isNull() or geom.isEmpty():
                skipped_null_houses += 1
                continue
            pt = geom.asPoint()
            if house_name_idx >= 0:
                nm = str(feat[house_name_idx] or '').strip()
            else:
                nm = "HP_{}".format(feat.id())
            # Store all field values
            all_fields = {}
            for fn in house_field_names:
                all_fields[fn] = str(feat[fn] or '')
            houses[feat.id()] = {
                'point': pt,
                'name': nm,
                'fields': all_fields
            }
        feedback.pushInfo("  Indexed {} house points".format(len(houses)))
        if skipped_null_houses > 0:
            feedback.pushInfo("  Skipped {} house features with null/empty geometry".format(skipped_null_houses))

        # --- Index Poles: list of {'fid', 'point', 'name'} ---
        poles = []
        skipped_null_poles = 0
        for feat in pole_layer.getFeatures():
            geom = feat.geometry()
            if geom.isNull() or geom.isEmpty():
                skipped_null_poles += 1
                continue
            pt = geom.asPoint()
            nm = str(feat[pole_name_idx] or '') if pole_name_idx >= 0 else "Pole_{}".format(feat.id())
            poles.append({
                'fid': feat.id(),
                'point': pt,
                'name': nm
            })
        feedback.pushInfo("  Indexed {} poles".format(len(poles)))
        if skipped_null_poles > 0:
            feedback.pushInfo("  Skipped {} pole features with null/empty geometry".format(skipped_null_poles))

        # --- Index DC lines (for upstream tracing) ---
        dc_by_to = {}   # {to_node_name: {'from': from_node_name, 'dc_name': str}}
        if dc_layer and dc_layer.isValid():
            dc_from_idx = find_field(dc_layer, ['from_node', 'from', 'From_Node', 'FROM_NODE'])
            dc_to_idx = find_field(dc_layer, ['to_node', 'to', 'To_Node', 'TO_NODE'])
            dc_name_idx2 = find_field(dc_layer, ['name', 'Name', 'NAME'])
            for feat in dc_layer.getFeatures():
                fn = str(feat[dc_from_idx] or '') if dc_from_idx >= 0 else ''
                tn = str(feat[dc_to_idx] or '') if dc_to_idx >= 0 else ''
                dn = str(feat[dc_name_idx2] or '') if dc_name_idx2 >= 0 else ''
                if tn and fn:
                    dc_by_to[tn] = {'from': fn, 'dc_name': dn}
            feedback.pushInfo("  Indexed {} DC lines".format(len(dc_by_to)))

        # --- Index FJ points (for FJ pole lookup) ---
        fjs = {}   # {fj_name: {'point': QgsPointXY}}
        if fj_layer and fj_layer.isValid():
            fj_name_idx2 = find_field(fj_layer, ['name', 'Name', 'NAME'])
            for feat in fj_layer.getFeatures():
                nm = str(feat[fj_name_idx2] or '') if fj_name_idx2 >= 0 else ''
                if nm:
                    fjs[nm] = {'point': feat.geometry().asPoint()}
            feedback.pushInfo("  Indexed {} FJ points".format(len(fjs)))

        # ================================================================
        # Step 3 -- Match DJ Polygons to DJ Points
        # ================================================================
        feedback.pushInfo("")
        feedback.pushInfo("[Step 3] Matching DJ Polygons to DJ Points...")

        # {polygon_fid: dj_name}
        poly_to_dj = {}
        unmatched_polys = []

        for poly in dj_polygons:
            poly_geom = poly['geom']
            best_dj = None
            best_dist = float('inf')

            # Try 1: point-in-polygon (DJ point inside polygon)
            for dj_name, dj_info in djs.items():
                if poly_geom.contains(dj_info['point']):
                    best_dj = dj_name
                    break

            # Try 2: nearest DJ within 10m of polygon centroid
            if not best_dj:
                centroid = poly_geom.centroid().asPoint()
                for dj_name, dj_info in djs.items():
                    d = centroid.distance(dj_info['point'])
                    if d < best_dist and d <= 10.0:
                        best_dist = d
                        best_dj = dj_name

            if best_dj:
                poly_to_dj[poly['fid']] = best_dj
            else:
                unmatched_polys.append(poly['fid'])

        feedback.pushInfo("  Matched {} of {} DJ Polygons to DJ Points".format(
            len(poly_to_dj), len(dj_polygons)))
        if unmatched_polys:
            feedback.pushInfo("  WARNING: {} DJ Polygons could not be matched to any DJ Point".format(
                len(unmatched_polys)))

        # ================================================================
        # Step 4 -- Match Houses to DJ Polygons (PRIMARY method)
        # ================================================================
        feedback.pushInfo("")
        feedback.pushInfo("[Step 4] Matching Houses to DJ Polygons (point-in-polygon)...")

        # {dj_name: [house_info, ...]}
        dj_to_houses = {}
        unmatched_houses = {}   # houses not matched by any DJ Polygon

        for house_fid, house_info in houses.items():
            matched = False
            for poly in dj_polygons:
                dj_name = poly_to_dj.get(poly['fid'])
                if not dj_name:
                    continue
                if poly['geom'].contains(house_info['point']):
                    if dj_name not in dj_to_houses:
                        dj_to_houses[dj_name] = []
                    dj_to_houses[dj_name].append({
                        'fid': house_fid,
                        'name': house_info['name'],
                        'fields': house_info.get('fields', {}),
                        'point': house_info['point'],
                    })
                    matched = True
                    break
            if not matched:
                unmatched_houses[house_fid] = house_info

        total_matched = sum(len(hlist) for hlist in dj_to_houses.values())
        feedback.pushInfo("  Matched {} houses to DJs via DJ Polygons".format(total_matched))
        if unmatched_houses:
            feedback.pushInfo("  {} houses not inside any DJ Polygon".format(len(unmatched_houses)))

        # ================================================================
        # Step 5 -- Fallback: match unmatched houses via Drop Cables
        # ================================================================
        if unmatched_houses and drops_layer and drops_layer.isValid():
            feedback.pushInfo("")
            feedback.pushInfo("[Step 5] Fallback: using Drop Cables for {} unmatched houses...".format(
                len(unmatched_houses)))

            # Build temporary DJ index by fid for nearest-node lookup
            djs_for_nearest = {}
            for dj_name, dj_info in djs.items():
                djs_for_nearest[dj_info['fid']] = {
                    'point': dj_info['point'],
                    'name': dj_name
                }

            fallback_matched = 0
            skipped_drops = 0

            for drop_feat in drops_layer.getFeatures():
                if feedback.isCanceled():
                    return {self.PARAM_OUTPUT: output_path}

                try:
                    drop_fid = drop_feat.id()
                    drop_geom = drop_feat.geometry()
                    pt_a, pt_b = get_line_endpoints(drop_geom)
                    if pt_a is None or pt_b is None:
                        skipped_drops += 1
                        continue

                    # Strategy: one end is an unmatched house, other end is a DJ
                    # Try both combinations
                    # Option 1: A=unmatched_house, B=DJ
                    h_fid_1, h_info_1 = nearest_node(pt_a, unmatched_houses, snap_tol)
                    d_fid_1, d_info_1 = nearest_node(pt_b, djs_for_nearest, snap_tol)
                    opt1_ok = (h_fid_1 is not None and d_fid_1 is not None)

                    # Option 2: A=DJ, B=unmatched_house
                    d_fid_2, d_info_2 = nearest_node(pt_a, djs_for_nearest, snap_tol)
                    h_fid_2, h_info_2 = nearest_node(pt_b, unmatched_houses, snap_tol)
                    opt2_ok = (d_fid_2 is not None and h_fid_2 is not None)

                    house_fid = None
                    dj_name = None

                    if opt1_ok and opt2_ok:
                        dist1 = pt_a.distance(h_info_1['point']) + pt_b.distance(d_info_1['point'])
                        dist2 = pt_a.distance(d_info_2['point']) + pt_b.distance(h_info_2['point'])
                        if dist1 <= dist2:
                            house_fid, dj_name = h_fid_1, d_info_1['name']
                        else:
                            house_fid, dj_name = h_fid_2, d_info_2['name']
                    elif opt1_ok:
                        house_fid, dj_name = h_fid_1, d_info_1['name']
                    elif opt2_ok:
                        house_fid, dj_name = h_fid_2, d_info_2['name']
                    else:
                        skipped_drops += 1
                        continue

                    if house_fid and house_fid in unmatched_houses:
                        house_info = unmatched_houses[house_fid]
                        if dj_name not in dj_to_houses:
                            dj_to_houses[dj_name] = []
                        dj_to_houses[dj_name].append({
                            'fid': house_fid,
                            'name': house_info['name'],
                            'fields': house_info.get('fields', {}),
                            'point': house_info['point'],
                        })
                        del unmatched_houses[house_fid]
                        fallback_matched += 1

                except Exception as e:
                    skipped_drops += 1
                    continue

            feedback.pushInfo("  Fallback matched {} additional houses via Drop Cables".format(
                fallback_matched))
            if skipped_drops > 0:
                feedback.pushInfo("  Skipped {} drop cables".format(skipped_drops))

        if unmatched_houses:
            feedback.pushInfo("  WARNING: {} houses remain unmatched (no DJ Polygon, no Drop Cable)".format(
                len(unmatched_houses)))

        # ================================================================
        # Step 6 -- Find nearest Pole for each DJ
        # ================================================================
        feedback.pushInfo("")
        feedback.pushInfo("[Step 6] Finding nearest poles for each DJ...")

        # {dj_name: {'pole_fid', 'pole_name'}}
        dj_to_pole = {}
        for poly in dj_polygons:
            dj_name = poly_to_dj.get(poly['fid'])
            if not dj_name or dj_name not in djs:
                continue
            dj_info = djs[dj_name]
            pole_fid, pole_name = nearest_pole(dj_info['point'], poles, snap_tol)
            dj_to_pole[dj_name] = {
                'pole_fid': pole_fid or '',
                'pole_name': pole_name
            }

        feedback.pushInfo("  Found poles for {} DJs".format(len(dj_to_pole)))

        # ================================================================
        # Step 7 -- Trace upstream to FJ (if DC and FJ provided)
        # ================================================================
        feedback.pushInfo("")
        feedback.pushInfo("[Step 7] Tracing upstream to FJ...")

        # {dj_name: {'fj_name', 'fj_pole'}}
        dj_to_fj = {}
        if dc_by_to and fjs:
            for poly in dj_polygons:
                dj_name = poly_to_dj.get(poly['fid'])
                if not dj_name:
                    continue
                fj_name, fj_point = find_fj_for_dj(dj_name, dc_by_to, fjs)
                fj_pole_name = ''
                if fj_name and fj_point:
                    _, fj_pole_name = nearest_pole(fj_point, poles, snap_tol)
                dj_to_fj[dj_name] = {
                    'fj_name': fj_name,
                    'fj_pole': fj_pole_name
                }
            feedback.pushInfo("  Traced FJ for {} DJs".format(len(dj_to_fj)))
        else:
            feedback.pushInfo("  Skipped (DC Lines and/or FJ Points not provided)")

        # ================================================================
        # Step 8 -- Build output rows
        # ================================================================
        feedback.pushInfo("")
        feedback.pushInfo("[Step 8] Building output rows...")

        layer_crs = house_layer.crs()
        wgs84 = QgsCoordinateReferenceSystem('EPSG:4326')
        coord_transform = QgsCoordinateTransform(layer_crs, wgs84, context.project())
        feedback.pushInfo("  House layer CRS: {} -> transforming to EPSG:4326".format(layer_crs.authid()))

        output_rows = []
        row_counter = 0

        # Process ALL DJ Polygons in order (ensures ALL DJs appear, even with 0 houses)
        for poly in dj_polygons:
            dj_name = poly_to_dj.get(poly['fid'])
            if not dj_name:
                # Unmatched polygon -- skip (DJ point not found)
                continue

            dj_info = djs.get(dj_name)
            if not dj_info:
                continue

            pole_info = dj_to_pole.get(dj_name, {'pole_fid': '', 'pole_name': ''})
            fj_info = dj_to_fj.get(dj_name, {'fj_name': '', 'fj_pole': ''})
            ag_name = extract_ag(dj_info['dc_name'])
            block_name = poly.get('block', '')

            house_list = dj_to_houses.get(dj_name, [])

            if house_list:
                # Sort houses alphabetically
                house_list_sorted = sorted(house_list, key=lambda h: h['name'])
                for h in house_list_sorted:
                    row_counter += 1
                    # Transform house coordinates to EPSG:4326 (decimal degrees)
                    h_point = h.get('point', None)
                    if h_point:
                        try:
                            pt_4326 = coord_transform.transform(h_point)
                            h_lat = round(pt_4326.y(), 8)
                            h_lon = round(pt_4326.x(), 8)
                        except Exception:
                            h_lat = ''
                            h_lon = ''
                    else:
                        h_lat = ''
                        h_lon = ''
                    output_rows.append({
                        'fid': row_counter,
                        'Name': '',
                        'fid_2': h['fid'],
                        'Name_2': h['name'],
                        'fid_3': dj_info['fid'],
                        'name_3': dj_name,
                        'splitter': dj_info['splitter'],
                        'position': dj_info['position'],
                        'dc_name': dj_info['dc_name'],
                        'fid_2_2': pole_info['pole_fid'],
                        'name_2_2': pole_info['pole_name'],
                        'fj_name': fj_info['fj_name'],
                        'fj_pole': fj_info['fj_pole'],
                        'ag_name': ag_name,
                        'block': block_name,
                        'house_fields': h.get('fields', {}),
                        'house_lat': h_lat,
                        'house_lon': h_lon,
                    })
            else:
                # DJ has no houses -- still output one row with empty Name_2
                row_counter += 1
                output_rows.append({
                    'fid': row_counter,
                    'Name': '',
                    'fid_2': '',
                    'Name_2': '',
                    'fid_3': dj_info['fid'],
                    'name_3': dj_name,
                    'splitter': dj_info['splitter'],
                    'position': dj_info['position'],
                    'dc_name': dj_info['dc_name'],
                    'fid_2_2': pole_info['pole_fid'],
                    'name_2_2': pole_info['pole_name'],
                    'fj_name': fj_info['fj_name'],
                    'fj_pole': fj_info['fj_pole'],
                    'ag_name': ag_name,
                    'block': block_name,
                    'house_fields': {},
                    'house_lat': '',
                    'house_lon': '',
                })

        feedback.pushInfo("  Built {} output rows for {} DJs".format(
            len(output_rows), len([p for p in dj_polygons if p['fid'] in poly_to_dj])))

        # ================================================================
        # Step 9 -- Build Excel splicing diagram
        # ================================================================
        feedback.pushInfo("")
        feedback.pushInfo("[Step 9] Building Excel splicing diagram...")

        if not HAS_OPENPYXL:
            feedback.reportError("  ERROR: openpyxl not available. Cannot generate Excel.")
            feedback.reportError("  Install: python -m pip install openpyxl")
            return {self.PARAM_OUTPUT: output_path}

        try:
            self._build_splicing_excel(output_rows, output_path, feedback)
            feedback.pushInfo("  [OK] Excel written: {}".format(output_path))
        except Exception as e:
            feedback.reportError("  [ERROR] Building Excel: {}".format(e))
            import traceback
            feedback.reportError(traceback.format_exc())
            return {self.PARAM_OUTPUT: output_path}

        feedback.pushInfo("")
        feedback.pushInfo("=" * 60)
        feedback.pushInfo("SPLICING PLAN COMPLETE")
        feedback.pushInfo("  DJ Polygons processed : {}".format(len(dj_polygons)))
        feedback.pushInfo("  DJs matched           : {}".format(len(poly_to_dj)))
        feedback.pushInfo("  Houses matched (poly) : {}".format(total_matched))
        feedback.pushInfo("  Output rows           : {}".format(len(output_rows)))
        feedback.pushInfo("  DJs with houses       : {}".format(
            len([d for d in dj_to_houses if dj_to_houses[d]])))
        feedback.pushInfo("  DJs with 0 houses     : {}".format(
            len([d for d in dj_to_houses if not dj_to_houses[d]])))
        feedback.pushInfo("  Output file           : {}".format(output_path))

        return {self.PARAM_OUTPUT: output_path}


# =============================================================================
# FTTH Splicing Provider -- registers only FTTHSplicingPlanAlgorithm
# =============================================================================

class FTTHSplicingProvider(QgsProcessingProvider):
    """Processing provider that registers the FTTH Splicing Plan algorithm."""

    def loadAlgorithms(self, *args, **kwargs):
        self.addAlgorithm(FTTHSplicingPlanAlgorithm())

    def id(self, *args, **kwargs):
        return 'ftth_splicing'

    def name(self, *args, **kwargs):
        return 'FTTH Tools'

    def icon(self):
        return QgsProcessingProvider.icon(self)


# =============================================================================
# FTTH Splicing Plugin -- wrapper with menu items
# =============================================================================

class FTTHSplicingPlugin:
    """QGIS Plugin wrapper for FTTH Auto Splicing Plan.

    Provides a menu under Plugins > FTTH Splicing Plan with:
      - FTTH Splicing Plan  -> runs the processing algorithm
      - About               -> shows version info
    """

    def __init__(self, iface):
        self.iface = iface
        self.provider = FTTHSplicingProvider()
        self.menu = None
        self.actions = []

    def initGui(self):
        # Add Processing algorithms
        QgsApplication.processingRegistry().addProvider(self.provider)

        # Add menu under Plugins menu
        self.menu = QMenu('FTTH Splicing Plan', self.iface.mainWindow())

        # Action: Run Splicing Plan
        action_run = QAction('FTTH Splicing Plan', self.iface.mainWindow())
        action_run.triggered.connect(self.run_splicing_plan)
        self.menu.addAction(action_run)
        self.actions.append(action_run)

        self.menu.addSeparator()

        # Action: About
        action_about = QAction('About', self.iface.mainWindow())
        action_about.triggered.connect(self.show_about)
        self.menu.addAction(action_about)
        self.actions.append(action_about)

        # Add menu to Plugins menu
        self.iface.pluginMenu().addMenu(self.menu)

    def unload(self):
        # Remove Processing provider
        QgsApplication.processingRegistry().removeProvider(self.provider)

        # Remove menu actions
        for action in self.actions:
            action.deleteLater()
        self.actions = []

        if self.menu:
            self.menu.deleteLater()
            self.menu = None

    def run_splicing_plan(self):
        """Open the Processing algorithm dialog for FTTH Splicing Plan."""
        processing.execAlgorithmDialog('ftth_splicing:ftth_splicing_plan', {})

    def show_about(self):
        """Show the About dialog."""
        QMessageBox.information(
            self.iface.mainWindow(),
            'About FTTH Auto Splicing Plan',
            '<h3>FTTH Auto Splicing Plan v8.6.22</h3>'
            '<p>Generates FTTH splicing plan Excel (.xlsx) from labeled network data.</p>'
            '<p>Uses DJ Polygons as the PRIMARY method to match houses to DJs '
            '(point-in-polygon). Handles both GeoPackage (Integer) and '
            'Shapefile (Real) position fields.</p>'
            '<p><b>Author:</b> Mustafa M M Ellaham<br/>'
            '<b>Email:</b> Mustafaellaham@gmail.com</p>'
            '<p>Copyright (c) Mustafa M M Ellaham. All rights reserved.</p>'
        )
